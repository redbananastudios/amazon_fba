"""Tests for buy_plan integration in the output writers.

PRD §10 acceptance: buy_plan columns appear in real generated XLSX /
CSV / MD output for each verdict.

Mirrors the patterns in test_output_candidate_score.py.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from sourcing_engine.output.csv_writer import OUTPUT_COLUMNS, write_csv
from sourcing_engine.output.excel_writer import COLUMNS, write_excel
from sourcing_engine.output.markdown_report import write_report


def _row(
    *,
    asin: str = "B000BUY01TEST",
    decision: str = "SHORTLIST",
    opportunity_verdict: str = "BUY",
    buy_plan_status: str = "OK",
    order_qty: int | None = 13,
    capital: float | None = 52.0,
    p_units: int | None = 18,
    p_revenue: float | None = 303.30,
    p_profit: float | None = 150.30,
    payback: float | None = 21.7,
    target_buy: float | None = 9.50,
    target_stretch: float | None = 8.52,
    gap_gbp: float | None = None,
    gap_pct: float | None = None,
    **overrides,
) -> dict:
    base = {
        "asin": asin,
        "product_name": f"Product for {asin}",
        "supplier": "test-supplier",
        "supplier_sku": "SKU-A",
        "ean": "5012345678900",
        "match_type": "UNIT",
        "price_basis": "FBA",
        "decision": decision,
        "decision_reason": "ok",
        "buy_cost": 4.0,
        "market_price": 16.85,
        "profit_current": 9.0,
        "profit_conservative": 8.35,
        "fba_seller_count": 4,
        "gated": "N",
        "risk_flags": [],
        "opportunity_verdict": opportunity_verdict,
        "opportunity_score": 85,
        "opportunity_confidence": "HIGH",
        "opportunity_reasons": [],
        "opportunity_blockers": [],
        "next_action": "test action",
        "candidate_score": 85,
        "candidate_band": "STRONG",
        "data_confidence": "HIGH",
        "candidate_reasons": [],
        "data_confidence_reasons": [],
        # buy_plan
        "order_qty_recommended": order_qty,
        "capital_required": capital,
        "projected_30d_units": p_units,
        "projected_30d_revenue": p_revenue,
        "projected_30d_profit": p_profit,
        "payback_days": payback,
        "target_buy_cost_buy": target_buy,
        "target_buy_cost_stretch": target_stretch,
        "gap_to_buy_gbp": gap_gbp,
        "gap_to_buy_pct": gap_pct,
        "buy_plan_status": buy_plan_status,
    }
    base.update(overrides)
    return base


# ────────────────────────────────────────────────────────────────────────
# Excel
# ────────────────────────────────────────────────────────────────────────


class TestExcelBuyPlanColumns:
    def test_all_eleven_columns_in_columns_definition(self):
        names = {c[0] for c in COLUMNS}
        for col in (
            "order_qty_recommended", "capital_required",
            "projected_30d_units", "projected_30d_revenue",
            "projected_30d_profit", "payback_days",
            "target_buy_cost_buy", "target_buy_cost_stretch",
            "gap_to_buy_gbp", "gap_to_buy_pct", "buy_plan_status",
        ):
            assert col in names, f"{col} missing from excel_writer COLUMNS"

    def test_buy_columns_appear_after_velocity(self):
        names = [c[0] for c in COLUMNS]
        share_idx = names.index("predicted_velocity_share_source")
        order_idx = names.index("order_qty_recommended")
        # Buy plan block sits immediately after the velocity block.
        assert order_idx == share_idx + 1

    def test_buy_columns_render_for_buy_verdict(self, tmp_path: Path):
        df = pd.DataFrame([_row()])
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        names = [c[0] for c in COLUMNS]
        qty_col = names.index("order_qty_recommended") + 1
        cap_col = names.index("capital_required") + 1
        target_col = names.index("target_buy_cost_buy") + 1
        # Header row is row 2; first data row is row 3.
        assert ws.cell(row=3, column=qty_col).value == 13
        assert ws.cell(row=3, column=cap_col).value == 52.0
        assert ws.cell(row=3, column=target_col).value == 9.50

    def test_source_only_renders_target_costs_only(self, tmp_path: Path):
        df = pd.DataFrame([_row(
            asin="B000SRC01TEST",
            decision="REVIEW",
            opportunity_verdict="SOURCE_ONLY",
            buy_plan_status="NO_BUY_COST",
            order_qty=None, capital=None, payback=None,
            target_buy=4.85, target_stretch=4.10,
            p_units=42, p_revenue=710.00, p_profit=136.00,
        )])
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        names = [c[0] for c in COLUMNS]
        qty_col = names.index("order_qty_recommended") + 1
        target_col = names.index("target_buy_cost_buy") + 1
        status_col = names.index("buy_plan_status") + 1
        # Order qty blank, target populated, status NO_BUY_COST.
        assert ws.cell(row=3, column=qty_col).value is None
        assert ws.cell(row=3, column=target_col).value == 4.85
        assert ws.cell(row=3, column=status_col).value == "NO_BUY_COST"

    def test_negotiate_renders_gap(self, tmp_path: Path):
        df = pd.DataFrame([_row(
            asin="B000NEG01TEST",
            decision="REVIEW",
            opportunity_verdict="NEGOTIATE",
            buy_plan_status="OK",
            order_qty=None, capital=None, payback=None,
            target_buy=4.38, target_stretch=3.50,
            gap_gbp=0.62, gap_pct=0.124,
            buy_cost=5.00,
        )])
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        names = [c[0] for c in COLUMNS]
        gap_col = names.index("gap_to_buy_gbp") + 1
        gap_pct_col = names.index("gap_to_buy_pct") + 1
        assert ws.cell(row=3, column=gap_col).value == 0.62
        assert ws.cell(row=3, column=gap_pct_col).value == pytest.approx(0.124)


class TestExcelBuyPlanSort:
    def test_buy_rows_sorted_by_projected_30d_profit_desc(self, tmp_path: Path):
        # Per PRD §8.1 — within BUY tier, secondary sort is
        # projected_30d_profit desc (not candidate_score).
        rows = [
            _row(asin="B000LOW000", p_profit=20.0, candidate_score=95),
            _row(asin="B000HIGH00", p_profit=200.0, candidate_score=50),
            _row(asin="B000MID000", p_profit=80.0, candidate_score=80),
        ]
        df = pd.DataFrame(rows)
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        names = [c[0] for c in COLUMNS]
        asin_col = names.index("asin") + 1
        asins = [ws.cell(row=r, column=asin_col).value for r in (3, 4, 5)]
        assert asins == ["B000HIGH00", "B000MID000", "B000LOW000"]


# ────────────────────────────────────────────────────────────────────────
# CSV
# ────────────────────────────────────────────────────────────────────────


class TestCsvBuyPlanColumns:
    def test_eleven_columns_in_schema(self):
        for col in (
            "order_qty_recommended", "capital_required",
            "projected_30d_units", "projected_30d_revenue",
            "projected_30d_profit", "payback_days",
            "target_buy_cost_buy", "target_buy_cost_stretch",
            "gap_to_buy_gbp", "gap_to_buy_pct", "buy_plan_status",
        ):
            assert col in OUTPUT_COLUMNS, f"{col} missing from csv_writer schema"

    def test_buy_plan_columns_written_to_csv(self, tmp_path: Path):
        df = pd.DataFrame([_row()])
        out_path = tmp_path / "out.csv"
        write_csv(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        # Header row carries every column.
        for col in (
            "order_qty_recommended", "capital_required",
            "projected_30d_units", "buy_plan_status",
        ):
            assert col in body
        # Round-trip: values land in the right rows.
        df_back = pd.read_csv(out_path)
        assert df_back["order_qty_recommended"].iloc[0] == 13
        assert df_back["buy_plan_status"].iloc[0] == "OK"


class TestCsvExposesEngineFields:
    """The 2026-05-03 decision-data audit revealed the CSV writer was
    silently dropping ~30 populated columns. Operators saw blanks for
    fields the engine had computed and the buyer report was reading.
    Pin the fix: every signal the analyst depends on must be in the
    schema."""

    def test_decision_pipeline_outputs_in_schema(self):
        for col in (
            "candidate_score", "candidate_band", "data_confidence",
            "opportunity_verdict", "opportunity_score", "opportunity_confidence",
            "next_action", "opportunity_reasons", "opportunity_blockers",
        ):
            assert col in OUTPUT_COLUMNS, f"{col} missing from csv schema"

    def test_trend_signals_in_schema(self):
        for col in (
            "bsr_slope_30d", "bsr_slope_90d", "bsr_slope_365d",
            "bsr_drops_30d",
            "buy_box_oos_pct_90", "price_volatility_90d", "buy_box_drop_pct_90",
            "buy_box_min_365d", "buy_box_avg30", "buy_box_avg90",
            "fba_offer_count_90d_start", "fba_offer_count_90d_joiners",
            "listing_age_days", "amazon_bb_pct_90",
        ):
            assert col in OUTPUT_COLUMNS, f"{col} missing from csv schema"

    def test_economics_breakdown_in_schema(self):
        for col in (
            "roi_current", "roi_conservative", "breakeven_price",
            "amazon_price", "new_fba_price", "buy_box_price",
            "fba_seller_count", "amazon_status", "price_history_basis",
            "fba_pick_pack_fee", "referral_fee_pct",
        ):
            assert col in OUTPUT_COLUMNS, f"{col} missing from csv schema"

    def test_predicted_velocity_in_schema(self):
        for col in (
            "predicted_velocity_low", "predicted_velocity_mid",
            "predicted_velocity_high", "predicted_velocity_share_source",
            "buy_box_seller_stats",
        ):
            assert col in OUTPUT_COLUMNS, f"{col} missing from csv schema"

    def test_dict_field_serialised_to_json(self, tmp_path: Path):
        # buy_box_seller_stats is a dict — must round-trip through CSV
        # without breaking pandas.
        df = pd.DataFrame([_row(buy_box_seller_stats={
            "ABC": {"pct_won": 45.0, "is_fba": True},
            "DEF": {"pct_won": 25.0, "is_fba": True},
        })])
        out_path = tmp_path / "out.csv"
        write_csv(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        # JSON serialisation preserves the structure.
        assert '"ABC"' in body
        assert "45.0" in body

    def test_list_fields_join_with_semicolons(self, tmp_path: Path):
        df = pd.DataFrame([_row(
            opportunity_reasons=["healthy ROI", "few sellers", "stable price"],
            opportunity_blockers=["candidate_score < 75"],
        )])
        out_path = tmp_path / "out.csv"
        write_csv(df, str(out_path))
        df_back = pd.read_csv(out_path)
        assert "healthy ROI; few sellers; stable price" in df_back["opportunity_reasons"].iloc[0]
        assert df_back["opportunity_blockers"].iloc[0] == "candidate_score < 75"


# ────────────────────────────────────────────────────────────────────────
# Markdown
# ────────────────────────────────────────────────────────────────────────


class TestMarkdownBuyPlan:
    def test_buy_row_renders_order_plan_line(self, tmp_path: Path):
        df = pd.DataFrame([_row()])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        # Per PRD §8.3 BUY shape: "Order plan: ... units · £... capital · ..."
        assert "Order plan" in body
        assert "13 units" in body
        assert "£52.00 capital" in body

    def test_source_only_row_renders_source_target_line(self, tmp_path: Path):
        df = pd.DataFrame([_row(
            decision="REVIEW",
            opportunity_verdict="SOURCE_ONLY",
            buy_plan_status="NO_BUY_COST",
            order_qty=None, capital=None, payback=None,
            target_buy=4.85, target_stretch=4.10,
            p_units=42, p_revenue=710.00, p_profit=136.00,
        )])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        assert "Source target" in body
        assert "4.85" in body

    def test_negotiate_row_renders_negotiation_ask_line(self, tmp_path: Path):
        df = pd.DataFrame([_row(
            decision="REVIEW",
            opportunity_verdict="NEGOTIATE",
            buy_plan_status="OK",
            order_qty=None, capital=None, payback=None,
            target_buy=4.38, target_stretch=3.50,
            gap_gbp=0.62, gap_pct=0.124,
            buy_cost=5.00,
        )])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        assert "Negotiation ask" in body
        assert "0.62" in body

    def test_watch_row_does_not_render_buy_plan_line(self, tmp_path: Path):
        # PRD §8.3 — WATCH/KILL rows: no buy-plan line in markdown.
        df = pd.DataFrame([_row(
            decision="REVIEW",
            opportunity_verdict="WATCH",
            buy_plan_status="BLOCKED_BY_VERDICT",
            order_qty=None, capital=None, payback=None,
            gap_gbp=None, gap_pct=None,
        )])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        # No order-plan / source-target / negotiation-ask line for WATCH.
        assert "Order plan" not in body
        assert "Source target" not in body
        assert "Negotiation ask" not in body
