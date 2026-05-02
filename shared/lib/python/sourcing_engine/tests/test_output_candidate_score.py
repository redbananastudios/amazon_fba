"""Tests for candidate-score integration in the output writers.

HANDOFF WS3.6 — verify:
  - excel_writer adds the 5 candidate columns
  - excel_writer sorts by candidate_score desc within each decision band
  - excel_writer colour-codes candidate_band cells per the WS3.6 rules
  - markdown_report includes candidate columns + per-row leading lines
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

from sourcing_engine.output.excel_writer import COLUMNS, write_excel
from sourcing_engine.output.markdown_report import (
    _candidate_score_summary,
    write_report,
)


def _row(
    *,
    decision: str = "SHORTLIST",
    candidate_band: str = "STRONG",
    candidate_score: int = 80,
    data_confidence: str = "HIGH",
    asin: str = "B0AAA",
    product_name: str = "A test product",
    supplier: str = "test-supplier",
    profit_current: float = 5.0,
    **overrides,
) -> dict:
    base = {
        "asin": asin,
        "product_name": product_name,
        "supplier": supplier,
        "decision": decision,
        "candidate_band": candidate_band,
        "candidate_score": candidate_score,
        "data_confidence": data_confidence,
        "candidate_reasons": ["sales=200/mo→10", "BSR improving→10"],
        "data_confidence_reasons": [],
        "buy_cost": 5.0,
        "market_price": 15.0,
        "profit_current": profit_current,
        "profit_conservative": 4.0,
        "fba_seller_count": 3,
        "ean": "5012345678900",
        "match_type": "UNIT",
        "supplier_sku": "SKU-A",
        "gated": "N",
        "risk_flags": [],
        "decision_reason": "ok",
        "price_basis": "FBA",
    }
    base.update(overrides)
    return base


# ────────────────────────────────────────────────────────────────────────
# excel_writer
# ────────────────────────────────────────────────────────────────────────


class TestExcelColumns:
    def test_candidate_columns_present_in_columns_definition(self):
        names = {c[0] for c in COLUMNS}
        for col in (
            "candidate_band", "candidate_score", "data_confidence",
            "candidate_reasons", "data_confidence_reasons",
        ):
            assert col in names, f"{col} missing from excel_writer COLUMNS"

    def test_candidate_columns_appear_after_decision(self):
        names = [c[0] for c in COLUMNS]
        decision_idx = names.index("decision")
        band_idx = names.index("candidate_band")
        # band immediately after decision per the WS3.6 layout.
        assert band_idx == decision_idx + 1


class TestExcelSort:
    def test_shortlist_rows_sorted_by_candidate_score_desc(self, tmp_path: Path):
        rows = [
            _row(asin="B0LOW", candidate_score=40, candidate_band="OK"),
            _row(asin="B0HIGH", candidate_score=90, candidate_band="STRONG"),
            _row(asin="B0MED", candidate_score=60, candidate_band="OK"),
        ]
        df = pd.DataFrame(rows)
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))

        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        names = [c[0] for c in COLUMNS]
        asin_col = names.index("asin") + 1
        # First data row is row 3 (rows 1=title, 2=header).
        asins = [ws.cell(row=r, column=asin_col).value for r in (3, 4, 5)]
        assert asins == ["B0HIGH", "B0MED", "B0LOW"]

    def test_shortlist_sorted_above_review(self, tmp_path: Path):
        rows = [
            _row(asin="B0REV", decision="REVIEW", candidate_score=99),
            _row(asin="B0SHORT", decision="SHORTLIST", candidate_score=10),
        ]
        df = pd.DataFrame(rows)
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))

        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        names = [c[0] for c in COLUMNS]
        asin_col = names.index("asin") + 1
        # SHORTLIST first regardless of candidate_score.
        first = ws.cell(row=3, column=asin_col).value
        second = ws.cell(row=4, column=asin_col).value
        assert first == "B0SHORT"
        assert second == "B0REV"


class TestExcelColouring:
    def _band_cell_fill(self, ws, row: int) -> str:
        """Return the candidate_band cell's fill colour as hex."""
        names = [c[0] for c in COLUMNS]
        band_col = names.index("candidate_band") + 1
        cell = ws.cell(row=row, column=band_col)
        # openpyxl returns the rgb as "00<RRGGBB>"; strip the alpha prefix.
        rgb = cell.fill.start_color.rgb if cell.fill.start_color else None
        return rgb[-6:].upper() if rgb else ""

    def test_strong_high_gets_green(self, tmp_path: Path):
        df = pd.DataFrame([_row(candidate_band="STRONG", data_confidence="HIGH")])
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        # ABEBC6 = green per the writer.
        assert self._band_cell_fill(ws, 3) == "ABEBC6"

    def test_strong_low_gets_amber(self, tmp_path: Path):
        df = pd.DataFrame([_row(candidate_band="STRONG", data_confidence="LOW")])
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        # FAD7A0 = amber per the writer.
        assert self._band_cell_fill(ws, 3) == "FAD7A0"

    def test_fail_band_gets_grey(self, tmp_path: Path):
        # FAIL rows reach the XLSX only when they're SHORTLIST or
        # REVIEW (excel_writer drops REJECT). Decisions and bands
        # are independent, so a REVIEW + FAIL row is real.
        df = pd.DataFrame([_row(
            decision="REVIEW", candidate_band="FAIL", data_confidence="LOW",
        )])
        out_path = tmp_path / "out.xlsx"
        write_excel(df, str(out_path))
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        # D5D8DC = grey per the writer.
        assert self._band_cell_fill(ws, 3) == "D5D8DC"


# ────────────────────────────────────────────────────────────────────────
# markdown_report
# ────────────────────────────────────────────────────────────────────────


class TestMarkdownReport:
    def test_candidate_columns_in_table(self, tmp_path: Path):
        df = pd.DataFrame([_row()])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        assert "candidate_band" in body
        assert "candidate_score" in body
        assert "data_confidence" in body

    def test_per_row_leading_line_summary(self):
        df = pd.DataFrame([
            _row(
                candidate_band="STRONG", candidate_score=82,
                data_confidence="HIGH", asin="B0AAA",
                product_name="A test product",
            ),
        ])
        out = _candidate_score_summary(df)
        # Format from handoff: **STRONG** (HIGH confidence) — score 82/100
        assert "**STRONG**" in out
        assert "(HIGH confidence)" in out
        assert "82/100" in out
        assert "B0AAA" in out

    def test_summary_empty_string_when_no_band_column(self):
        df = pd.DataFrame([{
            "asin": "B0X", "product_name": "X", "decision": "SHORTLIST",
        }])
        assert _candidate_score_summary(df) == ""

    def test_summary_handles_nan_band_cell(self):
        """NaN-truthy regression: pandas fills missing keys with NaN,
        which is truthy for floats. The summary must coerce NaN to ""
        rather than emitting `**nan** (nan confidence)`."""
        # Two rows: one with the column populated (gets a row), one
        # with NaN in the band cell (skipped silently).
        df = pd.DataFrame([
            {"asin": "B0NAN", "product_name": "X",
             "candidate_band": float("nan"),
             "candidate_score": float("nan"),
             "data_confidence": float("nan")},
            {"asin": "B0OK", "product_name": "Real product",
             "candidate_band": "STRONG", "candidate_score": 80,
             "data_confidence": "HIGH"},
        ])
        out = _candidate_score_summary(df)
        # Real row appears.
        assert "B0OK" in out
        assert "**STRONG**" in out
        # NaN row does NOT leak through as the literal string "nan".
        assert "**nan**" not in out
        assert "B0NAN" not in out

    def test_summary_empty_string_for_empty_df(self):
        df = pd.DataFrame()
        assert _candidate_score_summary(df) == ""

    def test_review_section_includes_summary(self, tmp_path: Path):
        df = pd.DataFrame([
            _row(decision="REVIEW", candidate_band="OK", candidate_score=55),
        ])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        # Manual Review section + leading line + table
        assert "Manual Review" in body
        assert "**OK**" in body
        assert "55/100" in body

    def test_falls_through_when_candidate_score_absent(self, tmp_path: Path):
        # Older runs without candidate-score columns must still produce
        # a valid report (just without the score-aware sections).
        df = pd.DataFrame([{
            "supplier": "test", "asin": "B0X",
            "product_name": "Old product", "decision": "SHORTLIST",
            "ean": "5012345678900", "match_type": "UNIT",
            "buy_cost": 5.0, "profit_conservative": 4.0,
            "margin_conservative": 0.30, "sales_estimate": 100,
            "decision_reason": "ok",
        }])
        out_path = tmp_path / "report.md"
        write_report(df, str(out_path))
        body = out_path.read_text(encoding="utf-8")
        # Report still produced; no score header but the structure is intact.
        assert "Supplier Shortlist Report" in body
        assert "B0X" in body
