"""Final styled Excel report for Abgee sourcing results.

Produces a single, clean XLSX with:
- Colour-coded rows (green REVIEW, red REJECT)
- Amazon URL column
- Formatted currency and percentage columns
- Frozen header row and auto-filters
- Sorted by profit descending
"""
import logging

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Columns for the final report (ordered for readability)
REPORT_COLUMNS = [
    "decision",
    "supplier",
    "product_name",
    "asin",
    "amazon_url",
    "ean",
    "match_type",
    "buy_cost",
    "market_price",
    "profit_conservative",
    "margin_conservative",
    "fees_conservative",
    "sales_estimate",
    "profit_current",
    "margin_current",
    "fees_current",
    "supplier_price_basis",
    "case_qty",
    "risk_flags",
    "decision_reason",
]

# Human-friendly column headers
COLUMN_LABELS = {
    "decision": "Decision",
    "supplier": "Supplier",
    "product_name": "Product",
    "asin": "ASIN",
    "amazon_url": "Amazon Link",
    "ean": "EAN",
    "match_type": "Match",
    "buy_cost": "Buy Cost",
    "market_price": "Sell Price",
    "profit_conservative": "Profit (Cons.)",
    "margin_conservative": "Margin (Cons.)",
    "fees_conservative": "Fees (Cons.)",
    "sales_estimate": "Sales/mo",
    "profit_current": "Profit (Current)",
    "margin_current": "Margin (Current)",
    "fees_current": "Fees (Current)",
    "supplier_price_basis": "Price Basis",
    "case_qty": "Case Qty",
    "risk_flags": "Risk Flags",
    "decision_reason": "Reason",
}

# Column widths
COLUMN_WIDTHS = {
    "decision": 10,
    "supplier": 18,
    "product_name": 45,
    "asin": 12,
    "amazon_url": 14,
    "ean": 15,
    "match_type": 7,
    "buy_cost": 10,
    "market_price": 10,
    "profit_conservative": 12,
    "margin_conservative": 12,
    "fees_conservative": 11,
    "sales_estimate": 10,
    "profit_current": 12,
    "margin_current": 12,
    "fees_current": 11,
    "supplier_price_basis": 10,
    "case_qty": 8,
    "risk_flags": 30,
    "decision_reason": 35,
}


def write_final_report(df: pd.DataFrame, path: str, include_rejects: bool = False):
    """Write the final styled Excel report.

    Args:
        df: Full pipeline output DataFrame.
        path: Output .xlsx path.
        include_rejects: If True, include REJECT rows in a second sheet.
    """
    try:
        # Add Amazon URL column
        df = df.copy()
        df["amazon_url"] = df["asin"].apply(
            lambda x: f"https://www.amazon.co.uk/dp/{x}" if pd.notna(x) and str(x).startswith("B") else ""
        )

        # Format risk_flags
        if "risk_flags" in df.columns:
            df["risk_flags"] = df["risk_flags"].apply(
                lambda x: ", ".join(x) if isinstance(x, list) else str(x) if pd.notna(x) else ""
            )

        # Select and order columns
        cols = [c for c in REPORT_COLUMNS if c in df.columns]

        # Split by decision
        review = df[df["decision"] == "REVIEW"].sort_values("profit_conservative", ascending=False)
        profitable_reject = df[
            (df["decision"] == "REJECT") & (df.get("profit_conservative", pd.Series(dtype=float)) > 0)
        ].sort_values("profit_conservative", ascending=False)
        other_reject = df[
            (df["decision"] == "REJECT") & (~(df.get("profit_conservative", pd.Series(dtype=float)) > 0))
        ]

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # Sheet 1: Candidates (REVIEW + profitable REJECT)
            candidates = pd.concat([review, profitable_reject], ignore_index=True)
            if not candidates.empty:
                _write_sheet(writer, candidates[cols], "Candidates", cols)

            # Sheet 2: All rejects (if requested)
            if include_rejects and not other_reject.empty:
                _write_sheet(writer, other_reject[cols].head(500), "Rejected", cols)

        logger.info("Final report written: %s (%d candidates)", path, len(candidates))
        return len(candidates)

    except Exception:
        logger.exception("Failed to write final report: %s", path)
        return 0


def _write_sheet(writer, df, sheet_name, cols):
    """Write a single styled sheet."""
    # Write data with renamed headers
    labels = [COLUMN_LABELS.get(c, c) for c in cols]
    out = df.copy()
    out.columns = labels
    out.to_excel(writer, index=False, sheet_name=sheet_name)

    ws = writer.sheets[sheet_name]

    # --- Styles ---
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    amber_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
    )
    body_font = Font(size=10)
    currency_fmt = '£#,##0.00'
    pct_fmt = '0.0%'

    # Header styling
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    for col_idx, col_name in enumerate(cols, 1):
        width = COLUMN_WIDTHS.get(col_name, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Find column indices for formatting
    col_indices = {c: i + 1 for i, c in enumerate(cols)}
    currency_cols = {"buy_cost", "market_price", "profit_conservative", "profit_current",
                     "fees_conservative", "fees_current"}
    pct_cols = {"margin_conservative", "margin_current"}
    decision_col = col_indices.get("decision")

    # Row styling
    for row_idx in range(2, len(df) + 2):
        # Decision-based row colour
        if decision_col:
            decision_val = ws.cell(row=row_idx, column=decision_col).value
            if decision_val == "REVIEW":
                row_fill = green_fill
            elif decision_val == "REJECT":
                row_fill = amber_fill
            else:
                row_fill = None
        else:
            row_fill = None

        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.border = thin_border

            if row_fill:
                cell.fill = row_fill

            # Number formatting
            if col_name in currency_cols:
                cell.number_format = currency_fmt
                cell.alignment = Alignment(horizontal="right")
            elif col_name in pct_cols:
                cell.number_format = pct_fmt
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "sales_estimate":
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "product_name":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_name == "amazon_url":
                # Make it a clickable hyperlink
                url = cell.value
                if url and str(url).startswith("http"):
                    cell.hyperlink = url
                    cell.value = "View"
                    cell.font = Font(color="0563C1", underline="single", size=10)
                    cell.alignment = Alignment(horizontal="center")

    # Freeze top row + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Row height
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 20
