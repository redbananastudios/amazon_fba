"""Excel output — styled workbook with SHORTLIST + REVIEW only."""
import logging
import pandas as pd

logger = logging.getLogger(__name__)

# Column definitions: (internal_name, display_header, width, format)
# format: "gbp" = £0.00, "pct" = 0.0%, "int" = whole number, "text" = string, "url" = hyperlink
COLUMNS = [
    ("product_name",            "Product Name",         45, "text"),
    ("amazon_url",              "Amazon URL",           38, "url"),
    ("asin",                    "ASIN",                 14, "text"),
    ("supplier",                "Supplier",             18, "text"),
    ("supplier_sku",            "Supplier Part No.",    16, "text"),
    ("decision",                "Decision",             14, "text"),
    ("gated",                   "Gated",                10, "text"),
    ("buy_cost",                "Cost inc VAT",         14, "gbp"),
    ("market_price",            "Buy Box",              12, "gbp"),
    ("profit_current",          "Profit (Current)",     16, "gbp"),
    ("profit_conservative",     "Profit (Conservative)",18, "gbp"),
    ("margin_current",          "Margin (Current)",     16, "pct"),
    ("margin_conservative",     "Margin (Conservative)",18, "pct"),
    ("sales_estimate",          "Est. Sales/Month",     16, "int"),
    ("bought_past_month",       "Bought/Month (Keepa)", 18, "int"),
    ("sales_rank",              "BSR",                  14, "int"),
    ("fba_seller_count",        "FBA Sellers",          12, "int"),
    ("amazon_on_listing",       "Amazon On Listing",    16, "text"),
    ("amazon_bb_pct_90",        "Amazon Share % (90d)", 18, "text"),
    ("fees_current",            "Total Fees (Current)", 16, "gbp"),
    ("fees_conservative",       "Total Fees (Conserv.)",18, "gbp"),
    ("raw_conservative_price",  "Conservative Price",   16, "gbp"),
    ("max_buy_price",           "Max Buy Price (20% ROI)", 20, "gbp"),
    ("buy_box_avg90",           "Buy Box 90d Avg",      14, "gbp"),
    ("buy_box_drop_pct_90",     "Price Drop % (90d)",   16, "text"),
    ("rating",                  "Rating",               10, "text"),
    ("review_count",            "Reviews",              12, "int"),
    ("price_basis",             "Fulfilment",           12, "text"),
    ("match_type",              "Match Type",           12, "text"),
    ("ean",                     "EAN",                  16, "text"),
    ("size_tier",               "Size Tier",            14, "text"),
    ("capital_exposure",        "Capital Exposure",     16, "gbp"),
    ("decision_reason",         "Decision Reason",      50, "text"),
    ("risk_flags",              "Risk Flags",           40, "text"),
    # SP-API preflight columns (informational only — appended at end per spec).
    # Populated when MCP CLI is built and SP_API creds are set; blank otherwise.
    ("restriction_status",      "Restriction Status",   18, "text"),
    ("restriction_reasons",     "Restriction Reasons",  30, "text"),
    ("restriction_links",       "Ungate Links",         60, "url"),
    ("fba_eligible",            "FBA Eligible",         12, "text"),
    ("fba_ineligibility",       "FBA Ineligibility",    24, "text"),
    ("live_buy_box",            "Live Buy Box",         14, "gbp"),
    ("live_buy_box_seller",     "Live BB Seller",       14, "text"),
    ("live_offer_count_new",    "Offers (new)",         12, "int"),
    ("live_offer_count_fba",    "Offers (FBA)",         12, "int"),
    ("catalog_brand",           "Catalog Brand",        18, "text"),
    ("keepa_brand",             "Keepa Brand",          18, "text"),
    ("catalog_hazmat",          "Hazmat",               10, "text"),
    ("preflight_errors",        "Preflight Errors",     40, "text"),
]


def write_excel(
    df: pd.DataFrame,
    path: str,
    market_data: dict | None = None,
    supplier_label: str | None = None,
):
    """Write the styled Excel workbook.

    Args:
        df: pipeline output DataFrame (will exclude REJECT rows automatically)
        path: output xlsx path
        market_data: optional Keepa market data dict for enrichment
        supplier_label: friendly name to show in the title bar.
            If None, derived from the most-frequent value of df['supplier'].
            Falls back to 'Supplier' if neither is available.
    """
    try:
        out = df.copy()

        # Exclude REJECT rows
        if "decision" in out.columns:
            out = out[out["decision"] != "REJECT"].reset_index(drop=True)

        if out.empty:
            logger.warning("No SHORTLIST/REVIEW rows — skipping Excel")
            return

        # Flatten risk_flags list to string
        if "risk_flags" in out.columns:
            out["risk_flags"] = out["risk_flags"].apply(
                lambda x: "; ".join(x) if isinstance(x, list) else str(x))

        # Derive amazon_on_listing from risk_flags
        if "amazon_on_listing" not in out.columns:
            out["amazon_on_listing"] = out.get("risk_flags", "").apply(
                lambda x: "Y" if "AMAZON_ON_LISTING" in str(x) else "N")

        # Enrich from market_data if available
        if market_data:
            _enrich_from_keepa(out, market_data)

        # Build output in column order
        col_names = [c[0] for c in COLUMNS]
        for col in col_names:
            if col not in out.columns:
                out[col] = None

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Shortlist"

        # Colours
        header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        amber_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        border = Border(
            bottom=Side(style="hair", color="D5D8DC"),
            left=Side(style="hair", color="D5D8DC"),
            right=Side(style="hair", color="D5D8DC"),
        )
        green_font = Font(color="27AE60", bold=True, size=10)
        red_font = Font(color="C0392B", bold=True, size=10)
        link_font = Font(color="2980B9", underline="single", size=9)
        money_fmt = '£#,##0.00'
        pct_fmt = '0.0%'
        int_fmt = '#,##0'

        # --- Title row ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
        title_cell = ws.cell(row=1, column=1)
        shortlist_count = (out["decision"] == "SHORTLIST").sum()
        review_count = (out["decision"] == "REVIEW").sum()
        # Title — uses caller-provided label, else most-common supplier in data, else generic.
        if supplier_label is None:
            try:
                supplier_label = str(out["supplier"].mode().iat[0]) if "supplier" in out.columns else "Supplier"
            except Exception:
                supplier_label = "Supplier"
        title_cell.value = (
            f"{supplier_label} Analysis — {shortlist_count} Shortlisted, "
            f"{review_count} Review  |  {pd.Timestamp.now().strftime('%Y-%m-%d')}"
        )
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        # --- Header row ---
        for col_idx, (internal, display, width, fmt) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=display)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color="D5D8DC"))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = 32

        # --- Data rows ---
        decision_col_idx = next(i for i, c in enumerate(COLUMNS) if c[0] == "decision") + 1
        profit_col_idx = next(i for i, c in enumerate(COLUMNS) if c[0] == "profit_current") + 1

        for row_idx, (_, row) in enumerate(out.iterrows(), 3):
            decision_val = row.get("decision", "")
            is_even = (row_idx - 3) % 2 == 0

            for col_idx, (internal, display, width, fmt) in enumerate(COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = row.get(internal)

                # Handle NaN
                if pd.isna(val):
                    cell.value = None
                elif fmt == "url" and val:
                    cell.value = str(val)
                    # Multi-URL cells (e.g. restriction_links =
                    # "https://...; https://...") use the FIRST URL as
                    # the hyperlink target — Excel cells only support
                    # one hyperlink per cell. The full text remains
                    # visible / copyable; the operator clicks for the
                    # primary application URL.
                    first_url = str(val).split(";", 1)[0].strip()
                    if first_url:
                        cell.hyperlink = first_url
                        cell.font = link_font
                elif fmt == "gbp":
                    try:
                        cell.value = round(float(val), 2)
                        cell.number_format = money_fmt
                    except (ValueError, TypeError):
                        cell.value = val
                elif fmt == "pct":
                    try:
                        cell.value = float(val)
                        cell.number_format = pct_fmt
                    except (ValueError, TypeError):
                        cell.value = val
                elif fmt == "int":
                    try:
                        cell.value = int(float(val))
                        cell.number_format = int_fmt
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = str(val) if val is not None else None

                # Default font (unless already set for URLs)
                if fmt != "url":
                    cell.font = Font(size=10)

                # Alignment
                if internal in ("product_name", "decision_reason", "risk_flags"):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Border
                cell.border = border

            # Row fill — decision-based
            row_fill = green_fill if decision_val == "SHORTLIST" else amber_fill
            if is_even and decision_val == "SHORTLIST":
                # Slightly different green for zebra
                row_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

            # Profit colour — green if positive, red if negative
            profit_cell = ws.cell(row=row_idx, column=profit_col_idx)
            try:
                pv = float(row.get("profit_current", 0))
                profit_cell.font = green_font if pv >= 0 else red_font
            except (ValueError, TypeError):
                pass

            # Gated cell highlighting
            gated_col_idx = next(i for i, c in enumerate(COLUMNS) if c[0] == "gated") + 1
            gated_cell = ws.cell(row=row_idx, column=gated_col_idx)
            gated_val = str(row.get("gated", "")).upper()
            if gated_val == "Y":
                gated_cell.fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
                gated_cell.font = Font(bold=True, size=10, color="8E44AD")
            elif gated_val == "UNKNOWN":
                gated_cell.fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
                gated_cell.font = Font(bold=True, size=10, color="E67E22")

            # Amazon on listing highlighting
            amz_col_idx = next(i for i, c in enumerate(COLUMNS) if c[0] == "amazon_on_listing") + 1
            amz_cell = ws.cell(row=row_idx, column=amz_col_idx)
            if str(row.get("amazon_on_listing", "")).upper() == "Y":
                amz_cell.fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
                amz_cell.font = Font(bold=True, size=10, color="E67E22")

            ws.row_dimensions[row_idx].height = 22

        # Freeze panes — freeze title + header rows, and first 3 columns
        ws.freeze_panes = "D3"

        # Auto-filter on header row
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(out) + 2}"

        wb.save(path)
        logger.info("Excel written: %s (%d rows)", path, len(out))

    except Exception:
        logger.exception("Failed to write Excel: %s", path)


def _enrich_from_keepa(df, market_data):
    """Add extra Keepa fields to the output dataframe."""
    for idx, row in df.iterrows():
        ean = str(row.get("ean", ""))
        md = market_data.get(ean, {})
        if not md:
            continue
        df.at[idx, "amazon_bb_pct_90"] = md.get("amazon_bb_pct_90", "")
        df.at[idx, "rating"] = md.get("rating", "")
        df.at[idx, "review_count"] = md.get("review_count")
        df.at[idx, "sales_rank"] = md.get("sales_rank")
        df.at[idx, "bought_past_month"] = md.get("monthly_sales_estimate")
        df.at[idx, "buy_box_avg90"] = md.get("buy_box_avg90")
        df.at[idx, "buy_box_drop_pct_90"] = md.get("buy_box_drop_pct_90", "")
        df.at[idx, "fba_seller_count"] = md.get("fba_seller_count")
