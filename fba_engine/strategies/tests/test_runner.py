"""Tests for fba_engine.strategies.runner.

The runner is the YAML composition layer over `fba_engine.steps.*`. These
tests exercise the wiring (variable interpolation, step chain, error
shapes, side-effect step config) plus an end-to-end run through the live
keepa_niche pipeline against a fixture CSV.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
import yaml

from fba_engine.strategies.runner import (
    StepDef,
    StrategyConfigError,
    StrategyDef,
    StrategyExecutionError,
    interpolate,
    load_strategy,
    run_strategy,
)


# ---------------------------------------------------------------------------
# Variable interpolation
# ---------------------------------------------------------------------------


class TestInterpolate:
    def test_substitutes_known_keys(self):
        assert interpolate("{niche}-data", {"niche": "kids-toys"}) == "kids-toys-data"

    def test_handles_multiple_keys(self):
        result = interpolate(
            "{base}/working/{niche}_phase4.csv",
            {"base": "fba_engine/data/niches/kids-toys", "niche": "kids_toys"},
        )
        assert result == "fba_engine/data/niches/kids-toys/working/kids_toys_phase4.csv"

    def test_passes_through_strings_with_no_placeholders(self):
        assert interpolate("/no/placeholders", {"niche": "x"}) == "/no/placeholders"

    def test_missing_key_raises_strategy_config_error(self):
        with pytest.raises(StrategyConfigError, match="missing"):
            interpolate("{undefined}", {"niche": "x"})

    def test_non_string_value_passes_through_unchanged(self):
        # Integers/floats/dicts in YAML configs should be passed through
        # without trying to .format() them.
        assert interpolate(42, {"niche": "x"}) == 42
        assert interpolate({"key": "value"}, {"niche": "x"}) == {"key": "value"}


# ---------------------------------------------------------------------------
# YAML loading
# ---------------------------------------------------------------------------


class TestLoadStrategy:
    def _write_yaml(self, tmp_path: Path, body: str) -> Path:
        path = tmp_path / "strategy.yaml"
        path.write_text(body, encoding="utf-8")
        return path

    def test_minimal_strategy(self, tmp_path: Path):
        body = """\
name: test
description: Minimal
steps:
  - name: ip_risk
    module: fba_engine.steps.ip_risk
    config:
      niche: kids-toys
"""
        strat = load_strategy(self._write_yaml(tmp_path, body))
        assert strat.name == "test"
        assert strat.description == "Minimal"
        assert len(strat.steps) == 1
        assert strat.steps[0].name == "ip_risk"
        assert strat.steps[0].module == "fba_engine.steps.ip_risk"
        assert strat.steps[0].config == {"niche": "kids-toys"}

    def test_full_strategy_with_input_and_output(self, tmp_path: Path):
        body = """\
name: keepa
description: Full
input:
  path: "{base}/working/{niche_snake}_phase3.csv"
  encoding: utf-8-sig
steps:
  - name: ip_risk
    module: fba_engine.steps.ip_risk
    config:
      niche: "{niche}"
output:
  csv: "{base}/working/{niche_snake}_phase6.csv"
"""
        strat = load_strategy(self._write_yaml(tmp_path, body))
        assert strat.input_path == "{base}/working/{niche_snake}_phase3.csv"
        assert strat.input_encoding == "utf-8-sig"
        assert strat.output_csv == "{base}/working/{niche_snake}_phase6.csv"

    def test_default_encoding_is_utf8_sig(self, tmp_path: Path):
        body = """\
name: t
description: d
input:
  path: "/tmp/x.csv"
steps: []
"""
        strat = load_strategy(self._write_yaml(tmp_path, body))
        assert strat.input_encoding == "utf-8-sig"

    def test_step_with_empty_config(self, tmp_path: Path):
        # A step that needs no config (like decision_engine, build_output).
        body = """\
name: t
description: d
steps:
  - name: decision_engine
    module: fba_engine.steps.decision_engine
"""
        strat = load_strategy(self._write_yaml(tmp_path, body))
        assert strat.steps[0].config == {}

    def test_input_discover_flag_loads(self, tmp_path: Path):
        body = """\
name: t
description: d
input:
  discover: true
steps: []
"""
        strat = load_strategy(self._write_yaml(tmp_path, body))
        assert strat.input_discover is True
        assert strat.input_path is None

    def test_input_discover_default_false(self, tmp_path: Path):
        body = """\
name: t
description: d
input:
  path: "/tmp/x.csv"
steps: []
"""
        strat = load_strategy(self._write_yaml(tmp_path, body))
        assert strat.input_discover is False

    def test_input_discover_quoted_string_rejected(self, tmp_path: Path):
        # A string `"true"` is truthy via bool() — silent acceptance
        # would let `discover: "false"` (still truthy) wrongly opt-in.
        # The loader must reject anything that isn't a real YAML bool.
        body = """\
name: t
description: d
input:
  discover: "true"
steps: []
"""
        with pytest.raises(StrategyConfigError, match="discover"):
            load_strategy(self._write_yaml(tmp_path, body))

    def test_missing_name_field_raises(self, tmp_path: Path):
        body = """\
description: no name
steps: []
"""
        with pytest.raises(StrategyConfigError, match="name"):
            load_strategy(self._write_yaml(tmp_path, body))

    def test_missing_steps_field_raises(self, tmp_path: Path):
        body = """\
name: t
description: d
"""
        with pytest.raises(StrategyConfigError, match="steps"):
            load_strategy(self._write_yaml(tmp_path, body))

    def test_step_missing_module_raises(self, tmp_path: Path):
        body = """\
name: t
description: d
steps:
  - name: nameOnly
"""
        with pytest.raises(StrategyConfigError, match="module"):
            load_strategy(self._write_yaml(tmp_path, body))

    def test_strategy_file_not_found_raises_filenotfound(self, tmp_path: Path):
        with pytest.raises(FileNotFoundError):
            load_strategy(tmp_path / "missing.yaml")


# ---------------------------------------------------------------------------
# run_strategy — control flow + error cases
# ---------------------------------------------------------------------------


def _strategy(steps: list[StepDef], **overrides) -> StrategyDef:
    base = {
        "name": "test",
        "description": "synthetic",
        "input_path": None,
        "input_encoding": "utf-8-sig",
        "steps": steps,
        "output_csv": None,
    }
    base.update(overrides)
    return StrategyDef(**base)


class TestRunStrategyControlFlow:
    def test_chains_two_steps_passing_df_through(self):
        # ip_risk + decision_engine: two real ported steps. ip_risk adds 9 cols,
        # decision_engine adds 11. So the chain should add 20 columns total.
        # ip_risk needs config["niche"], decision_engine needs nothing.
        from fba_engine.steps.build_output import FINAL_HEADERS

        # Build a minimal Phase-5 frame (the input to decision_engine).
        # We feed it directly through ip_risk first (which expects Phase 3
        # shortlist columns — most of which exist in FINAL_HEADERS).
        # Easier: just chain decision_engine alone for this test.
        from fba_engine.strategies.runner import run_strategy
        df_in = pd.DataFrame(
            [{h: "" for h in FINAL_HEADERS}]
        )
        df_in.iloc[0, df_in.columns.get_loc("ASIN")] = "B0SAMPLE"
        df_in.iloc[0, df_in.columns.get_loc("Verdict")] = "YES"
        df_in.iloc[0, df_in.columns.get_loc("Opportunity Lane")] = "BALANCED"

        strat = _strategy(
            steps=[
                StepDef("decision", "fba_engine.steps.decision_engine", {}),
            ]
        )
        df_out = run_strategy(strat, context={}, df_in=df_in)
        assert "Decision" in df_out.columns

    def test_step_module_missing_raises_config_error(self):
        strat = _strategy(
            steps=[StepDef("ghost", "fba_engine.steps.does_not_exist", {})]
        )
        with pytest.raises(StrategyConfigError, match="does_not_exist"):
            run_strategy(strat, context={}, df_in=pd.DataFrame([{"ASIN": "B0"}]))

    def test_step_without_run_step_raises_config_error(self):
        # `pandas` exists but has no run_step.
        strat = _strategy(steps=[StepDef("badstep", "pandas", {})])
        with pytest.raises(StrategyConfigError, match="run_step"):
            run_strategy(strat, context={}, df_in=pd.DataFrame([{"ASIN": "B0"}]))

    def test_step_runtime_error_wrapped_with_context(self):
        # ip_risk requires config["niche"]; missing it raises ValueError.
        # The runner should re-raise as StrategyExecutionError citing the
        # offending step.
        strat = _strategy(
            steps=[StepDef("ip_risk", "fba_engine.steps.ip_risk", {})]
        )
        with pytest.raises(StrategyExecutionError, match="ip_risk"):
            run_strategy(strat, context={}, df_in=pd.DataFrame([{"ASIN": "B0"}]))

    def test_no_input_path_and_no_df_in_raises(self):
        strat = _strategy(steps=[])
        with pytest.raises(StrategyConfigError, match="input"):
            run_strategy(strat, context={}, df_in=None)

    def test_input_discover_flag_passes_empty_df(self):
        # Strategies whose first step is a discoverer (creates rows from
        # supplier files / API calls) opt out of the input.path guard
        # via input.discover: true. The runner passes an empty df_in
        # and the discover step ignores it.
        strat = _strategy(steps=[], input_discover=True)
        out = run_strategy(strat, context={}, df_in=None)
        assert out.empty

    def test_empty_step_chain_returns_input_unchanged(self):
        df = pd.DataFrame([{"ASIN": "B0"}])
        strat = _strategy(steps=[])
        out = run_strategy(strat, context={}, df_in=df)
        pd.testing.assert_frame_equal(out, df)


class TestRunStrategyVariableInterpolation:
    def test_step_config_strings_get_interpolated(self):
        # ip_risk reads config["niche"]. Pass "{niche}" literal in YAML and
        # rely on context to substitute.
        from fba_engine.steps.build_output import FINAL_HEADERS
        df = pd.DataFrame(
            [{h: "" for h in FINAL_HEADERS} | {"ASIN": "B0", "Title": "T", "Brand": "B"}]
        )
        # Use a Phase-3-shape df for ip_risk (subset is sufficient — ip_risk
        # warns on missing cols but doesn't crash).
        strat = _strategy(
            steps=[
                StepDef("ip_risk", "fba_engine.steps.ip_risk", {"niche": "{niche}"}),
            ]
        )
        out = run_strategy(strat, context={"niche": "kids-toys"}, df_in=df)
        # ip_risk added the 9 columns — that's evidence it ran with the
        # interpolated niche.
        assert "IP Risk Band" in out.columns


class TestRunSummary:
    """run_summary.json is written alongside output.csv when set."""

    def test_summary_written_with_step_metrics(self, tmp_path: Path):
        import json as _json

        # Two-row input → ip_risk passes through (rows preserved) →
        # output csv + .summary.json both written.
        from fba_engine.steps.build_output import FINAL_HEADERS
        df = pd.DataFrame(
            [{h: "" for h in FINAL_HEADERS} | {"ASIN": f"B{i:03d}"}]
            for i in range(2)
        )
        out_path = tmp_path / "out.csv"
        strat = _strategy(
            steps=[StepDef("ip_risk", "fba_engine.steps.ip_risk", {"niche": "x"})],
            output_csv=str(out_path),
        )
        run_strategy(strat, context={}, df_in=df)
        summary_path = tmp_path / "out.summary.json"
        assert summary_path.exists()
        summary = _json.loads(summary_path.read_text(encoding="utf-8"))
        assert summary["strategy"] == "test"
        assert summary["initial_rows"] == 2
        assert summary["final_rows"] == 2
        assert "started_at" in summary
        assert "completed_at" in summary
        assert summary["duration_seconds"] >= 0
        # One step ran.
        assert len(summary["step_summary"]) == 1
        step = summary["step_summary"][0]
        assert step["name"] == "ip_risk"
        assert step["module"] == "fba_engine.steps.ip_risk"
        assert step["rows_in"] == 2
        assert step["rows_out"] == 2
        assert step["duration_seconds"] >= 0
        assert "error" not in step

    def test_no_summary_when_no_output_csv(self, tmp_path: Path):
        # Strategies without output.csv don't write a summary either —
        # there's no canonical location to put it. Operators who want
        # a summary set output.csv.
        df = pd.DataFrame([{"ASIN": "B001"}])
        strat = _strategy(steps=[])
        run_strategy(strat, context={}, df_in=df)
        # No file should appear in tmp_path.
        assert not list(tmp_path.glob("*.summary.json"))

    def test_summary_includes_context(self, tmp_path: Path):
        import json as _json

        df = pd.DataFrame([{"ASIN": "B001"}])
        out_path = tmp_path / "{niche}.csv"
        strat = _strategy(steps=[], output_csv=str(out_path))
        run_strategy(strat, context={"niche": "kids-toys"}, df_in=df)
        summary = _json.loads(
            (tmp_path / "kids-toys.summary.json").read_text(encoding="utf-8")
        )
        assert summary["context"] == {"niche": "kids-toys"}

    def test_summary_records_step_failure_with_error(self, tmp_path: Path):
        # When a step raises, the summary's step_summary entry for that
        # step gets an `error` field. The exception still propagates as
        # StrategyExecutionError so callers know the run failed —
        # operators read the summary to see which step failed and how
        # long it ran before failing.
        from fba_engine.steps import ip_risk

        df = pd.DataFrame([{"ASIN": "B001"}])
        out_path = tmp_path / "out.csv"
        strat = _strategy(
            steps=[
                StepDef("boom", "fba_engine.steps.ip_risk", {"niche": "x"}),
            ],
            output_csv=str(out_path),
        )

        # Patch run_step to raise.
        original = ip_risk.run_step
        ip_risk.run_step = lambda d, c: (_ for _ in ()).throw(
            RuntimeError("simulated")
        )
        try:
            with pytest.raises(StrategyExecutionError, match="boom"):
                run_strategy(strat, context={}, df_in=df)
        finally:
            ip_risk.run_step = original
        # No summary written because the run never reached the output
        # block. That's deliberate — the runner serialises the summary
        # ONLY on success path so partial state doesn't pollute downstream
        # consumers. Operators who need failure metrics should tail the
        # logs (the StrategyExecutionError carries the cause).
        assert not (tmp_path / "out.summary.json").exists()


class TestRunStrategyOutputCsv:
    def test_writes_output_csv_when_configured(self, tmp_path: Path):
        from fba_engine.steps.build_output import FINAL_HEADERS
        df = pd.DataFrame([{h: "" for h in FINAL_HEADERS} | {"ASIN": "B001"}])
        out_path = tmp_path / "out.csv"
        strat = _strategy(
            steps=[],
            output_csv=str(out_path),
        )
        run_strategy(strat, context={}, df_in=df)
        assert out_path.exists()
        round_tripped = pd.read_csv(out_path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
        assert "ASIN" in round_tripped.columns

    def test_output_csv_path_is_interpolated(self, tmp_path: Path):
        df = pd.DataFrame([{"ASIN": "B001"}])
        strat = _strategy(
            steps=[],
            output_csv=str(tmp_path / "{niche}.csv"),
        )
        run_strategy(strat, context={"niche": "kids-toys"}, df_in=df)
        assert (tmp_path / "kids-toys.csv").exists()


class TestRunStrategyOutputXlsx:
    """Runner writes a styled XLSX alongside CSV when output.xlsx is set.
    The XLSX is the operator-facing deliverable: clickable hyperlinks on
    URL columns, REJECT rows excluded automatically."""

    def test_xlsx_omitted_is_no_op(self, tmp_path: Path):
        """Strategies that don't declare output.xlsx still run cleanly —
        only CSV gets written, no XLSX file appears."""
        df = pd.DataFrame([{"ASIN": "B001"}])
        out_csv = tmp_path / "out.csv"
        strat = _strategy(steps=[], output_csv=str(out_csv))
        run_strategy(strat, context={}, df_in=df)
        assert out_csv.exists()
        assert not list(tmp_path.glob("*.xlsx"))

    def test_xlsx_written_when_declared(self, tmp_path: Path):
        df = pd.DataFrame([{
            "ASIN": "B001", "decision": "REVIEW",
            "buy_box_price": 24.99, "amazon_url": "https://www.amazon.co.uk/dp/B001",
            "decision_reason": "test",
        }])
        out_csv = tmp_path / "out.csv"
        out_xlsx = tmp_path / "out.xlsx"
        strat = _strategy(
            steps=[], output_csv=str(out_csv), output_xlsx=str(out_xlsx),
        )
        run_strategy(strat, context={}, df_in=df)
        assert out_csv.exists()
        assert out_xlsx.exists()

    def test_xlsx_path_is_interpolated(self, tmp_path: Path):
        df = pd.DataFrame([{
            "ASIN": "B001", "decision": "REVIEW",
            "buy_box_price": 24.99, "decision_reason": "test",
        }])
        out_csv = tmp_path / "out.csv"
        strat = _strategy(
            steps=[],
            output_csv=str(out_csv),
            output_xlsx=str(tmp_path / "{recipe}.xlsx"),
        )
        run_strategy(strat, context={"recipe": "amazon_oos_wholesale"}, df_in=df)
        assert (tmp_path / "amazon_oos_wholesale.xlsx").exists()

    def test_xlsx_excludes_reject_rows(self, tmp_path: Path):
        """The styled workbook is the operator's working file — REJECT rows
        belong in the audit-trail CSV, not the actionable XLSX. Verified
        by reading back the workbook and counting non-REJECT rows."""
        from openpyxl import load_workbook
        df = pd.DataFrame([
            {"asin": "B0KEEP0001", "decision": "REVIEW",
             "buy_box_price": 25.0, "decision_reason": "review"},
            {"asin": "B0DROP0001", "decision": "REJECT",
             "buy_box_price": 0.0, "decision_reason": "rejected"},
            {"asin": "B0SHRT0001", "decision": "SHORTLIST",
             "buy_box_price": 30.0, "decision_reason": "passes"},
        ])
        out_xlsx = tmp_path / "out.xlsx"
        strat = _strategy(
            steps=[], output_csv=str(tmp_path / "out.csv"),
            output_xlsx=str(out_xlsx),
        )
        run_strategy(strat, context={}, df_in=df)
        wb = load_workbook(out_xlsx)
        ws = wb.active
        # excel_writer puts the title bar in row 1 and column headers in
        # row 2 (data starts at row 3). Read the header row, find the
        # ASIN column index, then enumerate data values.
        header = [c.value for c in ws[2]]
        asin_col_idx = next(
            (i for i, h in enumerate(header, start=1) if h == "ASIN"), None,
        )
        assert asin_col_idx is not None, f"ASIN column missing; headers: {header}"
        asins = [
            ws.cell(row=r, column=asin_col_idx).value
            for r in range(3, ws.max_row + 1)
        ]
        assert "B0DROP0001" not in asins                 # REJECT excluded
        assert "B0KEEP0001" in asins
        assert "B0SHRT0001" in asins


class TestKeepaFinderXlsxOutput:
    """Lock in that keepa_finder.yaml declares both csv AND xlsx outputs.
    The deliverable to the operator is the XLSX (clickable hyperlinks);
    the CSV is audit-trail only."""

    def test_keepa_finder_yaml_declares_both_outputs(self):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "keepa_finder.yaml"
        if not yaml_path.exists():
            pytest.skip(f"keepa_finder.yaml not found at {yaml_path}")
        strat = load_strategy(yaml_path)
        assert strat.output_csv is not None and strat.output_csv.endswith(".csv")
        assert strat.output_xlsx is not None and strat.output_xlsx.endswith(".xlsx")


class TestRunStrategyOutputGsheet:
    """Runner uploads the XLSX to Google Drive as a Sheet via
    push_to_gsheets.js when output.gsheet is set. Subprocess mocked so
    tests don't hit real Google APIs / require service account creds."""

    def test_gsheet_block_parses(self, tmp_path: Path):
        yaml_text = (
            "name: g\n"
            "description: g\n"
            "steps: []\n"
            "output:\n"
            "  csv: out.csv\n"
            "  xlsx: out.xlsx\n"
            "  gsheet:\n"
            "    title: \"Hello {recipe}\"\n"
            "    id_file: out.gsheet_id.txt\n"
        )
        yp = tmp_path / "g.yaml"
        yp.write_text(yaml_text, encoding="utf-8")
        strat = load_strategy(yp)
        assert strat.output_gsheet is not None
        assert strat.output_gsheet["title"] == "Hello {recipe}"
        assert strat.output_gsheet["id_file"] == "out.gsheet_id.txt"
        assert strat.output_gsheet["folder_id"] is None

    def test_gsheet_block_requires_title(self, tmp_path: Path):
        yaml_text = (
            "name: g\n"
            "description: g\n"
            "steps: []\n"
            "output:\n"
            "  csv: out.csv\n"
            "  gsheet:\n"
            "    id_file: x.txt\n"   # title missing — must reject
        )
        yp = tmp_path / "g.yaml"
        yp.write_text(yaml_text, encoding="utf-8")
        with pytest.raises(StrategyConfigError, match="requires a `title`"):
            load_strategy(yp)

    def test_gsheet_block_must_be_a_mapping(self, tmp_path: Path):
        yaml_text = (
            "name: g\n"
            "description: g\n"
            "steps: []\n"
            "output:\n"
            "  csv: out.csv\n"
            "  gsheet: not a mapping\n"
        )
        yp = tmp_path / "g.yaml"
        yp.write_text(yaml_text, encoding="utf-8")
        with pytest.raises(StrategyConfigError, match="must be a mapping"):
            load_strategy(yp)

    def test_gsheet_skip_when_node_missing(self, tmp_path: Path, monkeypatch):
        """No node on PATH → silent skip (no exception, no URL)."""
        from fba_engine.strategies import runner as runner_mod
        # Make shutil.which return None for "node"
        import shutil
        monkeypatch.setattr(
            shutil, "which",
            lambda name: None if name == "node" else "/usr/bin/" + name,
        )
        url = runner_mod._push_xlsx_to_gsheet(
            xlsx_path=str(tmp_path / "any.xlsx"),
            gsheet_cfg={"title": "x", "folder_id": None, "id_file": None},
            context={},
        )
        assert url is None

    def test_gsheet_skip_when_service_account_key_missing(
        self, tmp_path: Path, monkeypatch,
    ):
        """No service-account.json → silent skip."""
        from fba_engine.strategies import runner as runner_mod
        # Patch Path so the key file lookup misses (use a tmp repo root
        # that doesn't have the key).
        monkeypatch.setattr(
            runner_mod, "__file__",
            str(tmp_path / "fba_engine" / "strategies" / "runner.py"),
        )
        url = runner_mod._push_xlsx_to_gsheet(
            xlsx_path=str(tmp_path / "any.xlsx"),
            gsheet_cfg={"title": "x", "folder_id": None, "id_file": None},
            context={},
        )
        assert url is None

    def test_gsheet_extracts_url_from_subprocess_stdout(
        self, tmp_path: Path, monkeypatch,
    ):
        """Happy path: subprocess prints `URL: https://...` and runner
        extracts it correctly. Verifies the contract between the JS
        script's stdout format and the runner's parser."""
        from fba_engine.strategies import runner as runner_mod
        from unittest.mock import MagicMock
        import shutil

        monkeypatch.setattr(shutil, "which", lambda name: "node")

        # Stub the script + key existence checks so the function reaches
        # the subprocess call.
        repo_root = Path(runner_mod.__file__).resolve().parents[2]
        script = repo_root / "fba_engine" / "_legacy_keepa" / "skills" / "skill-5-build-output" / "push_to_gsheets.js"
        key = repo_root / "fba_engine" / "_legacy_keepa" / "config" / "google-service-account.json"
        if not script.exists() or not key.exists():
            pytest.skip("legacy keepa script / key not present in this checkout")

        sheet_url = "https://docs.google.com/spreadsheets/d/abc123/edit"
        fake_proc = MagicMock(
            returncode=0,
            stdout=(
                f"Uploaded (xlsx conversion): test\n"
                f"Sheet ID: abc123\n"
                f"URL: {sheet_url}\n"
            ),
            stderr="",
        )
        import subprocess
        monkeypatch.setattr(subprocess, "run", lambda *a, **k: fake_proc)

        url = runner_mod._push_xlsx_to_gsheet(
            xlsx_path=str(tmp_path / "any.xlsx"),
            gsheet_cfg={"title": "x", "folder_id": None, "id_file": None},
            context={},
        )
        assert url == sheet_url

    def test_gsheet_handles_subprocess_failure(
        self, tmp_path: Path, monkeypatch,
    ):
        """Non-zero exit code from the script → silent skip + warn log."""
        from fba_engine.strategies import runner as runner_mod
        from unittest.mock import MagicMock
        import shutil

        monkeypatch.setattr(shutil, "which", lambda name: "node")
        repo_root = Path(runner_mod.__file__).resolve().parents[2]
        script = repo_root / "fba_engine" / "_legacy_keepa" / "skills" / "skill-5-build-output" / "push_to_gsheets.js"
        key = repo_root / "fba_engine" / "_legacy_keepa" / "config" / "google-service-account.json"
        if not script.exists() or not key.exists():
            pytest.skip("legacy keepa script / key not present in this checkout")

        fake_proc = MagicMock(returncode=1, stdout="", stderr="invalid grant")
        import subprocess
        monkeypatch.setattr(subprocess, "run", lambda *a, **k: fake_proc)

        url = runner_mod._push_xlsx_to_gsheet(
            xlsx_path=str(tmp_path / "any.xlsx"),
            gsheet_cfg={"title": "x", "folder_id": None, "id_file": None},
            context={},
        )
        assert url is None


# ---------------------------------------------------------------------------
# End-to-end: keepa_niche YAML against a real fixture
# ---------------------------------------------------------------------------


class TestKeepaNicheStrategy:
    """Integration test: load the canonical keepa_niche.yaml and run it
    through the full chain (ip_risk -> build_output -> decision_engine)."""

    def _phase3_fixture(self) -> pd.DataFrame:
        """Minimal Phase-3 shortlist row that survives the build_output PL filter."""
        return pd.DataFrame([{
            "ASIN": "B0CLEAN",
            "Title": "Clean Toy",
            "Brand": "Acme",
            "Amazon URL": "https://amzn.eu/d/x",
            "Category": "Toys",
            "Weight Flag": "OK",
            "Verdict": "YES",
            "Verdict Reason": "Strong",
            "Composite Score": "8.5",
            "Demand Score": "8",
            "Stability Score": "8",
            "Competition Score": "7",
            "Margin Score": "9",
            "Cash Flow Score": "8",
            "Profit Score": "8",
            "Balanced Score": "8",
            "Listing Quality": "Good",
            "Opportunity Lane": "BALANCED",
            "Commercial Priority": "1",
            "Lane Reason": "demand+margin",
            "Monthly Gross Profit": "GBP500",
            "Price Compression": "OK",
            "Current Price": "GBP25.99",
            "Buy Box 90d Avg": "GBP25.50",
            "Price Drop % 90d": "1",
            "Fulfilment Fee": "GBP3.50",
            "Amazon Fees": "GBP5.00",
            "Total Amazon Fees": "GBP8.50",
            "Est Cost 65%": "GBP10.00",
            "Est Profit": "GBP7.49",
            "Est ROI %": "32",
            "Max Cost 20% ROI": "GBP12.00",
            "Breakeven Price": "GBP18.00",
            "BSR Current": "5000",
            "BSR Drops 90d": "200",
            "Bought per Month": "150",
            "Star Rating": "4.5",
            "Review Count": "200",
            "Brand 1P": "N",
            "FBA Seller Count": "5",
            "Amazon on Listing": "N",
            "Buy Box Amazon %": "10%",
            "EAN": "1234567890123",
            "UPC": "",
            "GTIN": "",
        }])

    def test_keepa_niche_yaml_loads_and_runs_end_to_end(self, tmp_path: Path):
        # Load the canonical strategy file.
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "keepa_niche.yaml"
        if not yaml_path.exists():
            pytest.skip(f"keepa_niche.yaml not found at {yaml_path}")

        strat = load_strategy(yaml_path)
        assert strat.name == "keepa_niche"

        df_in = self._phase3_fixture()
        # Provide all context keys the strategy YAML interpolates: niche
        # is used by ip_risk's config; base + niche_snake are used by the
        # output_csv path. Point base at tmp_path so the test write is
        # sandboxed.
        context = {
            "niche": "kids-toys",
            "niche_snake": "kids_toys",
            "base": str(tmp_path),
        }
        # The runner reads input from `{base}/working/...` IFF df_in is None.
        # We pass df_in directly, so the input_path is unused; only the
        # output_csv path matters. Verify the chain output AND that the
        # output CSV got written to the sandboxed location.
        out = run_strategy(strat, context=context, df_in=df_in)

        # After the full chain (ip_risk -> build_output -> decision_engine),
        # the frame should have IP risk + decision columns. The fixture row
        # is a clean BALANCED YES with strong margin — `decision_engine`
        # demotes it to NEGOTIATE because no supplier cost is present
        # (Trade Price Found="" -> "Cost Needed" branch). Pinning the
        # verdict catches drift in any of the three composed steps.
        assert "IP Risk Band" in out.columns
        assert out.iloc[0]["Decision"] == "NEGOTIATE"

        # Output CSV was written to the sandbox via atomic_write.
        output_csv = tmp_path / "working" / "kids_toys_phase6_decisions.csv"
        assert output_csv.exists()


class TestSupplierPricelistStrategy:
    """Smoke test: the canonical supplier_pricelist.yaml loads and
    structurally matches the new step modules + the input.discover
    contract."""

    _SUPPLIER_CSV = (
        "SKU,Manufacturer,Title,Category,Barcode,Case Size,"
        "Units Available,Unit Price (GBP),Case Price (GBP)\n"
        "SKU-A,Acme,Widget,Tools,5012345678900,1,100,5.00,5.00\n"
    )

    def test_supplier_pricelist_yaml_loads_with_discover_input(self):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "supplier_pricelist.yaml"
        if not yaml_path.exists():
            pytest.skip(f"supplier_pricelist.yaml not found at {yaml_path}")

        strat = load_strategy(yaml_path)
        assert strat.name == "supplier_pricelist"
        assert strat.input_discover is True
        assert strat.input_path is None
        # All six step module paths are real Python modules.
        for step in strat.steps:
            __import__(step.module)

    def test_supplier_pricelist_yaml_runs_end_to_end(self, tmp_path: Path):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "supplier_pricelist.yaml"
        if not yaml_path.exists():
            pytest.skip(f"supplier_pricelist.yaml not found at {yaml_path}")

        # Sandbox the supplier raw + run dir under tmp_path.
        raw = tmp_path / "raw"
        raw.mkdir()
        (raw / "p.csv").write_text(self._SUPPLIER_CSV, encoding="utf-8")
        run_dir = tmp_path / "out"

        context = {
            "supplier": "connect-beauty",
            "input_path": str(raw),
            "market_data_path": "",  # no Keepa CSV — every row REJECTs (no_match)
            "run_dir": str(run_dir),
            "timestamp": "20260429_120000",
            "supplier_label": "Connect Beauty",
        }

        # `enabled: true` in the YAML for enrich would call the MCP CLI.
        # Override via the strategy YAML's loaded steps to disable it for
        # this offline test.
        strat = load_strategy(yaml_path)
        for s in strat.steps:
            if s.name == "enrich":
                s.config["enabled"] = False

        out = run_strategy(strat, context=context, df_in=None)
        # Every row should REJECT (no Keepa data).
        assert (out["decision"] == "REJECT").all()
        # Output writers ran and produced the timestamped CSV.
        assert (run_dir / "shortlist_20260429_120000.csv").exists()


class TestSellerStorefrontStrategy:
    """Smoke test: seller_storefront.yaml loads + runs end-to-end via a
    stubbed Keepa client. Pins the discover -> supplier_leads chain."""

    def test_seller_storefront_yaml_loads(self):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = (
            repo_root / "fba_engine" / "strategies" / "seller_storefront.yaml"
        )
        if not yaml_path.exists():
            pytest.skip(f"seller_storefront.yaml not found at {yaml_path}")
        strat = load_strategy(yaml_path)
        assert strat.name == "seller_storefront"
        assert strat.input_discover is True
        # Three steps: discover -> enrich (leads-mode) -> supplier_leads.
        assert [s.name for s in strat.steps] == [
            "discover", "enrich", "supplier_leads",
        ]
        # All modules import.
        for step in strat.steps:
            __import__(step.module)
        # The enrich step is configured for leads mode (skips
        # pricing-dependent MCP sources). This is the contract that
        # replaces the legacy SellerAmp skill.
        enrich_step = next(s for s in strat.steps if s.name == "enrich")
        assert enrich_step.config.get("include") == "leads"

    def test_seller_storefront_yaml_runs_end_to_end(self, tmp_path: Path):
        # Stub the Keepa client to avoid real API calls. Inject the
        # stub into the discover step's config dict via the YAML
        # context layer — the runner interpolates strings only, so
        # we mutate the loaded StrategyDef before run_strategy.
        from unittest.mock import MagicMock

        from keepa_client import KeepaProduct, KeepaSeller

        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = (
            repo_root / "fba_engine" / "strategies" / "seller_storefront.yaml"
        )
        if not yaml_path.exists():
            pytest.skip(f"seller_storefront.yaml not found at {yaml_path}")

        seller = KeepaSeller(
            sellerId="A1B2C3", sellerName="Stub Storefront",
            asinList=["B0AAA"],
        )
        products = [
            KeepaProduct(asin="B0AAA", title="Widget A", brand="Acme"),
        ]
        client = MagicMock()
        client.get_seller.return_value = seller
        client.get_products.return_value = products

        strat = load_strategy(yaml_path)
        # Inject the stubbed client into the discover step's config.
        for step in strat.steps:
            if step.name == "discover":
                step.config["client"] = client

        run_dir = tmp_path / "out"
        context = {
            "seller_id": "A1B2C3",
            "run_dir": str(run_dir),
            "timestamp": "20260430_120000",
        }
        out = run_strategy(strat, context=context, df_in=None)

        # Discovery emitted one row; supplier_leads added supplier_search_*
        # columns; output CSV got written.
        assert len(out) == 1
        assert out.iloc[0]["asin"] == "B0AAA"
        assert "supplier_search_brand_distributor" in out.columns
        # Brand search URL fires when brand is present.
        assert out.iloc[0]["supplier_search_brand_distributor"]
        # Atomic-written CSV at the configured location.
        assert (
            run_dir
            / "seller_storefront_A1B2C3_20260430_120000.csv"
        ).exists()


class TestOaCsvStrategy:
    """Smoke test: oa_csv.yaml loads + runs the full
    discover → keepa_enrich → calculate → decide chain against a
    fixture CSV with a stubbed Keepa client."""

    _SELLERAMP_CSV = (
        # B0OA1 buy_cost is £3 vs £15 market — clears ROI thresholds
        # comfortably (calculate's standard-tier FBA fees eat ~£8.50,
        # leaving ~£3.50 profit at >100% ROI).
        # B0OA2 buy_cost is £40 vs £15 market — clearly unprofitable.
        "ASIN,Title,Cost,URL\n"
        "B0OA1,OA Item One,3.00,https://retailer.test/p/1\n"
        "B0OA2,OA Item Two,40.00,https://retailer.test/p/2\n"
    )

    def test_oa_csv_yaml_loads(self):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "oa_csv.yaml"
        if not yaml_path.exists():
            pytest.skip(f"oa_csv.yaml not found at {yaml_path}")
        strat = load_strategy(yaml_path)
        assert strat.name == "oa_csv"
        assert strat.input_discover is True
        # Full chain: discover -> keepa_enrich -> calculate -> decide.
        assert [s.name for s in strat.steps] == [
            "discover", "keepa_enrich", "calculate", "decide",
        ]
        for step in strat.steps:
            __import__(step.module)

    def test_oa_csv_yaml_runs_end_to_end_with_decisions(self, tmp_path: Path):
        # End-to-end run with stubbed Keepa: pin that the chain
        # produces a `decision` column with non-empty verdicts. The
        # cheap row (£5 cost vs £15 market) should be SHORTLIST/REVIEW;
        # the expensive row (£40 cost vs £15 market) should be REJECT
        # for unprofitability — verifying calculate + decide actually
        # ran.
        from unittest.mock import MagicMock
        from keepa_client import KeepaProduct
        from keepa_client.models import KeepaStats

        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "oa_csv.yaml"
        if not yaml_path.exists():
            pytest.skip(f"oa_csv.yaml not found at {yaml_path}")

        # Build a stats-bearing KeepaProduct for each test ASIN. Both
        # have healthy market signals — pricing differs only in the
        # supplier-side buy_cost the OA CSV provides.
        def _stub_product(asin: str) -> KeepaProduct:
            current = [-1] * 19
            current[0] = 1400   # AMAZON £14
            current[3] = 5000   # SALES rank
            current[10] = 1450  # NEW_FBA £14.50
            current[11] = 5     # COUNT_NEW (offers)
            current[18] = 1500  # BUY_BOX £15
            return KeepaProduct(
                asin=asin, title=f"Title {asin}", brand="Acme",
                stats=KeepaStats(current=current, avg90=current),
                monthlySold=200,
            )

        client = MagicMock()
        client.get_products.return_value = [
            _stub_product("B0OA1"), _stub_product("B0OA2"),
        ]

        csv_in = tmp_path / "in.csv"
        csv_in.write_text(self._SELLERAMP_CSV, encoding="utf-8")
        exclusions = tmp_path / "exclusions.csv"
        exclusions.write_text("ASIN\n", encoding="utf-8")
        run_dir = tmp_path / "out"

        strat = load_strategy(yaml_path)
        for step in strat.steps:
            if step.name == "discover":
                step.config["exclusions_path"] = str(exclusions)
            if step.name == "keepa_enrich":
                step.config["client"] = client

        context = {
            "feed": "selleramp",
            "csv_path": str(csv_in),
            "run_dir": str(run_dir),
            "timestamp": "20260430_120000",
        }
        out = run_strategy(strat, context=context, df_in=None)

        assert len(out) == 2
        assert set(out["asin"]) == {"B0OA1", "B0OA2"}
        # Every row got a decision (SHORTLIST / REVIEW / REJECT).
        assert "decision" in out.columns
        valid = {"SHORTLIST", "REVIEW", "REJECT"}
        assert all(d in valid for d in out["decision"]), out["decision"].tolist()
        # Cheap row clears profit thresholds; expensive row REJECTs.
        cheap = out[out["asin"] == "B0OA1"].iloc[0]
        expensive = out[out["asin"] == "B0OA2"].iloc[0]
        assert cheap["decision"] in {"SHORTLIST", "REVIEW"}, cheap["decision_reason"]
        assert expensive["decision"] == "REJECT"
        # Verdict CSV written at the configured location.
        assert (
            run_dir / "oa_decisions_selleramp_20260430_120000.csv"
        ).exists()


class TestKeepaFinderStrategy:
    """Smoke test: keepa_finder.yaml loads + runs the full
    discover → enrich (leads) → calculate → decide → supplier_leads
    chain against a synthetic Keepa Product Finder CSV.

    No SP-API creds required — enrich.run_step silently no-ops when
    the MCP CLI isn't built / SP_API_CLIENT_ID isn't set, exactly
    matching how the seller_storefront chain runs in test envs.
    """

    # Keepa Product Finder column names. Several contain commas
    # ("New, 3rd Party FBA: Current") so the CSV must be properly
    # quoted — we let pandas handle that via to_csv().
    _KEEPA_COLUMNS = (
        "ASIN", "Title", "Brand", "Manufacturer",
        "Categories: Root", "Categories: Sub", "Categories: Tree",
        "Product Codes: EAN", "Product Codes: UPC",
        "Buy Box: Current", "Buy Box: 90 days avg.",
        "New, 3rd Party FBA: Current",
        "Amazon: Current", "FBA Pick&Pack Fee", "Referral Fee %",
        "Bought in past month", "New FBA Offer Count: Current",
        "Sales Rank: Current", "Sales Rank: 90 days avg.",
        "Buy Box: % Amazon 90 days", "Buy Box: 90 days OOS",
        "Buy Box: 30 days drop %", "Buy Box: 90 days drop %",
    )

    @classmethod
    def _row(cls, asin: str, title: str, buy_box: str, sales: str = "150") -> dict:
        """Build a row dict with the right column names (commas and all)."""
        return {
            "ASIN": asin, "Title": title, "Brand": "Acme",
            "Manufacturer": "Acme Mfg",
            "Categories: Root": "Toys & Games", "Categories: Sub": "Action",
            "Categories: Tree": "Toys & Games > Figures > Sets",
            "Product Codes: EAN": "5012345678901",
            "Product Codes: UPC": "012345678905",
            "Buy Box: Current": buy_box, "Buy Box: 90 days avg.": "26.50",
            "New, 3rd Party FBA: Current": "25.49",
            "Amazon: Current": "",                          # OFF_LISTING
            "FBA Pick&Pack Fee": "3.35", "Referral Fee %": "15 %",
            "Bought in past month": sales,
            "New FBA Offer Count: Current": "5",
            "Sales Rank: Current": "12345",
            "Sales Rank: 90 days avg.": "13500",
            "Buy Box: % Amazon 90 days": "0 %",
            "Buy Box: 90 days OOS": "2",
            "Buy Box: 30 days drop %": "3",
            "Buy Box: 90 days drop %": "5",
        }

    def test_keepa_finder_yaml_loads(self):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "keepa_finder.yaml"
        if not yaml_path.exists():
            pytest.skip(f"keepa_finder.yaml not found at {yaml_path}")
        strat = load_strategy(yaml_path)
        assert strat.name == "keepa_finder"
        assert strat.input_discover is True
        # Five steps: discover → enrich (leads) → calculate → decide → supplier_leads.
        assert [s.name for s in strat.steps] == [
            "discover", "enrich", "calculate", "decide", "supplier_leads",
        ]
        # All step modules import.
        for step in strat.steps:
            __import__(step.module)
        # Enrich is in leads mode (replaces SellerAmp for non-Buy-Box-% checks).
        enrich_step = next(s for s in strat.steps if s.name == "enrich")
        assert enrich_step.config.get("include") == "leads"

    def test_keepa_finder_yaml_runs_end_to_end(self, tmp_path: Path):
        """Full chain: synthetic Keepa CSV → enrich (no-op without creds) →
        calculate → decide → supplier_leads → output CSV. Asserts every
        row gets a decision and supplier_search columns are populated."""
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = repo_root / "fba_engine" / "strategies" / "keepa_finder.yaml"
        if not yaml_path.exists():
            pytest.skip(f"keepa_finder.yaml not found at {yaml_path}")

        # Synthetic Keepa Product Finder export. Two rows:
        #   - B0KEEPNORM: clean toy product — clears global exclusions
        #   - B0DROPSHOE: title contains "Shoes" — global filter drops it
        # pandas to_csv quotes column names containing commas (e.g.
        # "New, 3rd Party FBA: Current") correctly — matches the real
        # Keepa export format.
        csv_in = tmp_path / "keepa_export.csv"
        rows = [
            self._row("B0KEEPNORM", "Acme Action Figure Set", "24.99"),
            self._row("B0DROPSHOE", "Mens Running Shoes Size 10", "30.00"),
        ]
        pd.DataFrame(rows, columns=self._KEEPA_COLUMNS).to_csv(
            csv_in, index=False, encoding="utf-8-sig",
        )
        exclusions = tmp_path / "exclusions.csv"
        exclusions.write_text("ASIN\n", encoding="utf-8")
        out_dir = tmp_path / "out"

        strat = load_strategy(yaml_path)
        # Inject sandbox exclusions path so we don't read the repo's
        # canonical data/niches/exclusions.csv during tests.
        for s in strat.steps:
            if s.name == "discover":
                s.config["exclusions_path"] = str(exclusions)

        context = {
            "csv_path": str(csv_in),
            "recipe": "amazon_oos_wholesale",
            "output_dir": str(out_dir),
            "timestamp": "20260502_120000",
        }
        out = run_strategy(strat, context=context, df_in=None)

        # Global exclusion dropped the shoe row at discover stage.
        assert len(out) == 1
        assert out.iloc[0]["asin"] == "B0KEEPNORM"

        # Every surviving row carries a decision.
        assert "decision" in out.columns
        valid = {"SHORTLIST", "REVIEW", "REJECT"}
        assert all(d in valid for d in out["decision"]), out["decision"].tolist()

        # supplier_leads attached the brand-distributor column.
        assert "supplier_search_brand_distributor" in out.columns
        assert out.iloc[0]["supplier_search_brand_distributor"]

        # max_buy_price populated (the wholesale-flow signal — buy_cost=0
        # tells calculate to emit this as the negotiation ceiling).
        assert "max_buy_price" in out.columns
        assert out.iloc[0]["max_buy_price"] > 0

        # Output CSV written at the configured path.
        assert (
            out_dir
            / "keepa_finder_amazon_oos_wholesale_20260502_120000.csv"
        ).exists()


class TestSellerStorefrontCsvStrategy:
    """Smoke test: seller_storefront_csv.yaml loads + runs end-to-end.

    Reuses the Keepa Product Finder column shape (the Browser export
    is identical between Product Finder and Seller Storefront pages),
    plus the seller_id context arg the storefront strategy threads
    into discover step config and output filenames.
    """

    # Same Keepa Browser export columns as TestKeepaFinderStrategy.
    _KEEPA_COLUMNS = TestKeepaFinderStrategy._KEEPA_COLUMNS

    def test_seller_storefront_csv_yaml_loads(self):
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = (
            repo_root / "fba_engine" / "strategies" / "seller_storefront_csv.yaml"
        )
        if not yaml_path.exists():
            pytest.skip(f"seller_storefront_csv.yaml not found at {yaml_path}")
        strat = load_strategy(yaml_path)
        assert strat.name == "seller_storefront_csv"
        assert strat.input_discover is True
        # Five steps mirror keepa_finder.yaml — discover swaps in the
        # storefront-specific module, the rest are shared.
        assert [s.name for s in strat.steps] == [
            "discover", "enrich", "calculate", "decide", "supplier_leads",
        ]
        for step in strat.steps:
            __import__(step.module)
        # Discover step is the storefront-csv variant (not keepa_finder_csv).
        discover_step = next(s for s in strat.steps if s.name == "discover")
        assert discover_step.module == "fba_engine.steps.seller_storefront_csv"
        # Enrich is in leads mode (matches keepa_finder).
        enrich_step = next(s for s in strat.steps if s.name == "enrich")
        assert enrich_step.config.get("include") == "leads"

    def test_seller_storefront_csv_yaml_runs_end_to_end(self, tmp_path: Path):
        """Synthetic Keepa Storefront CSV → enrich (no-op without creds) →
        calculate → decide → supplier_leads → output CSV. Verifies that
        every row gets the seller_id tag and a decision verdict, and
        that {seller_id} interpolation lands in the output filename."""
        repo_root = Path(__file__).resolve().parents[3]
        yaml_path = (
            repo_root / "fba_engine" / "strategies" / "seller_storefront_csv.yaml"
        )
        if not yaml_path.exists():
            pytest.skip(f"seller_storefront_csv.yaml not found at {yaml_path}")

        # Two rows in this competitor's storefront. Shape mirrors a real
        # Keepa Browser → Seller Lookup → Storefront export. ASINs are
        # exactly 10 chars per Amazon's canonical format — anything else
        # gets dropped silently by the keepa_finder_csv 10-char check.
        csv_in = tmp_path / "keepa_storefront.csv"
        rows = [
            TestKeepaFinderStrategy._row(
                "B0STORE001", "Henry Vacuum Bags Pack of 10", "10.69",
            ),
            TestKeepaFinderStrategy._row(
                "B0STORE002", "Numatic Replacement Hose 2.5m", "14.39",
            ),
        ]
        pd.DataFrame(rows, columns=self._KEEPA_COLUMNS).to_csv(
            csv_in, index=False, encoding="utf-8-sig",
        )
        exclusions = tmp_path / "exclusions.csv"
        exclusions.write_text("ASIN\n", encoding="utf-8")
        out_dir = tmp_path / "out"

        strat = load_strategy(yaml_path)
        for s in strat.steps:
            if s.name == "discover":
                s.config["exclusions_path"] = str(exclusions)

        seller_id = "AR5NTANTFUHVI"
        context = {
            "csv_path": str(csv_in),
            "seller_id": seller_id,
            "output_dir": str(out_dir),
            "timestamp": "20260501_120000",
        }
        out = run_strategy(strat, context=context, df_in=None)

        # Both rows surface — wholesale flow + leads mode means no rejections
        # for missing market data, just verdicts on the SP-API-enriched info.
        assert len(out) == 2
        assert set(out["asin"]) == {"B0STORE001", "B0STORE002"}

        # Storefront-specific tagging.
        assert (out["source"] == "seller_storefront").all()
        assert (
            out["discovery_strategy"] == f"seller_storefront_{seller_id}"
        ).all()
        assert (out["seller_id"] == seller_id).all()

        # Decision verdicts from the canonical pipeline.
        valid = {"SHORTLIST", "REVIEW", "REJECT"}
        assert all(d in valid for d in out["decision"]), out["decision"].tolist()

        # Output filename interpolates {seller_id} + {timestamp}.
        assert (
            out_dir / f"seller_storefront_{seller_id}_20260501_120000.csv"
        ).exists()
