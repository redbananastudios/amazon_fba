"""Push-to-Google-Sheets step (Phase 5 — formerly Skill 5 part 3 in the legacy Keepa pipeline).

Uploads a styled XLSX (produced by step 4c.2) to Google Drive, with a
three-strategy fallback chain:

  1. **xlsx-conversion upload** — `files.create` with `mimeType =
     application/vnd.google-apps.spreadsheet` so Drive auto-converts
     xlsx to a native Google Sheet (preserves all openpyxl styling).
  2. **raw xlsx upload** — same `files.create` but with the xlsx mimetype,
     uploading the file as-is. Falls back here when Strategy 1 hits a
     storage-quota error (typical for service accounts not in a Workspace
     domain).
  3. **Sheets API create + populate** — creates a new spreadsheet via the
     Sheets API, moves it to the target folder, writes CSV rows + legend
     data, applies basic formatting (freeze panes, bold header row, auto-
     resize). No styling parity with the openpyxl xlsx — this is the
     last-resort path when Drive uploads are blocked entirely.

Logic ported 1:1 from `fba_engine/_legacy_keepa/skills/skill-5-build-output/
push_to_gsheets.js` (311 LOC).

**Required Python packages (already installed in this env):**
  - google-api-python-client
  - google-auth

**Required external setup:**
  - Service account JSON key (default path: `config/google-service-account.json`)
  - Target Drive folder shared with the service account email

Standalone CLI invocation:

    python -m fba_engine.steps.push_to_gsheets \\
        --niche kids-toys \\
        --base fba_engine/data/niches/kids-toys
"""
from __future__ import annotations

import argparse
import math
import sys
import time
from datetime import date
from pathlib import Path

import pandas as pd

from fba_engine.steps._helpers import atomic_write

# ────────────────────────────────────────────────────────────────────────
# Constants — pinned to the legacy values for direct migration.
# ────────────────────────────────────────────────────────────────────────

DEFAULT_FOLDER_ID = "1uFYiz7rYFm5ZJHgkXJh86jKaigK9H6yd"
DEFAULT_KEY_PATH = "config/google-service-account.json"

XLSX_MIMETYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
SHEETS_MIMETYPE = "application/vnd.google-apps.spreadsheet"

DEFAULT_SCOPES = (
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
)

# Legend body for the Sheets-API fallback path. Must match the legacy JS
# byte-for-byte so the fallback sheet renders the same explanations.
LEGEND_DATA: list[list[str]] = [
    ["Verdict", "Meaning"],
    ["YES", "Composite 7+, all filters pass -- pursue this product"],
    ["MAYBE", "Composite 5-6, one concern -- review needed"],
    ["MAYBE-ROI", "ROI below 20% estimated -- may improve with real trade price"],
    ["BRAND APPROACH", "2-3 sellers, weak listing -- contact brand direct"],
    ["BUY THE DIP", "Price 30%+ below 90-day avg -- recovery pattern detected"],
    ["PRICE EROSION", "Consistent downward slope -- reject"],
    ["GATED", "Restricted listing -- flag for ungating decision"],
    ["NO", "Fails filter, reason stated"],
    ["", ""],
    ["Lane", "Meaning"],
    ["BALANCED", "Strong velocity + strong margin -- premium opportunity"],
    ["PROFIT", "Strong unit profit/ROI, may have lower velocity -- capital efficient"],
    ["CASH FLOW", "High velocity, acceptable margin -- turnover and cash generation"],
    ["", ""],
    ["Column", "Explanation"],
    ["Price Compression", "OK / SQUEEZED (10-20% below 90d avg) / COMPRESSED (20%+ below avg)"],
    ["Est Cost 65%", "Estimated cost at 65% of selling price (rough placeholder)"],
    ["Max Cost for 20% ROI", "Maximum you can pay for stock and still hit 20% ROI after FBA fees"],
    ["Price Stability", "STABLE / SLIGHT DIP / DROPPING / RISING / SURGING"],
    ["Listing Quality", "STRONG (6+ images, A+, 5+ bullets) / AVERAGE / WEAK"],
    ["Weight Flag", "OK / HEAVY (>5kg) / OVERSIZE (>45cm) / HEAVY+OVERSIZE"],
    ["Brand 1P", "Y = brand sells direct on Amazon (hard to compete)"],
]


class PushFailedError(RuntimeError):
    """All three upload strategies exhausted without success."""


# ────────────────────────────────────────────────────────────────────────
# Auth + client construction.
# ────────────────────────────────────────────────────────────────────────


def _build_clients(key_path: str | Path):
    """Return (drive, sheets) Google API client objects.

    Imported lazily so the module loads cleanly when googleapis isn't
    installed (the auth/build path is only reached when actually pushing).
    """
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    credentials = service_account.Credentials.from_service_account_file(
        str(key_path), scopes=list(DEFAULT_SCOPES)
    )
    drive = build("drive", "v3", credentials=credentials, cache_discovery=False)
    sheets = build(
        "sheets", "v4", credentials=credentials, cache_discovery=False
    )
    return drive, sheets


# ────────────────────────────────────────────────────────────────────────
# Helpers.
# ────────────────────────────────────────────────────────────────────────


def _http_status(err: Exception) -> int | None:
    """Best-effort extraction of HTTP status from googleapiclient errors."""
    resp = getattr(err, "resp", None)
    status = getattr(resp, "status", None) if resp is not None else None
    if status is None:
        return None
    try:
        return int(status)
    except (TypeError, ValueError):
        return None


# Canonical phrases Google returns for quota/rate problems. Substring match
# only on these (the bare word "quota" caught false positives like
# "API quota check timed out").
_QUOTA_PHRASES = ("storagequotaexceeded", "storagequota", "quotaexceeded")
# Distinguish quota from rate limits: rate limits are transient (retry),
# storage quota is not (fall through to next strategy).
_RATE_LIMIT_PHRASES = ("userratelimitexceeded", "ratelimitexceeded")

# HTTP statuses that are worth retrying with exponential backoff.
TRANSIENT_STATUSES = frozenset({429, 500, 502, 503, 504})


def _is_quota_error(err: Exception) -> bool:
    """True iff `err` represents a Drive *storage*-quota failure.

    Storage-quota errors are non-transient — the right response is to
    fall through to a different upload strategy, not retry. Rate-limit
    errors look similar (both can surface as 403) but are transient and
    should be handled by the retry helper.
    """
    msg = str(err).lower()
    if any(phrase in msg for phrase in _QUOTA_PHRASES):
        return True
    # 403 alone is too broad (covers permission denied), so we only
    # accept it when accompanied by a quota/storage hint.
    if _http_status(err) == 403 and "storage" in msg:
        return True
    return False


def _is_transient(err: Exception) -> bool:
    """True iff retrying after a backoff is likely to succeed."""
    status = _http_status(err)
    if status in TRANSIENT_STATUSES:
        return True
    msg = str(err).lower()
    if any(phrase in msg for phrase in _RATE_LIMIT_PHRASES):
        return True
    # Network-level transients — connection reset, broken pipe, DNS blip.
    name = type(err).__name__
    if name in {"ConnectionError", "ConnectionResetError", "TimeoutError",
                "ServerNotFoundError", "ProtocolError"}:
        return True
    return False


def _retry_with_backoff(call, *, max_attempts: int = 4, base_delay: float = 2.0,
                        sleep=time.sleep):
    """Call `call()` with exponential-backoff retry on transient errors.

    Backoff schedule: base * (2 ** attempt) → 2s, 4s, 8s, 16s. Non-transient
    errors propagate immediately. The `sleep` parameter is injected so
    tests can run without real delays.
    """
    last_err: Exception | None = None
    for attempt in range(max_attempts):
        try:
            return call()
        except Exception as err:  # noqa: BLE001 — re-raised on non-transient
            if not _is_transient(err) or attempt == max_attempts - 1:
                raise
            last_err = err
            delay = base_delay * (2 ** attempt)
            print(
                f"Transient error (attempt {attempt + 1}/{max_attempts}); "
                f"retrying in {delay}s: {err}",
                file=sys.stderr,
            )
            sleep(delay)
    # Defensive — should never reach here; loop either returns or raises.
    if last_err is not None:
        raise last_err
    raise RuntimeError("retry loop exited without result or exception")


def csv_rows_for_upload(df: pd.DataFrame) -> list[list[str]]:
    """Convert a DataFrame to the row-of-rows shape Sheets API expects.

    First row is the header. NaN cells become empty strings (NOT 'nan' the
    literal). All values are stringified — Sheets API accepts mixed types
    but explicit strings keep the upload deterministic.
    """
    rows: list[list[str]] = [list(df.columns)]
    for _, row in df.iterrows():
        out_row: list[str] = []
        for v in row:
            if v is None:
                out_row.append("")
                continue
            if isinstance(v, float) and math.isnan(v):
                out_row.append("")
                continue
            try:
                if pd.isna(v):
                    out_row.append("")
                    continue
            except (TypeError, ValueError):
                pass
            out_row.append(str(v))
        rows.append(out_row)
    return rows


def _delete_previous_sheet(drive, previous_id: str | None) -> bool:
    """Delete the prior sheet at `previous_id`. Returns True on success.

    Errors are logged to stderr and swallowed — a stale ID file shouldn't
    block a fresh upload.
    """
    if not previous_id:
        return False
    try:
        drive.files().delete(
            fileId=previous_id, supportsAllDrives=True
        ).execute()
        print(f"Deleted previous sheet: {previous_id}")
        return True
    except Exception as err:
        print(
            f"Could not delete previous sheet ({previous_id}): {err}",
            file=sys.stderr,
        )
        return False


# ────────────────────────────────────────────────────────────────────────
# Strategy 1 — xlsx-conversion upload.
# ────────────────────────────────────────────────────────────────────────


# Chunk size for resumable uploads. 5 MiB is the smallest size Drive will
# accept for non-final chunks and keeps memory pressure low for large XLSX.
_RESUMABLE_CHUNK_SIZE = 5 * 1024 * 1024


def _execute_resumable(request) -> dict:
    """Drive resumable uploads pump via `next_chunk()` until done.

    Each chunk is a separate HTTP request; transient errors mid-upload
    can be retried per-chunk via `_retry_with_backoff` without restarting
    the whole upload. Returns the final API response dict.
    """
    response = None
    while response is None:
        _status, response = _retry_with_backoff(request.next_chunk)
    return response


def _upload_with_conversion(
    drive, xlsx_path: Path, name: str, folder_id: str
) -> str:
    """Upload xlsx to Drive with auto-conversion to a Google Sheet.

    Returns the new spreadsheet's id. Raises whatever the API raises —
    callers detect quota errors via `_is_quota_error` and fall through
    to Strategy 2; transient errors are retried internally via the
    resumable chunk loop.
    """
    from googleapiclient.http import MediaFileUpload

    file_metadata = {"name": name, "mimeType": SHEETS_MIMETYPE}
    if folder_id:
        file_metadata["parents"] = [folder_id]

    media = MediaFileUpload(
        str(xlsx_path), mimetype=XLSX_MIMETYPE,
        resumable=True, chunksize=_RESUMABLE_CHUNK_SIZE,
    )
    request = drive.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True,
    )
    res = _execute_resumable(request)
    return res["id"]


# ────────────────────────────────────────────────────────────────────────
# Strategy 2 — raw xlsx upload (no conversion).
# ────────────────────────────────────────────────────────────────────────


def _upload_raw_xlsx(
    drive, xlsx_path: Path, name: str, folder_id: str
) -> str:
    """Upload xlsx to Drive as-is, without conversion. Returns the file id."""
    from googleapiclient.http import MediaFileUpload

    file_metadata = {"name": name, "mimeType": XLSX_MIMETYPE}
    if folder_id:
        file_metadata["parents"] = [folder_id]

    media = MediaFileUpload(
        str(xlsx_path), mimetype=XLSX_MIMETYPE,
        resumable=True, chunksize=_RESUMABLE_CHUNK_SIZE,
    )
    request = drive.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, webViewLink",
        supportsAllDrives=True,
    )
    res = _execute_resumable(request)
    return res["id"]


# ────────────────────────────────────────────────────────────────────────
# Strategy 3 — Sheets API create + populate.
# ────────────────────────────────────────────────────────────────────────


def _delete_orphan_sheet(drive, sheet_id: str) -> None:
    """Best-effort cleanup of a half-created sheet so it doesn't leak."""
    try:
        drive.files().delete(
            fileId=sheet_id, supportsAllDrives=True
        ).execute()
        print(
            f"Cleaned up orphan sheet {sheet_id} after population failure",
            file=sys.stderr,
        )
    except Exception as cleanup_err:  # noqa: BLE001
        print(
            f"Warning: could not clean up orphan sheet {sheet_id}: "
            f"{type(cleanup_err).__name__}",
            file=sys.stderr,
        )


def _create_via_sheets_api(
    drive, sheets, csv_rows: list[list[str]], name: str, folder_id: str
) -> str:
    """Last-resort: create a new Sheet via Sheets API + populate.

    Returns the spreadsheetId. No openpyxl styling parity — this path
    exists for service accounts with zero Drive storage quota.

    If population fails after `spreadsheets.create` succeeds, the
    half-populated sheet is deleted before re-raising so we don't leak
    orphan files in Drive.
    """
    create_res = _retry_with_backoff(
        sheets.spreadsheets().create(
            body={
                "properties": {"title": name},
                "sheets": [
                    {"properties": {"title": "Results"}},
                    {"properties": {"title": "Legend"}},
                ],
            }
        ).execute
    )
    sheet_id = create_res["spreadsheetId"]
    results_sheet_id = create_res["sheets"][0]["properties"]["sheetId"]

    try:
        # Move to target folder.
        if folder_id:
            try:
                _retry_with_backoff(
                    drive.files().update(
                        fileId=sheet_id,
                        addParents=folder_id,
                        fields="id, parents",
                        supportsAllDrives=True,
                    ).execute
                )
            except Exception as err:  # noqa: BLE001
                print(f"Could not move to folder: {err}", file=sys.stderr)

        # Write Results data.
        _retry_with_backoff(
            sheets.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range="Results!A1",
                valueInputOption="RAW",
                body={"values": csv_rows},
            ).execute
        )

        # Write Legend data.
        _retry_with_backoff(
            sheets.spreadsheets().values().update(
                spreadsheetId=sheet_id,
                range="Legend!A1",
                valueInputOption="RAW",
                body={"values": LEGEND_DATA},
            ).execute
        )

        # Clamp freeze + auto-resize bounds to the actual data width so a
        # narrow frame doesn't trigger a 400 "frozenColumnCount > grid width".
        ncols = len(csv_rows[0]) if csv_rows and csv_rows[0] else 1
        frozen_cols = min(3, ncols)
        end_index = max(ncols, 1)

        # Apply basic formatting via batchUpdate.
        _retry_with_backoff(
            sheets.spreadsheets().batchUpdate(
                spreadsheetId=sheet_id,
                body={
                    "requests": [
                        # Freeze first row + first N (≤3) columns.
                        {
                            "updateSheetProperties": {
                                "properties": {
                                    "sheetId": results_sheet_id,
                                    "gridProperties": {
                                        "frozenRowCount": 1,
                                        "frozenColumnCount": frozen_cols,
                                    },
                                },
                                "fields": (
                                    "gridProperties.frozenRowCount,"
                                    "gridProperties.frozenColumnCount"
                                ),
                            }
                        },
                        # Bold header row.
                        {
                            "repeatCell": {
                                "range": {
                                    "sheetId": results_sheet_id,
                                    "startRowIndex": 0,
                                    "endRowIndex": 1,
                                },
                                "cell": {
                                    "userEnteredFormat": {
                                        "textFormat": {"bold": True}
                                    }
                                },
                                "fields": "userEnteredFormat.textFormat.bold",
                            }
                        },
                        # Auto-resize columns up to the actual data width.
                        {
                            "autoResizeDimensions": {
                                "dimensions": {
                                    "sheetId": results_sheet_id,
                                    "dimension": "COLUMNS",
                                    "startIndex": 0,
                                    "endIndex": end_index,
                                }
                            }
                        },
                    ]
                },
            ).execute
        )
    except Exception:
        _delete_orphan_sheet(drive, sheet_id)
        raise

    return sheet_id


# ────────────────────────────────────────────────────────────────────────
# Top-level orchestration.
# ────────────────────────────────────────────────────────────────────────


# Number of styling rows `build_xlsx.compute_workbook` writes BEFORE the
# column-header row (row 1: title, row 2: group headers). When falling back
# to the xlsx for Strategy 3 input, we must skip these so the resulting
# Sheet has a clean column-header-first shape.
_XLSX_STYLING_PRELUDE_ROWS = 2


def _csv_rows_from_csv(csv_path: Path) -> list[list[str]]:
    """Read a CSV at `csv_path` as rows-of-strings.

    Uses utf-8-sig to round-trip cleanly with `build_output`'s write
    encoding (matches the BOM-aware reads elsewhere in the pipeline).
    """
    df = pd.read_csv(
        csv_path, dtype=str, keep_default_na=False, encoding="utf-8-sig"
    )
    return csv_rows_for_upload(df)


def _csv_rows_from_xlsx(xlsx_path: Path) -> list[list[str]]:
    """Read the 'Results' sheet of the styled XLSX, skipping styling prelude.

    `build_xlsx.compute_workbook` writes a 3-row header (title at row 1,
    group headers at row 2, column headers at row 3) before the data. We
    drop the first 2 rows so the returned list starts with the column
    headers — matching the shape `_create_via_sheets_api` expects. Without
    this, Strategy 3 produces a Sheet where the title text sits in row 1
    and the actual headers are in row 3 unstyled.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
        rows: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < _XLSX_STYLING_PRELUDE_ROWS:
                continue
            rows.append(["" if v is None else str(v) for v in row])
        return rows
    finally:
        wb.close()


def _csv_rows_from_path(xlsx_path: Path) -> list[list[str]]:
    """Resolve csv_rows for the Strategy 3 Sheets-API fallback.

    Prefers a sibling CSV (same stem, `.csv` extension at the same
    directory, or under `working/`) since it has clean headers + data
    with no styling prelude — and matches what the legacy JS reads.
    Falls back to the xlsx with a 2-row skip if no CSV is co-located.

    Search order:
      1. <xlsx_path>.csv  — sibling CSV (build_output writes both base/
         and working/ but operators sometimes only retain one).
      2. <xlsx_dir>/working/<stem>.csv
      3. <xlsx_path> itself, with the styling prelude stripped.
    """
    sibling_csv = xlsx_path.with_suffix(".csv")
    working_csv = xlsx_path.parent / "working" / sibling_csv.name
    for csv_path in (sibling_csv, working_csv):
        if csv_path.exists():
            return _csv_rows_from_csv(csv_path)
    return _csv_rows_from_xlsx(xlsx_path)


def push_to_gsheets(
    xlsx_path: str | Path,
    niche: str,
    key_path: str | Path = DEFAULT_KEY_PATH,
    folder_id: str = DEFAULT_FOLDER_ID,
    id_file_path: str | Path | None = None,
    csv_rows: list[list[str]] | None = None,
) -> str:
    """Push the styled XLSX to Google Drive. Returns the Sheet/file id.

    Walks the three-strategy fallback chain. Writes the resulting id to
    `id_file_path` if provided (atomically — only after the new upload
    succeeds, so a failed run doesn't orphan the previous sheet). Raises
    FileNotFoundError if either the xlsx or the service-account key is
    missing. Raises PushFailedError if all three strategies fail.

    `csv_rows` is used for Strategy 3 (the Sheets API fallback). If not
    provided, the rows are read from the XLSX itself so Strategy 3 can
    always run.

    Title is truncated to 200 chars to stay within Sheets API limits.
    """
    xlsx_path = Path(xlsx_path)
    key_path = Path(key_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
    if not key_path.exists():
        raise FileNotFoundError(
            f"Service account key not found: {key_path}\n"
            "Setup: download the JSON key from Google Cloud Console and "
            "save to config/google-service-account.json (or pass --key)."
        )

    drive, sheets = _build_clients(key_path)

    # Read (but don't yet delete) the previous id. We defer deletion
    # until the new upload succeeds — this avoids the half-deleted state
    # where prev-sheet is gone but the new upload also failed.
    previous_id: str | None = None
    id_file = Path(id_file_path) if id_file_path else None
    if id_file and id_file.exists():
        previous_id = id_file.read_text(encoding="utf-8").strip() or None

    title = (
        f"{niche.replace('-', ' ').title()} Final Results — "
        f"{date.today().isoformat()}"
    )
    title = title[:200]

    new_id: str | None = None
    success_msg: str | None = None

    # Strategy 1: xlsx-conversion upload. Fall through on quota error.
    try:
        new_id = _upload_with_conversion(drive, xlsx_path, title, folder_id)
        success_msg = f"Uploaded (xlsx conversion): {title}"
    except Exception as err:  # noqa: BLE001
        if not _is_quota_error(err):
            raise
        print(
            f"Xlsx conversion upload hit quota error; falling back to raw xlsx: {err}",
            file=sys.stderr,
        )

    # Strategy 2: raw xlsx upload.
    if new_id is None:
        try:
            new_id = _upload_raw_xlsx(drive, xlsx_path, title, folder_id)
            success_msg = (
                f"Uploaded (raw xlsx, not converted): {title}\n"
                "Note: file is xlsx in Drive. Open in Google Sheets to view."
            )
        except Exception as err:  # noqa: BLE001
            print(
                f"Raw upload also failed: {type(err).__name__}: {err}",
                file=sys.stderr,
            )

    # Strategy 3: Sheets API create + populate. Render csv_rows from the
    # xlsx if not supplied so this path always has data — the legacy
    # default behaviour silently skipped Strategy 3 here.
    #
    # `_csv_rows_from_path` prefers a sibling CSV (built by step 4c.1)
    # since it has clean headers + data; falls back to reading the styled
    # xlsx with the title/group-header rows stripped.
    if new_id is None:
        if csv_rows is None:
            try:
                csv_rows = _csv_rows_from_path(xlsx_path)
            except Exception as err:  # noqa: BLE001
                print(
                    f"Could not resolve csv_rows for Strategy 3: "
                    f"{type(err).__name__}",
                    file=sys.stderr,
                )
                csv_rows = None
        if csv_rows is not None:
            try:
                new_id = _create_via_sheets_api(
                    drive, sheets, csv_rows, title, folder_id
                )
                success_msg = (
                    f"Uploaded (Sheets API fallback): {title}\n"
                    "Note: styling is basic. Free up Drive quota for full styling."
                )
            except Exception as err:  # noqa: BLE001
                print(
                    f"Sheets API fallback also failed: {type(err).__name__}",
                    file=sys.stderr,
                )

    if new_id is None:
        # All strategies exhausted. Leave previous_id alone — better to
        # keep the stale sheet visible than orphan everything.
        print("", file=sys.stderr)
        print("=== UPLOAD FAILED ===", file=sys.stderr)
        print(
            "All three upload strategies exhausted. The service account may have "
            "zero Drive storage quota (common for Cloud-only service accounts).",
            file=sys.stderr,
        )
        print(
            f"Manual fallback: drag {xlsx_path} into "
            f"https://drive.google.com/drive/folders/{folder_id}",
            file=sys.stderr,
        )
        raise PushFailedError("All Google Drive / Sheets upload strategies failed.")

    # New upload succeeded. Now safe to clean up the previous sheet and
    # persist the new id atomically (shared helper from steps._helpers).
    _delete_previous_sheet(drive, previous_id)
    if id_file:
        atomic_write(id_file, lambda p: p.write_text(new_id, encoding="utf-8"))
    if success_msg:
        print(success_msg)
    print(f"Sheet ID: {new_id}")
    return new_id


# ────────────────────────────────────────────────────────────────────────
# Step contract.
# ────────────────────────────────────────────────────────────────────────


def run_step(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Step-runner-compatible wrapper. Passthrough on the DataFrame.

    If `config["xlsx_path"]` is set, attempt an upload — otherwise no-op.
    Required `config` keys when uploading:
      - xlsx_path: path to the styled XLSX produced by step 4c.2
      - niche: niche slug (used in the sheet title)
      - key_path: optional path to the service-account JSON key

    The DataFrame is returned unchanged (this step is a side-effect step).
    """
    if not config.get("xlsx_path"):
        return df

    push_to_gsheets(
        xlsx_path=config["xlsx_path"],
        niche=config.get("niche", "results"),
        key_path=config.get("key_path", DEFAULT_KEY_PATH),
        folder_id=config.get("folder_id", DEFAULT_FOLDER_ID),
        id_file_path=config.get("id_file_path"),
        csv_rows=csv_rows_for_upload(df) if config.get("enable_sheets_api_fallback") else None,
    )
    return df


# ────────────────────────────────────────────────────────────────────────
# CLI — mirrors legacy push_to_gsheets.js paths.
# ────────────────────────────────────────────────────────────────────────


def run(niche: str, base: Path, key_path: Path, folder_id: str) -> None:
    base = Path(base)
    niche_snake = niche.replace("-", "_")

    xlsx_path = base / f"{niche_snake}_final_results.xlsx"
    id_file_path = base / f"{niche_snake}_gsheet_id.txt"
    csv_path = base / f"{niche_snake}_final_results.csv"
    if not csv_path.exists():
        csv_path = base / "working" / f"{niche_snake}_final_results.csv"

    csv_rows = None
    if csv_path.exists():
        df = pd.read_csv(
            csv_path, dtype=str, keep_default_na=False, encoding="utf-8-sig"
        )
        csv_rows = csv_rows_for_upload(df)

    sheet_id = push_to_gsheets(
        xlsx_path=xlsx_path,
        niche=niche,
        key_path=key_path,
        folder_id=folder_id,
        id_file_path=id_file_path,
        csv_rows=csv_rows,
    )
    print(f"URL: https://docs.google.com/spreadsheets/d/{sheet_id}/edit")
    print(f"ID saved to: {id_file_path}")


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Phase 5 Google Sheets push — uploads the styled XLSX produced "
            "by step 4c.2 to Drive with a 3-strategy fallback chain."
        )
    )
    parser.add_argument("--niche", required=True, help="Niche slug")
    parser.add_argument(
        "--base", required=True, type=Path,
        help="Base directory containing {niche}_final_results.xlsx",
    )
    parser.add_argument(
        "--key", type=Path, default=Path(DEFAULT_KEY_PATH),
        help=f"Service account JSON key path (default: {DEFAULT_KEY_PATH})",
    )
    parser.add_argument(
        "--folder", default=DEFAULT_FOLDER_ID,
        help=f"Drive folder ID (default: {DEFAULT_FOLDER_ID})",
    )
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = _parse_args(sys.argv[1:])
    run(niche=args.niche, base=args.base, key_path=args.key, folder_id=args.folder)
