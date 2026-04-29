"""Decision Engine step (formerly Phase 6 / Skill 6 in the legacy Keepa pipeline).

Appends 11 operator-decision columns to a Phase-5 final_results frame:

  Decision             BUY / NEGOTIATE / WATCH / KILL
  Decision Score       0..100 (rounded, JS Math.round semantics)
  Decision Reason      pipe-joined audit trail
  Joinability Status   Joinable / Review / Unsafe
  Buy Readiness        Ready / Cost Needed / Review Needed / Reject
  Max Buy Price        "GBPx.xx" (from Max Cost 20% ROI)
  Target Buy Price     "GBPx.xx" (lane-specific buffer applied)
  Cost Gap             "GBP+x.xx" / "GBP-x.xx" if supplier cost known
  Margin Status        Safe / Tight / Fail / Unknown
  Action Note          short operator instruction
  Shortlist Flag       Y for BUY/NEGOTIATE, N otherwise

Logic ported 1:1 from `fba_engine/_legacy_keepa/skills/skill-6-decision-engine/
phase6_decision.js`. The order of decision rules in the elif chain is
load-bearing — it is preserved verbatim. Output rows are sorted by Decision
Score DESC, then Monthly Gross Profit DESC, then ASIN ASC.

Standalone CLI invocation (mirrors the legacy JS contract):

    python -m fba_engine.steps.decision_engine \\
        --niche kids-toys \\
        --base fba_engine/data/niches/kids-toys
"""
from __future__ import annotations

import argparse
import math
import re
import sys
import warnings
from datetime import date
from pathlib import Path

import openpyxl
import openpyxl.styles
import openpyxl.utils
import pandas as pd

# ────────────────────────────────────────────────────────────────────────
# Constants — kept identical to the legacy JS so verdict counts match.
# ────────────────────────────────────────────────────────────────────────

DECISION_HEADERS = [
    "Decision",
    "Decision Score",
    "Decision Reason",
    "Joinability Status",
    "Buy Readiness",
    "Max Buy Price",
    "Target Buy Price",
    "Cost Gap",
    "Margin Status",
    "Action Note",
    "Shortlist Flag",
]

BUY_SCORE = 80
NEGOTIATE_SCORE = 60
WATCH_SCORE = 40
COMMERCIALLY_STRONG_SCORE = 68
SAFE_ENOUGH_SCORE = 60
NEGOTIATE_TOLERANCE = -2.0
IMPOSSIBLE_GAP = -4.0
TARGET_BUY_DISCOUNT = 0.9
TARGET_BUY_BUFFERS: dict[str, float] = {
    "BALANCED": 2.0,
    "CASH FLOW": 1.5,
    "PROFIT": 2.5,
    "default": 1.25,
}

SHORTLIST_COLUMNS: list[tuple[str, int]] = [
    ("ASIN", 14),
    ("Product Name", 42),
    ("Brand", 18),
    ("Opportunity Lane", 16),
    ("Monthly Gross Profit", 18),
    ("Est Profit", 12),
    ("IP Risk Band", 12),
    ("Private Label Risk", 16),
    ("Decision", 14),
    ("Decision Score", 14),
    ("Max Buy Price", 14),
    ("Target Buy Price", 16),
    ("Supplier Price", 14),
    ("Cost Gap", 14),
    ("Decision Reason", 54),
    ("Action Note", 24),
]

# ────────────────────────────────────────────────────────────────────────
# Coercion helpers — pure, NaN-safe.
# ────────────────────────────────────────────────────────────────────────


def _coerce_str(raw: object) -> str:
    """Coerce a cell value to a clean string. None and pandas NaN -> ""."""
    if raw is None:
        return ""
    if isinstance(raw, float) and math.isnan(raw):
        return ""
    return str(raw).strip()


def _to_num(value: object) -> float:
    """JS-style Number() coercion. NaN/None/empty/garbage -> 0.

    Used for the score-ladder helpers where the JS code relies on
    `Number(x) || fallback` to demote bad input to a sentinel.
    """
    if value is None:
        return 0.0
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    try:
        n = float(value)
    except (TypeError, ValueError):
        return 0.0
    if math.isnan(n):
        return 0.0
    return n


def parse_money(value: object) -> float:
    """Mirror the JS parseMoney: strip GBP and any non-[0-9.-] chars, parse, NaN->0.

    Used for currency-like cells that may carry "GBP10.50", "£5", or even
    "10.5" depending on the upstream phase.
    """
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    s = str(value or "")
    s = re.sub(r"GBP", "", s, flags=re.IGNORECASE)
    s = re.sub(r"[^0-9.\-]", "", s).strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_pct(value: object) -> float:
    """Mirror the JS parsePct: strip %, parse, NaN->0."""
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    s = str(value or "")
    s = s.replace("%", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def gbp(value: object) -> str:
    """Format as `GBP{value:.2f}`. Empty/None/NaN/non-numeric -> ""."""
    if value == "" or value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    try:
        n = float(value)
    except (TypeError, ValueError):
        return ""
    if math.isnan(n):
        return ""
    return f"GBP{n:.2f}"


def as_upper(value: object) -> str:
    """Trim and uppercase. None/NaN -> ""."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip().upper()


def is_truthy_y(value: object) -> bool:
    """Y / YES / TRUE (case-insensitive) -> True."""
    return as_upper(value) in {"Y", "YES", "TRUE"}


def clamp(value: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, value))


# ────────────────────────────────────────────────────────────────────────
# Categorical scoring helpers.
# ────────────────────────────────────────────────────────────────────────


def stable_state(value: object) -> str:
    """Bucket a free-text price-stability description into GOOD/CAUTION/BAD/UNKNOWN.

    The legacy JS uses substring containment after upper-casing — preserved
    verbatim here so phrases like "STABLE up" still classify as GOOD.
    """
    v = as_upper(value)
    if "STABLE" in v or "RISING" in v:
        return "GOOD"
    if "SLIGHT DIP" in v:
        return "CAUTION"
    if "DROPPING" in v or "SURGING" in v or "COMPRESSED" in v:
        return "BAD"
    return "UNKNOWN"


def lane_base_score(lane: object) -> float:
    v = as_upper(lane)
    if v == "BALANCED":
        return 92
    if v == "CASH FLOW":
        return 84
    if v == "PROFIT":
        return 78
    return 42


def score_from_priority(priority: object) -> float:
    """Higher commercial priority -> lower score. Falsy/garbage -> 8 -> floored at 40."""
    p = _to_num(priority) or 8  # JS `Number(priority) || 8`
    return clamp(100 - (max(1, p) - 1) * 12, 40, 100)


def score_from_monthly_gross(value: object) -> float:
    return clamp(_to_num(value) / 6, 0, 100)


def score_from_bought(value: object) -> float:
    return clamp(_to_num(value) * 1.25, 0, 100)


def score_from_unit_profit(value: object) -> float:
    return clamp(_to_num(value) * 12, 0, 100)


def score_from_roi(value: object) -> float:
    return clamp(_to_num(value) * 2.2, 0, 100)


def risk_band_score(band: object) -> float:
    v = as_upper(band)
    if v == "LOW":
        return 92
    if v == "MEDIUM":
        return 55
    if v == "HIGH":
        return 8
    return 50


def pl_risk_score(risk: object) -> float:
    v = as_upper(risk)
    if v in {"UNLIKELY", "LOW"}:
        return 88
    if v in {"LIKELY", "HIGH"}:
        return 22
    return 55


def route_score(route_code: object) -> float:
    v = as_upper(route_code)
    if not v or v == "UNCLEAR":
        return 30
    return 78


def get_target_buffer(lane: object) -> float:
    return TARGET_BUY_BUFFERS.get(as_upper(lane), TARGET_BUY_BUFFERS["default"])


def calc_target_buy_price(max_buy_price: float, lane: object) -> float | str:
    """Lower of (90% discount, max minus lane buffer). 0/negative max -> ""."""
    if not max_buy_price or max_buy_price <= 0:
        return ""
    discount = max_buy_price * TARGET_BUY_DISCOUNT
    buffered = max_buy_price - get_target_buffer(lane)
    return max(0, min(discount, buffered))


# ────────────────────────────────────────────────────────────────────────
# Decision-logic helpers.
# ────────────────────────────────────────────────────────────────────────


def calc_joinability(row: dict) -> str:
    """Classify listing joinability. Returns Joinable / Review / Unsafe."""
    ip_risk = as_upper(row.get("IP Risk Band"))
    fortress = as_upper(row.get("Fortress Listing"))
    brand_seller = as_upper(row.get("Brand Seller Match"))
    brand_type_v = as_upper(row.get("Brand Type"))
    brand_store = as_upper(row.get("Brand Store Present"))
    pl_risk = as_upper(row.get("Private Label Risk"))
    gated = is_truthy_y(row.get("Gated"))

    if ip_risk == "HIGH":
        return "Unsafe"
    if fortress == "YES" and brand_seller in {"YES", "PARTIAL"}:
        return "Unsafe"
    if (
        brand_type_v == "ESTABLISHED"
        and brand_seller == "YES"
        and brand_store == "LIKELY"
    ):
        return "Unsafe"
    if (
        ip_risk == "MEDIUM"
        or pl_risk == "LIKELY"
        or gated
        or fortress == "YES"
    ):
        return "Review"
    return "Joinable"


def calc_margin_status(
    actual_supplier_price: float,
    target_buy_price: float | str,
    max_buy_price: float,
    est_profit: float,
    est_roi: float,
) -> str:
    """Classify margin viability. Returns Safe / Tight / Fail / Unknown."""
    if actual_supplier_price > 0:
        if target_buy_price != "" and actual_supplier_price <= target_buy_price:
            return "Safe"
        if max_buy_price > 0 and actual_supplier_price <= max_buy_price:
            return "Tight"
        return "Fail"
    if max_buy_price <= 0:
        return "Unknown"
    if est_profit < 0.75 or est_roi < 10:
        return "Fail"
    return "Unknown"


def calc_buy_readiness(
    joinability: str,
    margin_status: str,
    has_supplier_cost: bool,
    commercially_strong: bool,
    gated: bool,
) -> str:
    """Classify buy readiness. Returns Ready / Cost Needed / Review Needed / Reject."""
    if joinability == "Unsafe" or margin_status == "Fail":
        return "Reject"
    if not has_supplier_cost:
        return "Cost Needed" if commercially_strong else "Review Needed"
    if joinability == "Review" or gated:
        return "Review Needed"
    return "Ready"


def calc_action_note(
    decision: str,
    buy_readiness: str,
    stability_state_value: str,
    has_supplier_cost: bool,
) -> str:
    """Operator-friendly next-step instruction for a row."""
    if decision == "BUY":
        return "Place opening order"
    if decision == "NEGOTIATE" and not has_supplier_cost:
        return "Contact supplier for cost"
    if decision == "NEGOTIATE":
        return "Negotiate lower trade price"
    if decision == "WATCH" and buy_readiness == "Review Needed":
        return "Review listing safety"
    if decision == "WATCH" and stability_state_value == "BAD":
        return "Monitor for 7 days"
    if decision == "KILL" and buy_readiness in {"Reject", "Review Needed"}:
        return "Avoid listing"
    return "Review manually"


def build_decision_reason(
    decision: str,
    lane: object,
    ip_risk_band: object,
    pl_risk: object,
    price_stability: object,
    monthly_gross_profit: float,
    extra: object,
) -> str:
    """Pipe-joined audit string. Mirrors legacy formatting byte-for-byte."""
    parts = [
        decision,
        str(lane) if lane else "UNASSIGNED",
        f"{gbp(monthly_gross_profit)}/mo",
        f"{str(ip_risk_band or 'Unknown').lower()} IP risk",
        f"{str(pl_risk or 'Unknown').lower()} PL risk",
        str(price_stability or "unknown").lower(),
    ]
    if extra:
        parts.append(str(extra))
    return " | ".join(parts)


# ────────────────────────────────────────────────────────────────────────
# Per-row scoring + decision (the heart of the step).
# ────────────────────────────────────────────────────────────────────────


def _score_and_decide(row: dict) -> dict[str, object]:
    """Apply Phase 6 scoring + decision rules to one row dict.

    Returns the 11 decision-layer fields. Pure: never reads or writes I/O.
    """
    lane = _coerce_str(row.get("Opportunity Lane"))
    monthly_gross_profit = parse_money(row.get("Monthly Gross Profit"))
    bought_per_month = parse_money(row.get("Bought per Month"))
    est_profit = parse_money(row.get("Est Profit"))
    est_roi = parse_pct(row.get("Est ROI %"))
    real_roi = parse_pct(row.get("Real ROI %"))
    priority = parse_money(row.get("Commercial Priority"))
    max_buy_price = parse_money(row.get("Max Cost 20% ROI"))
    trade_price = parse_money(row.get("Trade Price"))
    trade_price_found = is_truthy_y(row.get("Trade Price Found"))
    ip_risk_band = _coerce_str(row.get("IP Risk Band"))
    pl_risk = _coerce_str(row.get("Private Label Risk"))
    price_stability = _coerce_str(row.get("Price Stability"))
    route_code = _coerce_str(row.get("Route Code"))
    gated = is_truthy_y(row.get("Gated"))

    has_supplier_cost = trade_price_found and trade_price > 0
    target_buy_price = calc_target_buy_price(max_buy_price, lane)
    cost_gap = max_buy_price - trade_price if has_supplier_cost else ""
    stab_state = stable_state(price_stability)

    # Joinability uses a normalised view of the row.
    join_input = {
        "IP Risk Band": ip_risk_band,
        "Fortress Listing": _coerce_str(row.get("Fortress Listing")),
        "Brand Seller Match": _coerce_str(row.get("Brand Seller Match")),
        "Brand Type": _coerce_str(row.get("Brand Type")),
        "Brand Store Present": _coerce_str(row.get("Brand Store Present")),
        "Private Label Risk": pl_risk,
        "Gated": "Y" if gated else "N",
    }
    joinability = calc_joinability(join_input)

    commercial_score = clamp(
        lane_base_score(lane) * 0.35
        + score_from_monthly_gross(monthly_gross_profit) * 0.25
        + score_from_bought(bought_per_month) * 0.20
        + score_from_unit_profit(est_profit) * 0.10
        + score_from_priority(priority) * 0.10,
        0,
        100,
    )
    commercially_strong = commercial_score >= COMMERCIALLY_STRONG_SCORE

    margin_status = calc_margin_status(
        trade_price, target_buy_price, max_buy_price, est_profit, est_roi
    )
    buy_readiness = calc_buy_readiness(
        joinability, margin_status, has_supplier_cost, commercially_strong, gated
    )

    if has_supplier_cost:
        feasibility_base = clamp(((cost_gap + 5) / 10) * 100, 0, 100)
    else:
        feasibility_base = 52 if commercially_strong else 36

    if margin_status == "Safe":
        feasibility_base += 10
    elif margin_status == "Tight":
        feasibility_base += 2

    if buy_readiness == "Ready":
        readiness_score = 95
    elif buy_readiness == "Cost Needed":
        readiness_score = 55
    elif buy_readiness == "Review Needed":
        readiness_score = 45
    else:
        readiness_score = 10

    feasibility_score = clamp(
        feasibility_base * 0.55
        + route_score(route_code) * 0.20
        + readiness_score * 0.25,
        0,
        100,
    )

    if joinability == "Joinable":
        join_score = 95
    elif joinability == "Review":
        join_score = 55
    else:
        join_score = 5
    safety_score = clamp(
        risk_band_score(ip_risk_band) * 0.45
        + pl_risk_score(pl_risk) * 0.25
        + join_score * 0.30,
        0,
        100,
    )

    if margin_status == "Safe":
        margin_safety_base = 95
    elif margin_status == "Tight":
        margin_safety_base = 60
    elif margin_status == "Unknown":
        margin_safety_base = 50
    else:
        margin_safety_base = 5
    margin_safety_score = clamp(
        margin_safety_base * 0.45
        + score_from_roi(real_roi or est_roi) * 0.30
        + score_from_unit_profit(est_profit) * 0.25,
        0,
        100,
    )

    safe_enough = safety_score >= SAFE_ENOUGH_SCORE
    decision_score = clamp(
        commercial_score * 0.35
        + feasibility_score * 0.25
        + safety_score * 0.25
        + margin_safety_score * 0.15,
        0,
        100,
    )

    ip_risk_upper = as_upper(ip_risk_band)
    pl_risk_upper = as_upper(pl_risk)
    lane_upper = as_upper(lane)

    # Order is load-bearing — preserved verbatim from the legacy JS.
    if (
        ip_risk_upper == "HIGH"
        or joinability == "Unsafe"
        or margin_status == "Fail"
        or (has_supplier_cost and cost_gap < IMPOSSIBLE_GAP)
    ):
        decision = "KILL"
        decision_score = min(decision_score, 35)
    elif (
        commercially_strong
        and safe_enough
        and has_supplier_cost
        and margin_status == "Safe"
        and joinability == "Joinable"
        and ip_risk_upper == "LOW"
        and pl_risk_upper != "LIKELY"
    ):
        decision = "BUY"
        decision_score = max(decision_score, BUY_SCORE)
    elif (
        (ip_risk_upper == "MEDIUM" and not has_supplier_cost)
        or (pl_risk_upper == "-" and joinability != "Joinable")
        or stab_state == "BAD"
    ):
        decision = "WATCH"
        decision_score = min(
            max(decision_score, WATCH_SCORE), NEGOTIATE_SCORE - 1
        )
    elif (
        commercially_strong
        and safe_enough
        and (
            not has_supplier_cost
            or margin_status == "Tight"
            or (
                has_supplier_cost
                and cost_gap < 0
                and cost_gap >= NEGOTIATE_TOLERANCE
            )
        )
    ):
        decision = "NEGOTIATE"
        decision_score = max(
            min(decision_score, BUY_SCORE - 1), NEGOTIATE_SCORE
        )
    elif decision_score >= BUY_SCORE:
        decision = "BUY" if has_supplier_cost else "NEGOTIATE"
    elif decision_score >= NEGOTIATE_SCORE:
        decision = "NEGOTIATE"
    elif decision_score >= WATCH_SCORE:
        decision = "WATCH"
    else:
        # Defensive fallthrough — unreachable in practice for any input that
        # avoided the KILL override (safety_score floors near 80 once IP isn't
        # HIGH and joinability isn't Unsafe, keeping decision_score above 40).
        decision = "KILL"

    # Post-checks: BUY can be demoted but never promoted at this stage.
    if decision == "BUY" and buy_readiness != "Ready":
        decision = "NEGOTIATE" if has_supplier_cost else "WATCH"

    if decision == "BUY" and lane_upper not in {"BALANCED", "CASH FLOW"}:
        decision = "NEGOTIATE"

    if has_supplier_cost:
        sign = "+" if cost_gap >= 0 else ""
        extra_reason = f"cost gap {sign}{gbp(cost_gap)}"
    else:
        extra_reason = "supplier cost missing"
    decision_reason = build_decision_reason(
        decision, lane, ip_risk_band, pl_risk, price_stability,
        monthly_gross_profit, extra_reason,
    )

    action_note = calc_action_note(
        decision, buy_readiness, stab_state, has_supplier_cost
    )
    shortlist_flag = "Y" if decision in {"BUY", "NEGOTIATE"} else "N"

    # Commercial rounding (JS Math.round semantics) — Python's built-in round
    # uses banker's rounding, so 0.5 -> 0 instead of 1. Score is non-negative
    # in [0, 100], so floor(x + 0.5) matches Math.round.
    decision_score_int = int(math.floor(decision_score + 0.5))

    return {
        "Decision": decision,
        "Decision Score": str(decision_score_int),
        "Decision Reason": decision_reason,
        "Joinability Status": joinability,
        "Buy Readiness": buy_readiness,
        "Max Buy Price": gbp(max_buy_price) if max_buy_price > 0 else "",
        "Target Buy Price": gbp(target_buy_price) if target_buy_price != "" else "",
        "Cost Gap": gbp(cost_gap) if has_supplier_cost else "",
        "Margin Status": margin_status,
        "Action Note": action_note,
        "Shortlist Flag": shortlist_flag,
    }


# ────────────────────────────────────────────────────────────────────────
# DataFrame entry point.
# ────────────────────────────────────────────────────────────────────────


# Columns the scoring genuinely depends on. Missing any of these means rows
# get default zeros/empties — warn so a misshapen upstream isn't silent.
_REQUIRED_INPUT_COLUMNS = (
    "Opportunity Lane",
    "Monthly Gross Profit",
    "IP Risk Band",
    "Trade Price",
    "Trade Price Found",
)


def compute_decisions(df: pd.DataFrame) -> pd.DataFrame:
    """Append the 11 decision-layer columns to a Phase-5 final_results frame.

    Pure: does not read or write disk, does not mutate the input. Sort order
    matches the legacy JS (Decision Score DESC, Monthly Gross Profit DESC,
    ASIN ASC).
    """
    if df.empty:
        return df.assign(
            **{header: pd.Series(dtype=object) for header in DECISION_HEADERS}
        )

    missing = [c for c in _REQUIRED_INPUT_COLUMNS if c not in df.columns]
    if missing:
        warnings.warn(
            f"compute_decisions: missing input columns {missing}; rows will be "
            f"scored using defaults (0/empty), which can produce misleading "
            f"verdicts.",
            stacklevel=2,
        )

    out = df.copy()
    rows = [_score_and_decide(row.to_dict()) for _, row in out.iterrows()]
    enriched = pd.DataFrame(rows, index=out.index)
    combined = pd.concat([out, enriched], axis=1)

    # Sort: Decision Score DESC, Monthly Gross Profit DESC, ASIN ASC.
    sort_keys: list[str] = []
    sort_asc: list[bool] = []

    # Decision Score is always populated by _score_and_decide (stringified int),
    # so a simple float cast is sufficient.
    combined["_sort_score"] = combined["Decision Score"].astype(float)
    sort_keys.append("_sort_score")
    sort_asc.append(False)

    if "Monthly Gross Profit" in combined.columns:
        combined["_sort_mgp"] = combined["Monthly Gross Profit"].apply(parse_money)
        sort_keys.append("_sort_mgp")
        sort_asc.append(False)

    if "ASIN" in combined.columns:
        combined["_sort_asin"] = combined["ASIN"].apply(_coerce_str)
        sort_keys.append("_sort_asin")
        sort_asc.append(True)

    combined = combined.sort_values(
        sort_keys, ascending=sort_asc, kind="mergesort"
    ).reset_index(drop=True)
    combined = combined.drop(columns=[k for k in sort_keys if k.startswith("_sort_")])

    return combined


# ────────────────────────────────────────────────────────────────────────
# Stats + handoff text.
# ────────────────────────────────────────────────────────────────────────


def _count_by(df: pd.DataFrame, key: str) -> dict[str, int]:
    """Count rows by `key` value. Empty/NaN -> "-" bucket. Insertion-ordered."""
    if key not in df.columns:
        return {}
    out: dict[str, int] = {}
    for raw in df[key]:
        v = _coerce_str(raw)
        if not v:
            v = "-"
        out[v] = out.get(v, 0) + 1
    return out


def build_stats(df: pd.DataFrame, niche: str) -> str:
    decision_counts = {
        v: int((df["Decision"] == v).sum())
        for v in ["BUY", "NEGOTIATE", "WATCH", "KILL"]
    }
    lane_counts = _count_by(df, "Opportunity Lane")
    ip_counts = _count_by(df, "IP Risk Band")
    pl_counts = _count_by(df, "Private Label Risk")
    shortlist_count = int((df["Shortlist Flag"] == "Y").sum())

    df_sorted = df.copy()
    df_sorted["_score_num"] = df_sorted["Decision Score"].apply(
        lambda x: float(x) if x and str(x).strip() else 0.0
    )
    top15 = df_sorted.sort_values(
        "_score_num", ascending=False, kind="mergesort"
    ).head(15)
    top15_lines = [
        f"  {i + 1}. {row.get('ASIN', '')} | {row.get('Decision', '')} | "
        f"{row.get('Decision Score', '')} | {row.get('Decision Reason', '')}"
        for i, (_, row) in enumerate(top15.iterrows())
    ]

    lines = [
        f"Niche: {niche}",
        f"Date: {date.today().isoformat()}",
        f"Input: {len(df)} products from Phase 5 final results",
        "",
        "Decision distribution:",
        f"  BUY: {decision_counts['BUY']}",
        f"  NEGOTIATE: {decision_counts['NEGOTIATE']}",
        f"  WATCH: {decision_counts['WATCH']}",
        f"  KILL: {decision_counts['KILL']}",
        "",
        "Opportunity Lane:",
        *(f"  {k}: {v}" for k, v in lane_counts.items()),
        "",
        "IP Risk Band:",
        *(f"  {k}: {v}" for k, v in ip_counts.items()),
        "",
        "Private Label Risk:",
        *(f"  {k}: {v}" for k, v in pl_counts.items()),
        "",
        f"Shortlist rows: {shortlist_count}",
        "",
        "Top 15 by Decision Score:",
        *top15_lines,
    ]
    return "\n".join(lines) + "\n"


def build_handoff(
    df: pd.DataFrame,
    niche: str,
    decision_csv_path: str,
    stats_path: str,
    handoff_path: str,
    shortlist_xlsx_path: str,
) -> str:
    decision_counts = {
        v: int((df["Decision"] == v).sum())
        for v in ["BUY", "NEGOTIATE", "WATCH", "KILL"]
    }
    return (
        f"# Phase 6 Handoff -- {niche}\n\n"
        f"Generated: {date.today().isoformat()}\n\n"
        f"Outputs:\n"
        f"- {decision_csv_path}\n"
        f"- {stats_path}\n"
        f"- {handoff_path}\n"
        f"- {shortlist_xlsx_path}\n\n"
        f"Decision counts:\n"
        f"- BUY: {decision_counts['BUY']}\n"
        f"- NEGOTIATE: {decision_counts['NEGOTIATE']}\n"
        f"- WATCH: {decision_counts['WATCH']}\n"
        f"- KILL: {decision_counts['KILL']}\n\n"
        f"Next step:\n"
        f"- Review the shortlist workbook first, then use the full decision CSV "
        f"for audit detail.\n"
    )


# ────────────────────────────────────────────────────────────────────────
# Shortlist XLSX workbook.
# ────────────────────────────────────────────────────────────────────────


def _set_bold(cell: openpyxl.cell.cell.Cell) -> None:
    cell.font = openpyxl.styles.Font(bold=True)


def build_shortlist_xlsx(df: pd.DataFrame, output_path: str | Path) -> None:
    """Write a two-sheet workbook: Shortlist (BUY/NEGOTIATE rows) + Summary (counts)."""
    wb = openpyxl.Workbook()

    shortlist_ws = wb.active
    shortlist_ws.title = "Shortlist"

    headers = [name for name, _ in SHORTLIST_COLUMNS]
    shortlist_ws.append(headers)
    for col_idx, (_, width) in enumerate(SHORTLIST_COLUMNS, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        shortlist_ws.column_dimensions[col_letter].width = width
        _set_bold(shortlist_ws.cell(row=1, column=col_idx))
    shortlist_ws.freeze_panes = "A2"

    if "Shortlist Flag" in df.columns:
        shortlist_df = df[df["Shortlist Flag"] == "Y"]
    else:
        shortlist_df = df.iloc[0:0]

    for _, row in shortlist_df.iterrows():
        cost_gap = row.get("Cost Gap", "") or ""
        has_supplier_cost = bool(cost_gap)
        supplier_price_raw = parse_money(row.get("Trade Price", ""))
        supplier_price = (
            gbp(supplier_price_raw)
            if has_supplier_cost and supplier_price_raw > 0
            else ""
        )

        # Re-format Monthly Gross Profit and Est Profit as GBP if numeric;
        # fall back to the raw cell otherwise so non-numeric strings survive.
        mgp_raw = parse_money(row.get("Monthly Gross Profit", ""))
        mgp_display = gbp(mgp_raw) if mgp_raw else _coerce_str(
            row.get("Monthly Gross Profit", "")
        )
        est_profit_raw = parse_money(row.get("Est Profit", ""))
        est_profit_display = gbp(est_profit_raw) if est_profit_raw else _coerce_str(
            row.get("Est Profit", "")
        )

        score_str = _coerce_str(row.get("Decision Score", ""))
        try:
            score_int = int(score_str) if score_str else 0
        except ValueError:
            score_int = 0

        shortlist_ws.append(
            [
                _coerce_str(row.get("ASIN", "")),
                _coerce_str(row.get("Product Name", "")),
                _coerce_str(row.get("Brand", "")),
                _coerce_str(row.get("Opportunity Lane", "")),
                mgp_display,
                est_profit_display,
                _coerce_str(row.get("IP Risk Band", "")),
                _coerce_str(row.get("Private Label Risk", "")),
                _coerce_str(row.get("Decision", "")),
                score_int,
                _coerce_str(row.get("Max Buy Price", "")),
                _coerce_str(row.get("Target Buy Price", "")),
                supplier_price,
                cost_gap,
                _coerce_str(row.get("Decision Reason", "")),
                _coerce_str(row.get("Action Note", "")),
            ]
        )

    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions["A"].width = 24
    summary_ws.column_dimensions["B"].width = 12

    cursor = 1
    for title, key in [
        ("Decision", "Decision"),
        ("Opportunity Lane", "Opportunity Lane"),
        ("IP Risk Band", "IP Risk Band"),
        ("Private Label Risk", "Private Label Risk"),
    ]:
        title_cell = summary_ws.cell(row=cursor, column=1, value=title)
        _set_bold(title_cell)
        cursor += 1

        _set_bold(summary_ws.cell(row=cursor, column=1, value="Value"))
        _set_bold(summary_ws.cell(row=cursor, column=2, value="Count"))
        cursor += 1

        for k, v in _count_by(df, key).items():
            summary_ws.cell(row=cursor, column=1, value=k)
            summary_ws.cell(row=cursor, column=2, value=v)
            cursor += 1

        cursor += 1  # blank row between blocks

    wb.save(output_path)


# ────────────────────────────────────────────────────────────────────────
# Step contract.
# ────────────────────────────────────────────────────────────────────────


def run_step(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Step-runner-compatible wrapper.

    The decision step does not require any keys from `config` — it operates
    on the columns already present in the DataFrame. The signature is
    preserved so the step 5 YAML runner can invoke any step uniformly.
    """
    return compute_decisions(df)


# ────────────────────────────────────────────────────────────────────────
# CLI — mirrors legacy phase6_decision.js paths.
# ────────────────────────────────────────────────────────────────────────


def run(niche: str, base: Path) -> None:
    """End-to-end: read Phase 5 final_results, write Phase 6 outputs."""
    base = Path(base)
    working = base / "working"
    niche_snake = niche.replace("-", "_")

    # Match JS: prefer base/{niche}_final_results.csv, fallback to working/.
    primary = base / f"{niche_snake}_final_results.csv"
    fallback = working / f"{niche_snake}_final_results.csv"
    if primary.exists():
        input_path = primary
    elif fallback.exists():
        input_path = fallback
    else:
        print(
            f"Phase 5 final results CSV not found: {primary} or {fallback}",
            file=sys.stderr,
        )
        sys.exit(1)

    # Ensure working/ exists — legacy JS would crash here on a fresh niche dir.
    working.mkdir(parents=True, exist_ok=True)

    output_path = working / f"{niche_snake}_phase6_decisions.csv"
    stats_path = working / f"{niche_snake}_phase6_stats.txt"
    handoff_path = working / f"{niche_snake}_phase6_handoff.md"
    shortlist_xlsx_path = base / f"{niche_snake}_phase6_shortlist.xlsx"

    # utf-8-sig strips a leading BOM if present (legacy JS did the same).
    df = pd.read_csv(
        input_path, dtype=str, keep_default_na=False, encoding="utf-8-sig"
    )
    enriched = compute_decisions(df)
    enriched.to_csv(output_path, index=False)
    stats_path.write_text(build_stats(enriched, niche), encoding="utf-8")
    handoff_path.write_text(
        build_handoff(
            enriched,
            niche,
            str(output_path),
            str(stats_path),
            str(handoff_path),
            str(shortlist_xlsx_path),
        ),
        encoding="utf-8",
    )
    build_shortlist_xlsx(enriched, shortlist_xlsx_path)
    print(f"Saved: {output_path}")
    print(f"Saved: {stats_path}")
    print(f"Saved: {handoff_path}")
    print(f"Saved: {shortlist_xlsx_path}")


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Decision Engine step (Phase 6) — produces "
            "BUY/NEGOTIATE/WATCH/KILL verdicts on Phase 5 output."
        )
    )
    parser.add_argument(
        "--niche", required=True, help="Niche slug (e.g. kids-toys, sports-goods)"
    )
    parser.add_argument(
        "--base",
        required=True,
        type=Path,
        help="Base directory containing Phase 5 final_results.csv (and a working/ subfolder).",
    )
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = _parse_args(sys.argv[1:])
    run(niche=args.niche, base=args.base)
