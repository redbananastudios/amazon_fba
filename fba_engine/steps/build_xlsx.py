"""Build XLSX step (Phase 5 — formerly Skill 5 part 2 in the legacy Keepa pipeline).

Renders a styled Excel workbook from a 67-column `final_results` DataFrame.

Logic ported from `fba_engine/_legacy_keepa/skills/skill-5-build-output/
build_final_xlsx.js` (529 LOC). The JS uses ExcelJS; this port uses openpyxl.

**Deliberate deviation from the JS:** the legacy script hard-coded styling
to 64 columns (autofilter, title merge, COL_WIDTHS, GROUPS) but the upstream
`final_results.csv` has had 67 columns since step 4c.1 (the 3 product-code
columns EAN/UPC/GTIN were added without updating the XLSX builder). The
Python port extends styling to 67 columns and adds a "Product Codes" group.

Standalone CLI invocation:

    python -m fba_engine.steps.build_xlsx \\
        --niche kids-toys \\
        --base fba_engine/data/niches/kids-toys
"""
from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path

import openpyxl
import openpyxl.styles
import openpyxl.utils
import pandas as pd

# ────────────────────────────────────────────────────────────────────────
# Style constants — ARGB hex strings, openpyxl-compatible.
# ────────────────────────────────────────────────────────────────────────

COLOURS: dict[str, str] = {
    "header_bg":     "FF1B2A4A",  # dark navy
    "header_font":   "FFFFFFFF",
    "yes_green":     "FFD5F5E3",
    "maybe_yellow":  "FFFEF9E7",
    "no_red":        "FFFADBD8",
    "brand_blue":    "FFD6EAF8",
    "dip_orange":    "FFFDEBD0",
    "erosion_red":   "FFF5B7B1",
    "gated_purple":  "FFE8DAEF",
    "zebra_even":    "FFF8F9FA",
    "border_grey":   "FFD5D8DC",
    "green_text":    "FF27AE60",
    "red_text":      "FFC0392B",
    "link_blue":     "FF2980B9",
    "score_bg_high": "FF82E0AA",
    "score_bg_mid":  "FFF9E79F",
    "score_bg_low":  "FFF1948A",
}

# Group headers for merged cells at row 2. Extended to 67 cols with a
# "Product Codes" group (vs the legacy 64-col stop).
GROUPS: list[dict] = [
    {"label": "Product",                     "start": 1,  "end": 6},
    {"label": "Verdict & Scores",            "start": 7,  "end": 22},
    {"label": "Pricing & Margins",           "start": 23, "end": 33},
    {"label": "Demand, Competition & Risk",  "start": 34, "end": 54},
    {"label": "Supplier & Sourcing",         "start": 55, "end": 64},
    {"label": "Product Codes",               "start": 65, "end": 67},
]

# Column widths (1-based). Cols 65-67 added by the port.
COL_WIDTHS: dict[int, int] = {
    1: 14,   # ASIN
    2: 45,   # Product Name
    3: 18,   # Brand
    4: 36,   # Amazon URL
    5: 14,   # Category
    6: 12,   # Weight Flag
    7: 16,   # Verdict
    8: 40,   # Verdict Reason
    9: 16,   # Opportunity Lane
    10: 8,   # Commercial Priority
    11: 40,  # Lane Reason
    12: 10,  # Composite
    13: 8,   # Demand
    14: 8,   # Stability
    15: 8,   # Competition
    16: 8,   # Margin
    17: 8,   # Cash Flow Score
    18: 8,   # Profit Score
    19: 8,   # Balanced Score
    20: 14,  # Monthly Gross Profit
    21: 14,  # Price Compression
    22: 14,  # Listing Quality
    23: 12,  # Current Price
    24: 12,  # BB 90d avg
    25: 14,  # Price Stability
    26: 12,  # Fulfilment Fee
    27: 12,  # Amazon Fees
    28: 14,  # Total Amazon Fees
    29: 12,  # Est Cost
    30: 10,  # Est Profit
    31: 10,  # Est ROI%
    32: 14,  # Max Cost
    33: 12,  # Breakeven
    34: 10,  # BSR
    35: 10,  # BSR Drops
    36: 12,  # Bought/mo
    37: 10,  # Star Rating
    38: 12,  # Review Count
    39: 8,   # Brand 1P
    40: 10,  # Sellers
    41: 10,  # Amazon
    42: 12,  # BB Share
    43: 10,  # PL Risk
    44: 14,  # Brand Seller Match
    45: 14,  # Fortress Listing
    46: 14,  # Brand Type
    47: 14,  # A+ Content Present
    48: 14,  # Brand Store Present
    49: 14,  # Category Risk Level
    50: 10,  # IP Risk Score
    51: 10,  # IP Risk Band
    52: 36,  # IP Reason
    53: 8,   # Gated
    54: 14,  # SAS Flags
    55: 14,  # Route
    56: 22,  # Supplier
    57: 22,  # Website
    58: 22,  # Contact
    59: 12,  # MOQ
    60: 10,  # Trade Found
    61: 12,  # Trade Price
    62: 10,  # Real ROI
    63: 30,  # Notes
    64: 18,  # Outreach
    # Extended by the Python port:
    65: 16,  # EAN
    66: 16,  # UPC
    67: 16,  # GTIN
}

# Numeric columns (1-based) — formatted as numbers. Three deliberate fixes
# vs the legacy JS, all of the same class (numeric col missing from
# NUMERIC_COLS so its number_format never applies):
#   - col 61 (Trade Price): legacy was in GBP_COLS only, so the GBP format
#     never fired. Added to NUMERIC_COLS.
#   - col 40 (FBA Seller Count): pure integer count — was rendered as a
#     left-aligned string. Added.
#   - col 42 (Buy Box Amazon %): a percentage value — was missing from
#     both NUMERIC_COLS and PCT_COLS, so '"10%"'-style strings never got
#     numeric coercion or % formatting. Added to both.
NUMERIC_COLS: frozenset[int] = frozenset(
    {10, 12, 13, 14, 15, 16, 17, 18, 19, 20, 23, 24, 26, 27, 28, 29, 30, 31,
     32, 33, 34, 35, 36, 37, 38, 40, 42, 50, 61, 62}
)
PCT_COLS: frozenset[int] = frozenset({31, 42, 62})
GBP_COLS: frozenset[int] = frozenset({20, 23, 24, 26, 27, 28, 29, 30, 32, 33, 61})

# Excel number formats.
_FMT_GBP = '"GBP"#,##0.00'
_FMT_PCT = '0.0"%"'
_FMT_INT = "#,##0"
_FMT_FLOAT = "0.0"

# Column indices for cells that need word-wrap + top alignment.
_WRAP_COLS = frozenset({2, 8, 11} | set(range(46, 68)))
# Columns whose data should be aligned center / middle.
_NUMERIC_RE = re.compile(r"[^0-9.\-]")

# ────────────────────────────────────────────────────────────────────────
# Pure helpers.
# ────────────────────────────────────────────────────────────────────────


def verdict_fill(verdict: object) -> str | None:
    """Return the ARGB fill colour for a verdict cell, or None for unknown."""
    if verdict is None:
        return None
    v = str(verdict).strip().upper()
    if not v:
        return None
    if v == "YES":
        return COLOURS["yes_green"]
    if v.startswith("MAYBE"):
        return COLOURS["maybe_yellow"]
    if v == "BRAND APPROACH":
        return COLOURS["brand_blue"]
    if v == "BUY THE DIP":
        return COLOURS["dip_orange"]
    if v == "PRICE EROSION":
        return COLOURS["erosion_red"]
    if v == "GATED":
        return COLOURS["gated_purple"]
    if v == "NO":
        return COLOURS["no_red"]
    return None


def score_fill(value: object) -> str | None:
    """Return a high/mid/low fill colour for a numeric score cell."""
    if value is None or value == "":
        return None
    try:
        n = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(n):
        return None
    if n >= 8:
        return COLOURS["score_bg_high"]
    if n >= 5:
        return COLOURS["score_bg_mid"]
    return COLOURS["score_bg_low"]


def _parse_numeric(value: object) -> float | None:
    """Strip GBP and percent + non-numeric chars, return a float or None."""
    if value is None or value == "" or value == "-":
        return None
    s = str(value)
    s = re.sub(r"GBP", "", s, flags=re.IGNORECASE)
    s = _NUMERIC_RE.sub("", s).strip()
    if not s:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    if math.isnan(n):
        return None
    return n


def _coerce_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _solid_fill(argb: str) -> openpyxl.styles.PatternFill:
    return openpyxl.styles.PatternFill(
        fill_type="solid", start_color=argb, end_color=argb
    )


def _bold_font(size: int = 10, color: str | None = None) -> openpyxl.styles.Font:
    kwargs: dict = {"bold": True, "size": size}
    if color is not None:
        kwargs["color"] = color
    return openpyxl.styles.Font(**kwargs)


def _hair_border() -> openpyxl.styles.Border:
    side = openpyxl.styles.Side(style="hair", color=COLOURS["border_grey"])
    return openpyxl.styles.Border(bottom=side, left=side, right=side)


def _niche_title(niche: str) -> str:
    """'kids-toys' -> 'Kids Toys'."""
    return niche.replace("-", " ").title()


# ────────────────────────────────────────────────────────────────────────
# Data-cell write — handles numeric coercion and per-type formatting.
# ────────────────────────────────────────────────────────────────────────


def _write_cell(cell, value: object, col_idx: int) -> None:
    """Write a cell with appropriate type coercion and number format."""
    if col_idx in NUMERIC_COLS and value not in ("", "-", None):
        n = _parse_numeric(value)
        if n is not None:
            cell.value = n
            if col_idx in PCT_COLS:
                cell.number_format = _FMT_PCT
            elif col_idx in GBP_COLS:
                cell.number_format = _FMT_GBP
            elif n == int(n):
                cell.number_format = _FMT_INT
            else:
                cell.number_format = _FMT_FLOAT
            return
    cell.value = value


# ────────────────────────────────────────────────────────────────────────
# Per-row conditional formatting.
# ────────────────────────────────────────────────────────────────────────


def _apply_conditional_fills(ws, excel_row: int, row: dict) -> None:
    """Apply per-cell colour rules for one data row.

    The legacy JS applies ~10 distinct cell-level rules (verdict, lane,
    composite score band, price compression, ROI text colour, IP risk band,
    Real ROI text colour, Amazon-on-listing flag, Gated flag, Listing
    Quality, Brand 1P, Weight Flag). All preserved here.
    """
    # Verdict (col 7).
    vfill = verdict_fill(row.get("Verdict"))
    if vfill:
        cell = ws.cell(excel_row, 7)
        cell.fill = _solid_fill(vfill)
        cell.font = _bold_font(10)

    # Opportunity Lane (col 9).
    lane = _coerce_str(row.get("Opportunity Lane")).upper()
    if lane == "BALANCED":
        cell = ws.cell(excel_row, 9)
        cell.fill = _solid_fill(COLOURS["yes_green"])
        cell.font = _bold_font(10)
    elif lane == "PROFIT":
        cell = ws.cell(excel_row, 9)
        cell.fill = _solid_fill(COLOURS["brand_blue"])
        cell.font = _bold_font(10)
    elif "CASH" in lane:
        cell = ws.cell(excel_row, 9)
        cell.fill = _solid_fill(COLOURS["dip_orange"])
        cell.font = _bold_font(10)

    # Composite Score (col 12) — high/mid/low band fill.
    sfill = score_fill(row.get("Composite Score"))
    if sfill:
        cell = ws.cell(excel_row, 12)
        cell.fill = _solid_fill(sfill)
        cell.font = _bold_font(11)

    # Price Compression (col 21).
    pc = _coerce_str(row.get("Price Compression")).upper()
    if pc == "COMPRESSED":
        cell = ws.cell(excel_row, 21)
        cell.fill = _solid_fill(COLOURS["no_red"])
        cell.font = _bold_font(10)
    elif pc == "SQUEEZED":
        cell = ws.cell(excel_row, 21)
        cell.fill = _solid_fill(COLOURS["dip_orange"])
        cell.font = _bold_font(10)

    # Listing Quality (col 22) — WEAK in orange.
    if _coerce_str(row.get("Listing Quality")).upper() == "WEAK":
        cell = ws.cell(excel_row, 22)
        cell.fill = _solid_fill(COLOURS["dip_orange"])
        cell.font = _bold_font(10)

    # Est ROI % (col 31) — green/red text by sign.
    roi = _parse_numeric(row.get("Est ROI %"))
    if roi is not None:
        cell = ws.cell(excel_row, 31)
        cell.font = _bold_font(
            10, color=COLOURS["green_text"] if roi >= 0 else COLOURS["red_text"]
        )

    # Brand 1P (col 39) — Y in red.
    if _coerce_str(row.get("Brand 1P")).upper() == "Y":
        cell = ws.cell(excel_row, 39)
        cell.fill = _solid_fill(COLOURS["no_red"])
        cell.font = _bold_font(10)

    # Amazon on Listing (col 41) — Y in orange.
    if _coerce_str(row.get("Amazon on Listing")).upper() == "Y":
        cell = ws.cell(excel_row, 41)
        cell.fill = _solid_fill(COLOURS["dip_orange"])
        cell.font = _bold_font(10)

    # IP Risk Band (col 51).
    ip = _coerce_str(row.get("IP Risk Band")).upper()
    if ip == "HIGH":
        cell = ws.cell(excel_row, 51)
        cell.fill = _solid_fill(COLOURS["no_red"])
        cell.font = _bold_font(10)
    elif ip == "MEDIUM":
        cell = ws.cell(excel_row, 51)
        cell.fill = _solid_fill(COLOURS["dip_orange"])
        cell.font = _bold_font(10)
    elif ip == "LOW":
        cell = ws.cell(excel_row, 51)
        cell.fill = _solid_fill(COLOURS["yes_green"])
        cell.font = _bold_font(10)

    # Gated (col 53) — Y in purple.
    if _coerce_str(row.get("Gated")).upper() == "Y":
        cell = ws.cell(excel_row, 53)
        cell.fill = _solid_fill(COLOURS["gated_purple"])
        cell.font = _bold_font(10)

    # Real ROI % (col 62) — green/red text by sign.
    real_roi = _parse_numeric(row.get("Real ROI %"))
    if real_roi is not None:
        cell = ws.cell(excel_row, 62)
        cell.font = _bold_font(
            10,
            color=COLOURS["green_text"] if real_roi >= 0 else COLOURS["red_text"],
        )

    # Weight Flag (col 6) — HEAVY/OVERSIZE in orange.
    wf = _coerce_str(row.get("Weight Flag")).upper()
    if "HEAVY" in wf or "OVERSIZE" in wf:
        cell = ws.cell(excel_row, 6)
        cell.fill = _solid_fill(COLOURS["dip_orange"])
        cell.font = _bold_font(10)


# ────────────────────────────────────────────────────────────────────────
# Header rows + workbook scaffolding.
# ────────────────────────────────────────────────────────────────────────


def _write_title_row(ws, niche: str, product_count: int, last_col: int) -> None:
    from datetime import date

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1)
    cell.value = (
        f"{_niche_title(niche)} -- Final Results ({product_count} products)  "
        f"|  Generated {date.today().isoformat()}"
    )
    cell.font = openpyxl.styles.Font(
        bold=True, size=14, color=COLOURS["header_font"]
    )
    cell.fill = _solid_fill(COLOURS["header_bg"])
    cell.alignment = openpyxl.styles.Alignment(
        horizontal="left", vertical="center"
    )
    ws.row_dimensions[1].height = 30


def _write_group_headers(ws) -> None:
    # White thin sides give visual separation between groups (matches JS).
    white_side = openpyxl.styles.Side(style="thin", color="FFFFFFFF")
    group_border = openpyxl.styles.Border(left=white_side, right=white_side)
    for g in GROUPS:
        ws.merge_cells(
            start_row=2, start_column=g["start"],
            end_row=2, end_column=g["end"],
        )
        cell = ws.cell(2, g["start"])
        cell.value = g["label"]
        cell.font = openpyxl.styles.Font(
            bold=True, size=11, color=COLOURS["header_font"]
        )
        cell.fill = _solid_fill("FF2C3E50")
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )
        cell.border = group_border
    ws.row_dimensions[2].height = 22


def _write_column_headers(ws, headers: list[str]) -> None:
    # Medium grey bottom + thin navy sides — matches the legacy JS row-3 style.
    bottom_side = openpyxl.styles.Side(style="medium", color="FF7F8C8D")
    side_side = openpyxl.styles.Side(style="thin", color="FF34495E")
    header_border = openpyxl.styles.Border(
        bottom=bottom_side, left=side_side, right=side_side
    )
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(3, i)
        cell.value = h
        cell.font = openpyxl.styles.Font(
            bold=True, size=10, color=COLOURS["header_font"]
        )
        cell.fill = _solid_fill(COLOURS["header_bg"])
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = header_border
    ws.row_dimensions[3].height = 36


def _set_column_widths(ws) -> None:
    for col, width in COL_WIDTHS.items():
        letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[letter].width = width


# ────────────────────────────────────────────────────────────────────────
# Legend sheet.
# ────────────────────────────────────────────────────────────────────────


_LEGEND_DATA: list[tuple[str, str]] = [
    ("Verdict", "Meaning"),
    ("YES", "Composite 7+, all filters pass -- pursue this product"),
    ("MAYBE", "Composite 5-6, one concern -- review needed"),
    ("MAYBE-ROI", "ROI below 20% estimated -- may improve with real trade price"),
    ("BRAND APPROACH", "2-3 sellers, weak listing -- contact brand direct"),
    ("BUY THE DIP", "Price 30%+ below 90-day avg -- recovery pattern detected"),
    ("PRICE EROSION", "Consistent downward slope -- reject"),
    ("GATED", "Restricted listing -- flag for ungating decision"),
    ("NO", "Fails filter, reason stated"),
    ("", ""),
    ("Lane", "Meaning"),
    ("BALANCED", "Strong velocity + strong margin -- premium opportunity"),
    ("PROFIT", "Strong unit profit/ROI, may have lower velocity -- capital efficient"),
    ("CASH FLOW", "High velocity, acceptable margin -- turnover and cash generation"),
    ("", ""),
    ("Column", "Explanation"),
    ("Price Compression", "OK / SQUEEZED (price 10-20% below 90d avg) / COMPRESSED (price 20%+ below avg)"),
    ("Est Cost 65%", "Estimated cost at 65% of selling price (rough placeholder)"),
    ("Max Cost for 20% ROI", "Maximum you can pay for stock and still hit 20% ROI after FBA fees"),
    ("Breakeven Price", "Minimum selling price to cover FBA fees + cost (at 65% estimate)"),
    ("Price Stability", "STABLE / SLIGHT DIP / DROPPING / RISING / SURGING -- based on 90-day trend"),
    ("Route Code", "EXISTING ACCOUNT / DISTRIBUTOR / BRAND DIRECT / TRADE PLATFORM / UNCLEAR"),
    ("FBA Seller Count", "Number of FBA sellers on listing (2-20 target range)"),
    ("Amazon Buy Box Share", "Percentage of time Amazon holds the Buy Box (flag if >70%)"),
    ("Listing Quality", "STRONG (6+ images, A+, 5+ bullets) / AVERAGE / WEAK"),
    ("Weight Flag", "OK / HEAVY (>5kg) / OVERSIZE (>45cm) / HEAVY+OVERSIZE"),
    ("Brand 1P", "Y = brand sells direct on Amazon (hard to compete)"),
    ("Star Rating", "Average customer rating 1.0-5.0"),
    ("Review Count", "Total customer reviews (>500 = proven, <20 = risky)"),
]
# Indices of header rows in _LEGEND_DATA (0-based).
_LEGEND_HEADER_ROWS = {0, 10, 16}


def _write_legend(legend_ws) -> None:
    legend_ws.column_dimensions["A"].width = 20
    legend_ws.column_dimensions["B"].width = 60

    for i, (a, b) in enumerate(_LEGEND_DATA):
        excel_row = i + 1
        ca = legend_ws.cell(excel_row, 1, value=a)
        cb = legend_ws.cell(excel_row, 2, value=b)

        if i in _LEGEND_HEADER_ROWS:
            ca.font = openpyxl.styles.Font(
                bold=True, size=11, color=COLOURS["header_font"]
            )
            cb.font = openpyxl.styles.Font(
                bold=True, size=11, color=COLOURS["header_font"]
            )
            ca.fill = _solid_fill(COLOURS["header_bg"])
            cb.fill = _solid_fill(COLOURS["header_bg"])
            continue

        # Verdict rows (1..8): apply verdict fill.
        if 1 <= i <= 8:
            fill = verdict_fill(a)
            if fill:
                ca.fill = _solid_fill(fill)
                ca.font = _bold_font(10)
            continue

        # Lane rows.
        if i == 11:  # BALANCED
            ca.fill = _solid_fill(COLOURS["yes_green"])
            ca.font = _bold_font(10)
        elif i == 12:  # PROFIT
            ca.fill = _solid_fill(COLOURS["brand_blue"])
            ca.font = _bold_font(10)
        elif i == 13:  # CASH FLOW
            ca.fill = _solid_fill(COLOURS["dip_orange"])
            ca.font = _bold_font(10)


# ────────────────────────────────────────────────────────────────────────
# DataFrame entry point.
# ────────────────────────────────────────────────────────────────────────


def compute_workbook(df: pd.DataFrame, niche: str) -> openpyxl.Workbook:
    """Render a styled openpyxl Workbook from a 67-column final_results frame.

    Pure: does not mutate the input. Caller is responsible for `wb.save(path)`
    or `build_xlsx(...)` for the file-write contract.

    Validates the input frame's schema — the conditional-fill logic
    addresses cells by hard-coded column index (col 7 = Verdict, col 12 =
    Composite Score, etc.), so a reordered frame would silently paint
    fills on the wrong cells.
    """
    # Late import to avoid a circular dep — both modules are sibling steps
    # at the same package level.
    from fba_engine.steps.build_output import FINAL_HEADERS

    headers = list(df.columns)
    if headers != FINAL_HEADERS:
        # Identify the first divergence to make the error actionable.
        diffs = []
        for i, (got, want) in enumerate(zip(headers, FINAL_HEADERS), start=1):
            if got != want:
                diffs.append(f"col {i}: got {got!r}, expected {want!r}")
                if len(diffs) >= 3:
                    break
        if len(headers) != len(FINAL_HEADERS):
            diffs.append(
                f"length mismatch: got {len(headers)}, expected "
                f"{len(FINAL_HEADERS)}"
            )
        raise ValueError(
            "compute_workbook: input frame schema does not match "
            "FINAL_HEADERS. " + "; ".join(diffs or ["unknown divergence"])
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    last_col = len(headers)

    _write_title_row(ws, niche, len(df), last_col)
    _write_group_headers(ws)
    _write_column_headers(ws, headers)
    _set_column_widths(ws)
    ws.freeze_panes = "D4"

    # Cell fonts and borders shared by every data cell.
    default_font = openpyxl.styles.Font(size=10)
    centre_align = openpyxl.styles.Alignment(
        horizontal="center", vertical="center"
    )
    wrap_align = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    hair = _hair_border()
    zebra = _solid_fill(COLOURS["zebra_even"])

    # Data rows start at Excel row 4. Row index `ri` is 0-based for the
    # zebra-stripe parity (matches JS `ri % 2 === 0`).
    for ri, (_, row) in enumerate(df.iterrows()):
        excel_row = ri + 4
        is_even = ri % 2 == 0

        for ci, header in enumerate(headers, start=1):
            cell = ws.cell(excel_row, ci)
            value = row[header]
            # NaN-safe value coercion before write.
            try:
                if pd.isna(value):
                    value = ""
            except (TypeError, ValueError):
                pass

            if ci == 4 and value:
                # Amazon URL — clickable hyperlink with link-blue underline.
                cell.value = value
                cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=cell.coordinate, target=str(value)
                )
                cell.font = openpyxl.styles.Font(
                    color=COLOURS["link_blue"], underline="single", size=9
                )
            else:
                _write_cell(cell, value, ci)
                cell.font = default_font

            cell.alignment = wrap_align if ci in _WRAP_COLS else centre_align
            if is_even:
                cell.fill = zebra
            cell.border = hair

        # Per-row conditional formatting overrides (applied after the base
        # styles so they win on contested cells).
        _apply_conditional_fills(ws, excel_row, row.to_dict())

        ws.row_dimensions[excel_row].height = 22

    # AutoFilter spans all 67 columns and the data range.
    last_letter = openpyxl.utils.get_column_letter(last_col)
    last_row = max(3, len(df) + 3)
    ws.auto_filter.ref = f"A3:{last_letter}{last_row}"

    legend_ws = wb.create_sheet("Legend")
    _write_legend(legend_ws)

    return wb


def build_xlsx(df: pd.DataFrame, niche: str, output_path: str | Path) -> None:
    """Render the workbook + write it to disk."""
    wb = compute_workbook(df, niche)
    wb.save(output_path)


# ────────────────────────────────────────────────────────────────────────
# Step contract.
# ────────────────────────────────────────────────────────────────────────


def run_step(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Step-runner-compatible wrapper.

    Returns the input DataFrame UNCHANGED (passthrough) since the XLSX is a
    side-output, not a transformation. If `config["output_path"]` is set,
    writes the workbook to that path. `config["niche"]` is used for the
    title bar; defaults to "results" if missing.
    """
    if config.get("output_path"):
        niche = config.get("niche", "results")
        build_xlsx(df, niche=niche, output_path=config["output_path"])
    return df


# ────────────────────────────────────────────────────────────────────────
# CLI — mirrors legacy build_final_xlsx.js paths.
# ────────────────────────────────────────────────────────────────────────


def run(niche: str, base: Path) -> None:
    base = Path(base)
    niche_snake = niche.replace("-", "_")

    # Prefer base/{niche}_final_results.csv, fallback to working/.
    primary = base / f"{niche_snake}_final_results.csv"
    fallback = base / "working" / f"{niche_snake}_final_results.csv"
    if primary.exists():
        input_path = primary
    elif fallback.exists():
        input_path = fallback
    else:
        print(
            f"Final results CSV not found for niche '{niche}'.\n"
            f"Looked in:\n  {primary}\n  {fallback}",
            file=sys.stderr,
        )
        sys.exit(1)

    output_path = base / f"{niche_snake}_final_results.xlsx"
    df = pd.read_csv(
        input_path, dtype=str, keep_default_na=False, encoding="utf-8-sig"
    )
    build_xlsx(df, niche=niche, output_path=output_path)
    print(f"Saved: {output_path}")
    print(f"Rows: {len(df)} products")
    print("Sheets: Results + Legend")


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Phase 5 XLSX builder — renders a styled Excel workbook from "
            "the 67-column final_results.csv produced by step 4c.1."
        )
    )
    parser.add_argument(
        "--niche", required=True, help="Niche slug (e.g. kids-toys)"
    )
    parser.add_argument(
        "--base", required=True, type=Path,
        help="Base directory containing final_results.csv (or working/).",
    )
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = _parse_args(sys.argv[1:])
    run(niche=args.niche, base=args.base)
