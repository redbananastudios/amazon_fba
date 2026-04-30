"""Tests for fba_engine.steps.build_xlsx.

Logic ported from `fba_engine/_legacy_keepa/skills/skill-5-build-output/
build_final_xlsx.js` (529 LOC). The JS uses ExcelJS; this port uses openpyxl.

Most styling tests verify *structure and key fills*, not every ARGB string —
brittle pixel-level pinning would break on small visual tweaks without
catching real regressions. The structural tests + the verdict / IP-risk /
score-band fill tests catch the high-signal regressions.

One deliberate deviation from the JS source: legacy hard-codes 64 columns
(autofilter, title merge, COL_WIDTHS) but the upstream `final_results.csv`
has 67 columns since step 4c.1. The Python port extends styling to 67
columns and adds a "Product Codes" group for EAN/UPC/GTIN. This is documented
in the module docstring.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from fba_engine.steps.build_xlsx import (
    COL_WIDTHS,
    COLOURS,
    GBP_COLS,
    GROUPS,
    NUMERIC_COLS,
    PCT_COLS,
    build_xlsx,
    compute_workbook,
    run_step,
    score_fill,
    verdict_fill,
)
from fba_engine.steps.build_output import FINAL_HEADERS, compute_phase5


# Reuse the step 4c.1 fixture builder so tests stay coupled to a real
# Phase-5 schema rather than duplicating column lists.
def _make_phase4_row(**overrides) -> dict:
    base = {
        "ASIN": "B0SAMPLE",
        "Title": "Sample Product",
        "Brand": "Acme",
        "Amazon URL": "https://amzn.eu/d/sample",
        "Category": "Toys",
        "Weight Flag": "OK",
        "Verdict": "YES",
        "Verdict Reason": "Strong demand and margin",
        "Composite Score": "8.5",
        "Demand Score": "8",
        "Stability Score": "8",
        "Competition Score": "7",
        "Margin Score": "9",
        "Cash Flow Score": "8",
        "Profit Score": "8",
        "Balanced Score": "8",
        "Listing Quality": "Good",
        "Opportunity Lane": "BALANCED",
        "Commercial Priority": "1",
        "Lane Reason": "Demand and margin balanced",
        "Monthly Gross Profit": "GBP500",
        "Price Compression": "OK",
        "Current Price": "GBP25.99",
        "Buy Box 90d Avg": "GBP25.50",
        "Price Drop % 90d": "1",
        "Fulfilment Fee": "GBP3.50",
        "Amazon Fees": "GBP5.00",
        "Total Amazon Fees": "GBP8.50",
        "Est Cost 65%": "GBP10.00",
        "Est Profit": "GBP7.49",
        "Est ROI %": "32",
        "Max Cost 20% ROI": "GBP12.00",
        "Breakeven Price": "GBP18.00",
        "BSR Current": "5000",
        "BSR Drops 90d": "200",
        "Bought per Month": "150",
        "Star Rating": "4.5",
        "Review Count": "200",
        "Brand 1P": "N",
        "FBA Seller Count": "5",
        "Amazon on Listing": "N",
        "Buy Box Amazon %": "10%",
        "Brand Seller Match": "NO",
        "Fortress Listing": "NO",
        "Brand Type": "GENERIC",
        "A+ Content Present": "Y",
        "Brand Store Present": "UNLIKELY",
        "Category Risk Level": "MEDIUM",
        "IP Risk Score": "3",
        "IP Risk Band": "Low",
        "IP Reason": "A+ content",
        "Gated": "N",
        "SAS Flags": "",
        "EAN": "1234567890123",
        "UPC": "",
        "GTIN": "",
    }
    base.update(overrides)
    return base


def _make_final_df(rows: list[dict] | None = None) -> pd.DataFrame:
    """Run the step 4c.1 pipeline to produce a real 67-col final_results frame."""
    raw_rows = rows if rows is not None else [_make_phase4_row()]
    df = pd.DataFrame(raw_rows)
    final_df, _, _ = compute_phase5(df)
    return final_df


# ---------------------------------------------------------------------------
# Module constants
# ---------------------------------------------------------------------------


class TestConstants:
    def test_colours_has_required_keys(self):
        # Pin the keys but not the values — values may be tweaked in design
        # iterations, but missing keys would crash the styling code.
        for key in [
            "header_bg", "header_font", "yes_green", "maybe_yellow", "no_red",
            "brand_blue", "dip_orange", "erosion_red", "gated_purple",
            "zebra_even", "border_grey", "green_text", "red_text", "link_blue",
            "score_bg_high", "score_bg_mid", "score_bg_low",
        ]:
            assert key in COLOURS, key

    def test_col_widths_covers_all_67_columns(self):
        # Reviewer concern: legacy JS only widths cols 1-64. Python port
        # must extend to 67 to match the actual final CSV schema.
        for i in range(1, 68):
            assert i in COL_WIDTHS, f"missing width for col {i}"

    def test_groups_cover_1_to_67(self):
        starts_ends = [(g["start"], g["end"]) for g in GROUPS]
        assert min(s for s, _ in starts_ends) == 1
        assert max(e for _, e in starts_ends) == 67

    def test_numeric_cols_and_subsets_are_disjoint(self):
        # PCT and GBP cols should be subsets of NUMERIC.
        assert PCT_COLS.issubset(NUMERIC_COLS)
        assert GBP_COLS.issubset(NUMERIC_COLS)


# ---------------------------------------------------------------------------
# Verdict fills + score fills
# ---------------------------------------------------------------------------


class TestVerdictFill:
    @pytest.mark.parametrize(
        "verdict,expected_key",
        [
            ("YES", "yes_green"),
            ("yes", "yes_green"),
            ("MAYBE", "maybe_yellow"),
            ("MAYBE-ROI", "maybe_yellow"),
            ("BRAND APPROACH", "brand_blue"),
            ("BUY THE DIP", "dip_orange"),
            ("PRICE EROSION", "erosion_red"),
            ("GATED", "gated_purple"),
            ("NO", "no_red"),
        ],
    )
    def test_known_verdicts(self, verdict, expected_key):
        assert verdict_fill(verdict) == COLOURS[expected_key]

    def test_unknown_returns_none(self):
        assert verdict_fill("WEIRD") is None
        assert verdict_fill("") is None
        assert verdict_fill(None) is None


class TestScoreFill:
    def test_high_score(self):
        assert score_fill("8") == COLOURS["score_bg_high"]
        assert score_fill("9.5") == COLOURS["score_bg_high"]

    def test_mid_score(self):
        assert score_fill("5") == COLOURS["score_bg_mid"]
        assert score_fill("7.9") == COLOURS["score_bg_mid"]

    def test_low_score(self):
        assert score_fill("3") == COLOURS["score_bg_low"]
        assert score_fill("4.9") == COLOURS["score_bg_low"]

    def test_non_numeric_returns_none(self):
        assert score_fill("") is None
        assert score_fill("N/A") is None
        assert score_fill(None) is None


# ---------------------------------------------------------------------------
# Workbook structure
# ---------------------------------------------------------------------------


class TestWorkbookStructure:
    def test_returns_workbook_with_results_and_legend_sheets(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        assert "Results" in wb.sheetnames
        assert "Legend" in wb.sheetnames

    def test_results_sheet_has_three_header_rows(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # Row 1: title; row 2: group headers; row 3: column headers; row 4+: data.
        assert ws.cell(1, 1).value is not None  # title
        assert ws.cell(2, 1).value is not None  # group label
        assert ws.cell(3, 1).value == "ASIN"  # column header

    def test_data_starts_at_row_4(self):
        df = _make_final_df([_make_phase4_row(ASIN="B001"), _make_phase4_row(ASIN="B002")])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        assert ws.cell(4, 1).value == "B001"
        assert ws.cell(5, 1).value == "B002"

    def test_title_row_includes_niche_and_count(self):
        df = _make_final_df([_make_phase4_row(ASIN=f"B{i:03d}") for i in range(3)])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        title = ws.cell(1, 1).value
        assert "Kids Toys" in title
        assert "3" in title  # product count

    def test_column_headers_match_final_headers(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        for i, header in enumerate(FINAL_HEADERS, start=1):
            assert ws.cell(3, i).value == header, (
                f"col {i}: expected {header!r}, got {ws.cell(3, i).value!r}"
            )

    def test_freeze_panes_set_to_d4(self):
        # 3 rows + 3 columns frozen → freeze panes anchor is D4.
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        assert ws.freeze_panes == "D4"

    def test_autofilter_spans_full_data_range(self):
        df = _make_final_df([_make_phase4_row(ASIN=f"B{i:03d}") for i in range(5)])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # 67 cols -> last column letter "BO". 5 data rows + 3 header rows = row 8.
        assert ws.auto_filter.ref == "A3:BO8"


class TestColumnWidths:
    def test_column_one_width_set(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # ASIN col width.
        assert ws.column_dimensions["A"].width == COL_WIDTHS[1]

    def test_all_67_columns_have_width(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        for i in range(1, 68):
            col_letter = openpyxl.utils.get_column_letter(i)
            assert ws.column_dimensions[col_letter].width == COL_WIDTHS[i]


# ---------------------------------------------------------------------------
# Number formatting
# ---------------------------------------------------------------------------


class TestNumberFormats:
    def test_gbp_columns_get_gbp_format(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # "Current Price" is col 23, value GBP25.99 → numeric 25.99 with GBP format.
        cell = ws.cell(4, 23)
        assert "GBP" in (cell.number_format or "")

    def test_pct_columns_get_percent_format(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # "Est ROI %" is col 31, value "32.0%".
        cell = ws.cell(4, 31)
        assert "%" in (cell.number_format or "")

    def test_numeric_integer_format(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # "BSR Current" is col 34, integer.
        cell = ws.cell(4, 34)
        assert cell.number_format in {"#,##0", "0.0"}


# ---------------------------------------------------------------------------
# Conditional formatting (per-cell fills based on row content)
# ---------------------------------------------------------------------------


def _has_fill(cell) -> bool:
    """True iff cell has a non-default solid fill."""
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    val = getattr(fg, "rgb", None) or getattr(fg, "value", None)
    return val not in (None, "00000000", "FFFFFFFF")


def _fill_argb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    return getattr(fg, "rgb", None) or getattr(fg, "value", None)


class TestConditionalFills:
    def test_yes_verdict_cell_filled_green(self):
        df = _make_final_df([_make_phase4_row(**{"Verdict": "YES"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 7)  # Verdict column
        assert _fill_argb(cell) == COLOURS["yes_green"]

    def test_no_verdict_cell_filled_red(self):
        df = _make_final_df([_make_phase4_row(**{"Verdict": "NO"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 7)
        assert _fill_argb(cell) == COLOURS["no_red"]

    def test_balanced_lane_cell_green(self):
        df = _make_final_df([_make_phase4_row(**{"Opportunity Lane": "BALANCED"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 9)  # Opportunity Lane
        assert _fill_argb(cell) == COLOURS["yes_green"]

    def test_profit_lane_cell_blue(self):
        df = _make_final_df([_make_phase4_row(**{"Opportunity Lane": "PROFIT"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 9)
        assert _fill_argb(cell) == COLOURS["brand_blue"]

    def test_cash_flow_lane_cell_orange(self):
        df = _make_final_df([_make_phase4_row(**{"Opportunity Lane": "CASH FLOW"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 9)
        assert _fill_argb(cell) == COLOURS["dip_orange"]

    def test_high_composite_score_filled_green(self):
        df = _make_final_df([_make_phase4_row(**{"Composite Score": "9.0"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 12)  # Composite Score
        assert _fill_argb(cell) == COLOURS["score_bg_high"]

    def test_high_ip_risk_band_filled_red(self):
        df = _make_final_df([_make_phase4_row(**{"IP Risk Band": "High"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 51)  # IP Risk Band
        assert _fill_argb(cell) == COLOURS["no_red"]

    def test_low_ip_risk_band_filled_green(self):
        df = _make_final_df([_make_phase4_row(**{"IP Risk Band": "Low"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 51)
        assert _fill_argb(cell) == COLOURS["yes_green"]

    def test_amazon_on_listing_y_filled_orange(self):
        df = _make_final_df([_make_phase4_row(**{"Amazon on Listing": "Y"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 41)  # Amazon on Listing
        assert _fill_argb(cell) == COLOURS["dip_orange"]

    def test_gated_y_filled_purple(self):
        df = _make_final_df([_make_phase4_row(**{"Gated": "Y"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 53)  # Gated
        assert _fill_argb(cell) == COLOURS["gated_purple"]

    def test_brand_1p_y_filled_red(self):
        # Brand 1P=Y rows actually get rejected at step 4c.1 (PL), so build
        # a frame that bypasses that — pass a synthetic row directly.
        rows = [{h: "" for h in FINAL_HEADERS}]
        rows[0]["ASIN"] = "B0BRAND"
        rows[0]["Brand 1P"] = "Y"
        rows[0]["Verdict"] = "YES"
        df = pd.DataFrame(rows, columns=FINAL_HEADERS)
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 39)  # Brand 1P
        assert _fill_argb(cell) == COLOURS["no_red"]

    # Reviewer M3: missing conditional-fill rule coverage. Each rule below
    # had no cell-level test before — adding to lock against regressions.

    def test_price_compression_compressed_filled_red(self):
        df = _make_final_df([_make_phase4_row(**{"Price Compression": "COMPRESSED"})])
        wb = compute_workbook(df, niche="kids-toys")
        cell = wb["Results"].cell(4, 21)
        assert _fill_argb(cell) == COLOURS["no_red"]

    def test_price_compression_squeezed_filled_orange(self):
        df = _make_final_df([_make_phase4_row(**{"Price Compression": "SQUEEZED"})])
        wb = compute_workbook(df, niche="kids-toys")
        cell = wb["Results"].cell(4, 21)
        assert _fill_argb(cell) == COLOURS["dip_orange"]

    def test_listing_quality_weak_filled_orange(self):
        df = _make_final_df([_make_phase4_row(**{"Listing Quality": "WEAK"})])
        wb = compute_workbook(df, niche="kids-toys")
        cell = wb["Results"].cell(4, 22)
        assert _fill_argb(cell) == COLOURS["dip_orange"]

    def test_weight_flag_heavy_filled_orange(self):
        df = _make_final_df([_make_phase4_row(**{"Weight Flag": "HEAVY"})])
        wb = compute_workbook(df, niche="kids-toys")
        cell = wb["Results"].cell(4, 6)
        assert _fill_argb(cell) == COLOURS["dip_orange"]

    def test_weight_flag_oversize_filled_orange(self):
        df = _make_final_df([_make_phase4_row(**{"Weight Flag": "HEAVY+OVERSIZE"})])
        wb = compute_workbook(df, niche="kids-toys")
        cell = wb["Results"].cell(4, 6)
        assert _fill_argb(cell) == COLOURS["dip_orange"]

    def test_medium_ip_risk_band_filled_orange(self):
        df = _make_final_df([_make_phase4_row(**{"IP Risk Band": "Medium"})])
        wb = compute_workbook(df, niche="kids-toys")
        cell = wb["Results"].cell(4, 51)
        assert _fill_argb(cell) == COLOURS["dip_orange"]


class TestRoiTextColour:
    def test_positive_roi_green_text(self):
        df = _make_final_df([_make_phase4_row(**{"Est ROI %": "32"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 31)  # Est ROI %
        # openpyxl exposes font color as Color obj; the rgb attr is the ARGB string.
        font = cell.font
        rgb = getattr(font.color, "rgb", None) if font.color else None
        assert rgb == COLOURS["green_text"]

    def test_negative_roi_red_text(self):
        df = _make_final_df([_make_phase4_row(**{"Est ROI %": "-5"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 31)
        font = cell.font
        rgb = getattr(font.color, "rgb", None) if font.color else None
        assert rgb == COLOURS["red_text"]

    def test_real_roi_positive_green_text(self):
        # Real ROI is col 62 — placeholder is empty in step 4c.1 output, so
        # synthesise a row directly to exercise the rule.
        rows = [{h: "" for h in FINAL_HEADERS}]
        rows[0]["ASIN"] = "B0REAL"
        rows[0]["Real ROI %"] = "25%"
        rows[0]["Verdict"] = "YES"
        df = pd.DataFrame(rows, columns=FINAL_HEADERS)
        wb = compute_workbook(df, niche="kids-toys")
        font = wb["Results"].cell(4, 62).font
        rgb = getattr(font.color, "rgb", None) if font.color else None
        assert rgb == COLOURS["green_text"]

    def test_real_roi_negative_red_text(self):
        rows = [{h: "" for h in FINAL_HEADERS}]
        rows[0]["ASIN"] = "B0REAL"
        rows[0]["Real ROI %"] = "-10%"
        rows[0]["Verdict"] = "YES"
        df = pd.DataFrame(rows, columns=FINAL_HEADERS)
        wb = compute_workbook(df, niche="kids-toys")
        font = wb["Results"].cell(4, 62).font
        rgb = getattr(font.color, "rgb", None) if font.color else None
        assert rgb == COLOURS["red_text"]


class TestZebraStriping:
    def test_even_indexed_data_rows_have_zebra_fill(self):
        # The legacy JS uses ri%2==0 (0-indexed) for "even" fill — i.e.
        # the first data row (ri=0). The Python port must match.
        rows = [_make_phase4_row(ASIN=f"B{i:03d}") for i in range(2)]
        df = _make_final_df(rows)
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # First data row (Excel row 4): zebra. Cells WITHOUT verdict/etc fills
        # should show the zebra colour. Pick a column without conditional fills.
        # Col 1 (ASIN) has no conditional rule, so it's a clean check.
        first_data_asin = ws.cell(4, 1)
        # Either has zebra fill OR is the second data row (no zebra).
        # Determinism: the JS marks ri=0 (first data row) as even = zebra.
        assert _fill_argb(first_data_asin) == COLOURS["zebra_even"]

    def test_odd_indexed_rows_no_zebra(self):
        rows = [_make_phase4_row(ASIN=f"B{i:03d}") for i in range(2)]
        df = _make_final_df(rows)
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # Second data row (Excel row 5, ri=1): no zebra.
        second_data_asin = ws.cell(5, 1)
        assert _fill_argb(second_data_asin) != COLOURS["zebra_even"]


# ---------------------------------------------------------------------------
# Hyperlink for Amazon URL
# ---------------------------------------------------------------------------


class TestHyperlink:
    def test_amazon_url_cell_has_hyperlink(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 4)  # Amazon URL
        assert cell.hyperlink is not None
        # openpyxl Hyperlink has a target attribute.
        target = getattr(cell.hyperlink, "target", None) or str(cell.hyperlink)
        assert "https://amzn.eu/d/sample" in target


# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------


class TestLegend:
    def test_legend_sheet_exists(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        assert "Legend" in wb.sheetnames

    def test_legend_first_row_is_verdict_header(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        legend = wb["Legend"]
        assert legend.cell(1, 1).value == "Verdict"

    def test_legend_includes_all_known_verdicts(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        legend = wb["Legend"]
        # Read all values in column 1.
        col_a_values = [legend.cell(r, 1).value for r in range(1, 30)]
        for verdict in ["YES", "MAYBE", "BRAND APPROACH", "BUY THE DIP", "GATED", "NO"]:
            assert verdict in col_a_values

    def test_legend_section_headers_have_header_bg_fill(self):
        # Rows 1, 11, 17 (1-based) are the 3 section headers.
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        legend = wb["Legend"]
        for excel_row in (1, 11, 17):
            cell = legend.cell(excel_row, 1)
            assert _fill_argb(cell) == COLOURS["header_bg"], f"row {excel_row}"


# ---------------------------------------------------------------------------
# build_xlsx (file write) + run_step contract
# ---------------------------------------------------------------------------


class TestBuildXlsx:
    def test_build_xlsx_writes_file(self, tmp_path: Path):
        df = _make_final_df()
        out_path = tmp_path / "test.xlsx"
        build_xlsx(df, niche="kids-toys", output_path=out_path)
        assert out_path.exists()
        # Verify the file is a valid openpyxl workbook.
        wb = openpyxl.load_workbook(out_path)
        assert "Results" in wb.sheetnames


class TestRunStep:
    def test_run_step_returns_input_unchanged_when_no_output_path(self):
        df = _make_final_df()
        out = run_step(df, {})
        # XLSX is a side-output; run_step is a passthrough on the DataFrame.
        pd.testing.assert_frame_equal(out, df)

    def test_run_step_writes_xlsx_when_output_path_in_config(self, tmp_path: Path):
        df = _make_final_df()
        out_path = tmp_path / "via_runstep.xlsx"
        out = run_step(df, {"niche": "kids-toys", "output_path": str(out_path)})
        assert out_path.exists()
        # Returns the input df unchanged (passthrough contract).
        pd.testing.assert_frame_equal(out, df)


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    def test_empty_df_produces_workbook_with_only_header_rows(self):
        df = pd.DataFrame(columns=FINAL_HEADERS)
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # No data rows — just the 3 header rows.
        assert ws.cell(4, 1).value is None

    def test_unknown_verdict_does_not_crash(self):
        df = _make_final_df([_make_phase4_row(**{"Verdict": "WEIRDVERDICT"})])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        # Verdict cell should still have data, just no fill.
        assert ws.cell(4, 7).value == "WEIRDVERDICT"

    def test_unicode_brand_does_not_crash(self):
        df = _make_final_df([_make_phase4_row(Brand="Café Ø", Title="日本商品")])
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        assert ws.cell(4, 3).value == "Café Ø"

    def test_niche_label_title_case(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="afro-hair")
        ws = wb["Results"]
        # "afro-hair" → "Afro Hair" in title.
        title = ws.cell(1, 1).value
        assert "Afro Hair" in title


class TestSchemaValidation:
    """Regression: hard-coded conditional-fill column indices mean a
    reordered frame would silently paint fills on the wrong cells."""

    def test_reordered_frame_raises(self):
        # Swap two adjacent columns — same length, same columns, wrong order.
        bad_headers = list(FINAL_HEADERS)
        bad_headers[6], bad_headers[7] = bad_headers[7], bad_headers[6]
        rows = [{h: "" for h in bad_headers}]
        df = pd.DataFrame(rows, columns=bad_headers)
        with pytest.raises(ValueError, match="schema does not match"):
            compute_workbook(df, niche="kids-toys")

    def test_short_frame_raises(self):
        bad_headers = FINAL_HEADERS[:50]
        rows = [{h: "" for h in bad_headers}]
        df = pd.DataFrame(rows, columns=bad_headers)
        with pytest.raises(ValueError, match="schema does not match"):
            compute_workbook(df, niche="kids-toys")


class TestHeaderBorders:
    """Regression: row-2 group headers and row-3 column headers must have
    visible borders matching the legacy JS styling."""

    def test_group_header_has_white_side_borders(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(2, 1)  # First group's start cell.
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
        # Colour normalisation across openpyxl versions: accept "FFFFFFFF"
        # or "00FFFFFF" (RGB / ARGB pair) — the visual is the same.
        assert (cell.border.left.color.value or "").upper().endswith("FFFFFF")

    def test_column_header_has_bottom_border(self):
        df = _make_final_df()
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(3, 1)
        assert cell.border.bottom.style == "medium"
        assert cell.border.left.style == "thin"


class TestNumericColsExtended:
    """Regression: cols 40 (FBA Seller Count) and 42 (BB Amazon %) must
    apply numeric formatting — the legacy JS missed these."""

    def test_fba_seller_count_col_40_formatted_as_number(self):
        rows = [{h: "" for h in FINAL_HEADERS}]
        rows[0][FINAL_HEADERS[39]] = "5"  # col 40, 1-based
        df = pd.DataFrame(rows, columns=FINAL_HEADERS)
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 40)
        # Should be coerced to a number, not left as a string.
        assert cell.value == 5
        assert cell.number_format == "#,##0"

    def test_bb_amazon_pct_col_42_formatted_as_percent(self):
        rows = [{h: "" for h in FINAL_HEADERS}]
        rows[0][FINAL_HEADERS[41]] = "10%"  # col 42, 1-based
        df = pd.DataFrame(rows, columns=FINAL_HEADERS)
        wb = compute_workbook(df, niche="kids-toys")
        ws = wb["Results"]
        cell = ws.cell(4, 42)
        # _parse_numeric strips '%' and gives 10.0; PCT format applies.
        assert cell.value == 10
        assert cell.number_format == '0.0"%"'
