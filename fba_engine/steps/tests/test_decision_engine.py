"""Tests for fba_engine.steps.decision_engine.

Logic ported 1:1 from `fba_engine/_legacy_keepa/skills/skill-6-decision-engine/
phase6_decision.js` so these tests double as a regression contract for the
JS->Python port.

Coverage targets the same density as test_ip_risk.py: every helper exercised,
every decision-rule branch hit, NaN-safety and run_step contract pinned, plus
the JS-vs-Python porting traps (Math.round half-rounding, GBP coercion).
"""
from __future__ import annotations

import math
import warnings
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from fba_engine.steps.decision_engine import (
    DECISION_HEADERS,
    as_upper,
    build_decision_reason,
    build_handoff,
    build_shortlist_xlsx,
    build_stats,
    calc_action_note,
    calc_buy_readiness,
    calc_joinability,
    calc_margin_status,
    calc_target_buy_price,
    clamp,
    compute_decisions,
    gbp,
    get_target_buffer,
    is_truthy_y,
    lane_base_score,
    parse_money,
    parse_pct,
    pl_risk_score,
    risk_band_score,
    route_score,
    run_step,
    score_from_bought,
    score_from_monthly_gross,
    score_from_priority,
    score_from_roi,
    score_from_unit_profit,
    stable_state,
)


# ---------------------------------------------------------------------------
# Numeric / string coercion helpers
# ---------------------------------------------------------------------------


class TestParseMoney:
    def test_strips_gbp_prefix(self):
        assert parse_money("GBP1.50") == 1.5
        assert parse_money("gbp 2.25") == 2.25

    def test_strips_currency_symbols(self):
        assert parse_money("£5") == 5.0
        assert parse_money("$3.20") == 3.2

    def test_handles_negative(self):
        assert parse_money("-2.50") == -2.5

    def test_empty_or_garbage_returns_zero(self):
        assert parse_money("") == 0
        assert parse_money(None) == 0
        assert parse_money("abc") == 0

    def test_handles_nan(self):
        assert parse_money(float("nan")) == 0


class TestParsePct:
    def test_strips_percent_sign(self):
        assert parse_pct("20%") == 20.0

    def test_handles_decimal(self):
        assert parse_pct("20.5%") == 20.5

    def test_empty_returns_zero(self):
        assert parse_pct("") == 0
        assert parse_pct(None) == 0

    def test_garbage_returns_zero(self):
        assert parse_pct("abc%") == 0


class TestGbp:
    def test_formats_two_decimals(self):
        assert gbp(1.5) == "GBP1.50"
        assert gbp(0) == "GBP0.00"

    def test_negative_includes_sign(self):
        assert gbp(-2.5) == "GBP-2.50"

    def test_empty_returns_empty_string(self):
        assert gbp("") == ""
        assert gbp(None) == ""

    def test_non_numeric_returns_empty_string(self):
        assert gbp("abc") == ""
        assert gbp(float("nan")) == ""


class TestAsUpper:
    def test_uppercases_and_strips(self):
        assert as_upper("  hello  ") == "HELLO"

    def test_handles_none(self):
        assert as_upper(None) == ""

    def test_handles_numbers(self):
        assert as_upper(42) == "42"


class TestIsTruthyY:
    @pytest.mark.parametrize("token", ["Y", "y", "YES", "yes", "TRUE", "true"])
    def test_truthy_tokens(self, token):
        assert is_truthy_y(token) is True

    @pytest.mark.parametrize("token", ["N", "NO", "false", "", "maybe"])
    def test_falsy_tokens(self, token):
        assert is_truthy_y(token) is False

    def test_handles_none(self):
        assert is_truthy_y(None) is False


# ---------------------------------------------------------------------------
# Categorical scoring helpers
# ---------------------------------------------------------------------------


class TestStableState:
    @pytest.mark.parametrize("value", ["STABLE", "stable up", "RISING quickly"])
    def test_good_when_stable_or_rising(self, value):
        assert stable_state(value) == "GOOD"

    def test_caution_for_slight_dip(self):
        assert stable_state("Slight Dip recovering") == "CAUTION"

    @pytest.mark.parametrize("value", ["DROPPING", "surging", "compressed"])
    def test_bad_for_drop_surge_compress(self, value):
        assert stable_state(value) == "BAD"

    def test_unknown_for_anything_else(self):
        assert stable_state("normal") == "UNKNOWN"
        assert stable_state("") == "UNKNOWN"


class TestLaneBaseScore:
    @pytest.mark.parametrize(
        "lane,expected",
        [("BALANCED", 92), ("balanced", 92), ("CASH FLOW", 84), ("PROFIT", 78)],
    )
    def test_known_lanes(self, lane, expected):
        assert lane_base_score(lane) == expected

    def test_unknown_lane_default(self):
        assert lane_base_score("UNASSIGNED") == 42
        assert lane_base_score("") == 42


class TestScoreFromPriority:
    def test_priority_one_is_max(self):
        assert score_from_priority(1) == 100

    def test_priority_two_is_88(self):
        assert score_from_priority(2) == 88

    def test_priority_six_floors_at_forty(self):
        # 100 - (6-1)*12 = 40 (boundary).
        assert score_from_priority(6) == 40

    def test_high_priority_clamped_to_floor(self):
        assert score_from_priority(20) == 40

    def test_zero_or_missing_treated_as_eight(self):
        # JS `Number(priority) || 8` collapses 0/NaN/empty to 8 -> 40.
        assert score_from_priority(0) == 40
        assert score_from_priority(None) == 40
        assert score_from_priority("") == 40


class TestScoreFromMonthlyGross:
    def test_zero_returns_zero(self):
        assert score_from_monthly_gross(0) == 0

    def test_300_returns_50(self):
        assert score_from_monthly_gross(300) == 50

    def test_clamps_at_100(self):
        assert score_from_monthly_gross(10000) == 100

    def test_handles_garbage(self):
        assert score_from_monthly_gross("abc") == 0


class TestScoreFromBought:
    def test_zero_returns_zero(self):
        assert score_from_bought(0) == 0

    def test_eighty_caps_at_100(self):
        assert score_from_bought(80) == 100

    def test_forty_returns_50(self):
        assert score_from_bought(40) == 50


class TestScoreFromUnitProfit:
    def test_zero_returns_zero(self):
        assert score_from_unit_profit(0) == 0

    def test_under_cap(self):
        assert score_from_unit_profit(5) == 60

    def test_caps_at_100(self):
        assert score_from_unit_profit(20) == 100


class TestScoreFromRoi:
    def test_zero_returns_zero(self):
        assert score_from_roi(0) == 0

    def test_under_cap(self):
        assert score_from_roi(20) == pytest.approx(44.0)

    def test_caps_at_100(self):
        assert score_from_roi(100) == 100


class TestRiskBandScore:
    @pytest.mark.parametrize(
        "band,expected",
        [("LOW", 92), ("low", 92), ("MEDIUM", 55), ("HIGH", 8), ("", 50), ("Unknown", 50)],
    )
    def test_band_scores(self, band, expected):
        assert risk_band_score(band) == expected


class TestPlRiskScore:
    @pytest.mark.parametrize(
        "risk,expected",
        [
            ("UNLIKELY", 88),
            ("low", 88),
            ("LIKELY", 22),
            ("HIGH", 22),
            ("-", 55),
        ],
    )
    def test_pl_risk_scores(self, risk, expected):
        assert pl_risk_score(risk) == expected


class TestRouteScore:
    @pytest.mark.parametrize("value", ["", "UNCLEAR", "unclear"])
    def test_unclear_returns_30(self, value):
        assert route_score(value) == 30

    def test_known_route_returns_78(self):
        assert route_score("DIRECT") == 78


class TestGetTargetBuffer:
    @pytest.mark.parametrize(
        "lane,expected",
        [("BALANCED", 2.0), ("CASH FLOW", 1.5), ("PROFIT", 2.5), ("UNKNOWN", 1.25)],
    )
    def test_buffers_per_lane(self, lane, expected):
        assert get_target_buffer(lane) == expected


class TestCalcTargetBuyPrice:
    def test_zero_max_returns_empty(self):
        assert calc_target_buy_price(0, "BALANCED") == ""

    def test_negative_max_returns_empty(self):
        assert calc_target_buy_price(-5, "BALANCED") == ""

    def test_balanced_lane_uses_lower_of_discount_and_buffer(self):
        # max=10, BALANCED: discount=9, buffer=10-2=8 -> min=8.
        assert calc_target_buy_price(10, "BALANCED") == 8.0

    def test_profit_lane_uses_buffer(self):
        # max=10, PROFIT: discount=9, buffer=10-2.5=7.5 -> min=7.5.
        assert calc_target_buy_price(10, "PROFIT") == 7.5

    def test_returns_empty_when_buffer_consumes_full_max(self):
        # max=1, BALANCED: discount=0.9, buffer=1-2=-1 -> min=-1 <= 0
        # -> "" (deliberate fix: legacy leaked "GBP0.00" here).
        assert calc_target_buy_price(1, "BALANCED") == ""

    def test_returns_empty_when_buffer_exactly_consumes_max(self):
        # max=2, BALANCED (buffer=2): buffered=0, discount=1.8 -> min=0
        # -> "" (no useful Target Buy Price).
        assert calc_target_buy_price(2, "BALANCED") == ""

    def test_profit_lane_returns_empty_for_small_max(self):
        # max=2, PROFIT (buffer=2.5): buffered=-0.5 -> "".
        assert calc_target_buy_price(2, "PROFIT") == ""


class TestClamp:
    def test_within(self):
        assert clamp(5, 0, 10) == 5

    def test_below_min(self):
        assert clamp(-1, 0, 10) == 0

    def test_above_max(self):
        assert clamp(15, 0, 10) == 10


# ---------------------------------------------------------------------------
# Decision-logic helpers
# ---------------------------------------------------------------------------


class TestCalcJoinability:
    def test_high_ip_risk_is_unsafe(self):
        row = {"IP Risk Band": "High"}
        assert calc_joinability(row) == "Unsafe"

    def test_fortress_with_brand_seller_match_yes_unsafe(self):
        row = {
            "IP Risk Band": "Low",
            "Fortress Listing": "YES",
            "Brand Seller Match": "YES",
        }
        assert calc_joinability(row) == "Unsafe"

    def test_fortress_with_brand_seller_partial_unsafe(self):
        row = {
            "IP Risk Band": "Low",
            "Fortress Listing": "YES",
            "Brand Seller Match": "PARTIAL",
        }
        assert calc_joinability(row) == "Unsafe"

    def test_established_brand_with_seller_match_and_store_unsafe(self):
        row = {
            "IP Risk Band": "Low",
            "Fortress Listing": "NO",
            "Brand Seller Match": "YES",
            "Brand Type": "ESTABLISHED",
            "Brand Store Present": "LIKELY",
        }
        assert calc_joinability(row) == "Unsafe"

    def test_medium_ip_risk_is_review(self):
        row = {"IP Risk Band": "Medium", "Fortress Listing": "NO"}
        assert calc_joinability(row) == "Review"

    def test_likely_pl_risk_is_review(self):
        row = {
            "IP Risk Band": "Low",
            "Fortress Listing": "NO",
            "Private Label Risk": "LIKELY",
        }
        assert calc_joinability(row) == "Review"

    def test_gated_is_review(self):
        row = {"IP Risk Band": "Low", "Fortress Listing": "NO", "Gated": "Y"}
        assert calc_joinability(row) == "Review"

    def test_clean_listing_joinable(self):
        row = {
            "IP Risk Band": "Low",
            "Fortress Listing": "NO",
            "Brand Seller Match": "NO",
            "Private Label Risk": "UNLIKELY",
            "Gated": "N",
        }
        assert calc_joinability(row) == "Joinable"


class TestCalcMarginStatus:
    def test_supplier_at_or_below_target_safe(self):
        # actual=8, target=8 -> Safe.
        assert calc_margin_status(8, 8, 10, 5, 30) == "Safe"

    def test_supplier_above_target_below_max_tight(self):
        assert calc_margin_status(9, 8, 10, 5, 30) == "Tight"

    def test_supplier_above_max_fail(self):
        assert calc_margin_status(11, 8, 10, 5, 30) == "Fail"

    def test_no_supplier_no_max_unknown(self):
        assert calc_margin_status(0, "", 0, 5, 30) == "Unknown"

    def test_no_supplier_low_profit_fails(self):
        # estProfit < 0.75 -> Fail.
        assert calc_margin_status(0, "", 10, 0.5, 30) == "Fail"

    def test_no_supplier_low_roi_fails(self):
        # estRoi < 10 -> Fail.
        assert calc_margin_status(0, "", 10, 5, 5) == "Fail"

    def test_no_supplier_acceptable_unknown(self):
        # has max + decent profit/roi -> Unknown (not Fail).
        assert calc_margin_status(0, "", 10, 5, 30) == "Unknown"

    def test_supplier_with_empty_target_skips_safe_check(self):
        # When targetBuyPrice is empty (max was 0), Safe branch skipped;
        # max is 0 so Tight branch skipped -> Fail.
        assert calc_margin_status(5, "", 0, 5, 30) == "Fail"


class TestCalcBuyReadiness:
    def test_unsafe_or_fail_rejects(self):
        assert (
            calc_buy_readiness("Unsafe", "Safe", True, True, False) == "Reject"
        )
        assert calc_buy_readiness("Joinable", "Fail", True, True, False) == "Reject"

    def test_no_cost_strong_signals_cost_needed(self):
        # commerciallyStrong but no cost -> Cost Needed.
        assert (
            calc_buy_readiness("Joinable", "Unknown", False, True, False)
            == "Cost Needed"
        )

    def test_no_cost_weak_review_needed(self):
        assert (
            calc_buy_readiness("Joinable", "Unknown", False, False, False)
            == "Review Needed"
        )

    def test_review_or_gated_review_needed(self):
        assert (
            calc_buy_readiness("Review", "Safe", True, True, False)
            == "Review Needed"
        )
        assert (
            calc_buy_readiness("Joinable", "Safe", True, True, True)
            == "Review Needed"
        )

    def test_clean_safe_supply_ready(self):
        assert (
            calc_buy_readiness("Joinable", "Safe", True, True, False) == "Ready"
        )


class TestCalcActionNote:
    def test_buy_action(self):
        assert calc_action_note("BUY", "Ready", "GOOD", True) == "Place opening order"

    def test_negotiate_no_cost(self):
        assert (
            calc_action_note("NEGOTIATE", "Cost Needed", "GOOD", False)
            == "Contact supplier for cost"
        )

    def test_negotiate_with_cost(self):
        assert (
            calc_action_note("NEGOTIATE", "Ready", "GOOD", True)
            == "Negotiate lower trade price"
        )

    def test_watch_review_needed(self):
        assert (
            calc_action_note("WATCH", "Review Needed", "GOOD", True)
            == "Review listing safety"
        )

    def test_watch_bad_stability(self):
        assert (
            calc_action_note("WATCH", "Ready", "BAD", True) == "Monitor for 7 days"
        )

    def test_kill_unsafe(self):
        assert calc_action_note("KILL", "Reject", "GOOD", True) == "Avoid listing"
        assert (
            calc_action_note("KILL", "Review Needed", "GOOD", True)
            == "Avoid listing"
        )

    def test_default_review_manually(self):
        assert calc_action_note("KILL", "Ready", "GOOD", True) == "Review manually"


class TestBuildDecisionReason:
    def test_includes_all_pipe_separated_parts(self):
        out = build_decision_reason(
            "BUY", "BALANCED", "Low", "Unlikely", "Stable", 500, "cost gap +GBP1.00"
        )
        assert out == (
            "BUY | BALANCED | GBP500.00/mo | low IP risk | unlikely PL risk | "
            "stable | cost gap +GBP1.00"
        )

    def test_unassigned_lane_default(self):
        out = build_decision_reason("WATCH", "", "Medium", "-", "unknown", 0, None)
        assert "UNASSIGNED" in out

    def test_extra_omitted_when_falsy(self):
        out = build_decision_reason("BUY", "BALANCED", "Low", "Low", "Stable", 100, "")
        # Empty extra should not produce a trailing pipe.
        assert not out.endswith(" | ")
        assert "GBP100.00/mo" in out


# ---------------------------------------------------------------------------
# DataFrame entry point: compute_decisions
# ---------------------------------------------------------------------------


def _make_phase5_row(**overrides) -> dict:
    """Build a complete Phase-5 final_results row with sensible defaults.

    Defaults represent a clean BUY-eligible product so that test variations
    can simulate single-axis perturbations.
    """
    base = {
        "ASIN": "B0SAMPLE",
        "Product Name": "Sample Product",
        "Brand": "Acme",
        "BB Seller": "Different Co",
        "Opportunity Lane": "BALANCED",
        "Monthly Gross Profit": "GBP600.00",
        "Bought per Month": "100",
        "Est Profit": "GBP10.00",
        "Est ROI %": "45%",
        "Real ROI %": "45%",
        "Commercial Priority": "1",
        "Max Cost 20% ROI": "GBP10.00",
        "Trade Price": "GBP6.00",
        "Trade Price Found": "Y",
        "IP Risk Band": "Low",
        "Private Label Risk": "Unlikely",
        "Price Stability": "Stable",
        "Route Code": "DIRECT",
        "Gated": "N",
        "Fortress Listing": "NO",
        "Brand Seller Match": "NO",
        "Brand Type": "GENERIC",
        "Brand Store Present": "UNLIKELY",
    }
    base.update(overrides)
    return base


class TestComputeDecisions:
    def test_empty_df_returns_empty_with_decision_columns(self):
        df = pd.DataFrame(columns=["ASIN"])
        out = compute_decisions(df)
        for header in DECISION_HEADERS:
            assert header in out.columns
        assert len(out) == 0

    def test_input_df_is_not_mutated(self):
        df = pd.DataFrame([_make_phase5_row()])
        before = df.copy()
        _ = compute_decisions(df)
        pd.testing.assert_frame_equal(df, before)

    def test_appends_all_decision_headers(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = compute_decisions(df)
        for header in DECISION_HEADERS:
            assert header in out.columns

    def test_clean_row_is_buy(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "BUY"
        assert out.iloc[0]["Shortlist Flag"] == "Y"

    def test_sort_descending_by_decision_score(self):
        rows = [
            _make_phase5_row(ASIN="B001"),  # Strong BUY -> high score
            _make_phase5_row(
                ASIN="B002",
                **{
                    "IP Risk Band": "High",  # forces KILL
                    "Monthly Gross Profit": "GBP10.00",
                },
            ),
        ]
        df = pd.DataFrame(rows)
        out = compute_decisions(df)
        # KILL row (capped at 35) should land below the BUY row.
        assert (
            int(out.iloc[0]["Decision Score"]) >= int(out.iloc[1]["Decision Score"])
        )
        assert out.iloc[0]["ASIN"] == "B001"

    def test_sort_tiebreaker_by_monthly_gross_then_asin(self):
        # Identical decision profile, only ASIN/MGP differ.
        rows = [
            _make_phase5_row(ASIN="B002", **{"Monthly Gross Profit": "GBP100.00"}),
            _make_phase5_row(ASIN="B001", **{"Monthly Gross Profit": "GBP100.00"}),
        ]
        df = pd.DataFrame(rows)
        out = compute_decisions(df)
        # Tied scores + tied MGP -> ASIN ascending.
        assert list(out["ASIN"]) == ["B001", "B002"]


# ---------------------------------------------------------------------------
# Decision verdict branches
# ---------------------------------------------------------------------------


class TestDecisionVerdicts:
    def test_high_ip_risk_forces_kill(self):
        df = pd.DataFrame([_make_phase5_row(**{"IP Risk Band": "High"})])
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "KILL"
        assert int(out.iloc[0]["Decision Score"]) <= 35
        assert out.iloc[0]["Shortlist Flag"] == "N"

    def test_unsafe_joinability_forces_kill(self):
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "Fortress Listing": "YES",
                        "Brand Seller Match": "YES",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "KILL"

    def test_margin_fail_forces_kill(self):
        # Trade price above max, supplier confirmed.
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "Trade Price": "GBP15.00",  # above max GBP10
                        "Max Cost 20% ROI": "GBP10.00",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "KILL"

    def test_impossible_cost_gap_forces_kill(self):
        # cost gap = max - trade = 10 - 14.5 = -4.5 -> below impossibleGap (-4).
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "Trade Price": "GBP14.50",
                        "Max Cost 20% ROI": "GBP10.00",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "KILL"

    def test_buy_score_clamped_to_floor_eighty(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = compute_decisions(df)
        if out.iloc[0]["Decision"] == "BUY":
            assert int(out.iloc[0]["Decision Score"]) >= 80

    def test_negotiate_when_no_supplier_cost(self):
        # Strong product, no supplier cost -> NEGOTIATE.
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{"Trade Price Found": "N", "Trade Price": ""}
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "NEGOTIATE"
        assert int(out.iloc[0]["Decision Score"]) >= 60
        assert int(out.iloc[0]["Decision Score"]) <= 79

    def test_negotiate_when_margin_tight(self):
        # supplier between target and max -> Tight; no overrides hit -> NEGOTIATE.
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "Trade Price": "GBP9.00",  # between target 8 and max 10.
                        "Max Cost 20% ROI": "GBP10.00",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "NEGOTIATE"

    def test_watch_when_medium_ip_no_cost(self):
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "IP Risk Band": "Medium",
                        "Trade Price Found": "N",
                        "Trade Price": "",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "WATCH"
        # WATCH band: [40, 59].
        assert 40 <= int(out.iloc[0]["Decision Score"]) <= 59

    def test_watch_when_price_stability_bad(self):
        # Stability=BAD only forces WATCH if the row doesn't otherwise qualify
        # for BUY (BUY is checked first in the elif chain). Drop supplier cost
        # to disqualify BUY, then WATCH should fire on stability=BAD.
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "Trade Price Found": "N",
                        "Trade Price": "",
                        "Price Stability": "DROPPING",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "WATCH"

    def test_buy_demoted_to_negotiate_when_lane_is_profit(self):
        # Lane = PROFIT means even a buy-eligible row becomes NEGOTIATE.
        df = pd.DataFrame([_make_phase5_row(**{"Opportunity Lane": "PROFIT"})])
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "NEGOTIATE"

    def test_buy_demoted_to_negotiate_when_buy_readiness_not_ready(self):
        # Gated=Y forces joinability=Review -> buy_readiness=Review Needed.
        # Strict BUY branch fails (joinability != Joinable), but score-fallthrough
        # picks BUY (score >= 80 + has_supplier_cost). Post-check 1 demotes
        # BUY -> NEGOTIATE deterministically when supplier cost present.
        df = pd.DataFrame([_make_phase5_row(**{"Gated": "Y"})])
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "NEGOTIATE"

    @pytest.mark.parametrize("lane", ["PROFIT", "UNASSIGNED", "BRAND APPROACH"])
    def test_buy_demoted_to_negotiate_for_non_buy_lanes(self, lane):
        # Post-check 2: BUY rows on lanes that aren't BALANCED or CASH FLOW
        # are deterministically demoted to NEGOTIATE.
        df = pd.DataFrame([_make_phase5_row(**{"Opportunity Lane": lane})])
        out = compute_decisions(df)
        # An UNASSIGNED row may also fall short of commercially_strong, in which
        # case the score-floor branch decides — assert the rule rather than the
        # specific verdict. The key invariant: it must NOT be BUY.
        assert out.iloc[0]["Decision"] != "BUY"

    def test_watch_when_pl_risk_dash_and_not_joinable(self):
        # Reviewer-flagged WATCH branch: pl_risk == "-" AND joinability != Joinable.
        # ip=Medium pushes joinability=Review; pl="-" matches the dash sentinel.
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "IP Risk Band": "Medium",
                        "Private Label Risk": "-",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "WATCH"

    def test_low_score_lands_in_watch(self):
        # Weak commercial signals + no decisive override -> falls through to
        # the score-band branches. With ip=LOW + Joinable, safety_score floors
        # around ~80, which keeps decision_score >= 40 -> WATCH (not KILL).
        # The score-fallthrough KILL branch is defensive — see decision_engine.py
        # comment at the `else: decision = "KILL"` line.
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{
                        "Opportunity Lane": "UNASSIGNED",
                        "Monthly Gross Profit": "GBP0",
                        "Bought per Month": "0",
                        "Est Profit": "GBP0",
                        "Est ROI %": "0%",
                        "Real ROI %": "0%",
                        "Commercial Priority": "8",
                        "Trade Price Found": "Y",
                        "Trade Price": "GBP9.50",
                        "Max Cost 20% ROI": "GBP10.00",
                        "Route Code": "UNCLEAR",
                    }
                )
            ]
        )
        out = compute_decisions(df)
        assert out.iloc[0]["Decision"] == "WATCH"


class TestDecisionReason:
    def test_decision_reason_includes_cost_gap_when_supplier_cost_known(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = compute_decisions(df)
        assert "cost gap" in out.iloc[0]["Decision Reason"]

    def test_decision_reason_includes_supplier_missing_when_no_cost(self):
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    **{"Trade Price Found": "N", "Trade Price": ""}
                )
            ]
        )
        out = compute_decisions(df)
        assert "supplier cost missing" in out.iloc[0]["Decision Reason"]


# ---------------------------------------------------------------------------
# Step contract
# ---------------------------------------------------------------------------


class TestRunStep:
    def test_run_step_returns_enriched_dataframe(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = run_step(df, {})  # no required keys for decision step
        for header in DECISION_HEADERS:
            assert header in out.columns

    def test_run_step_accepts_optional_niche_key(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = run_step(df, {"niche": "kids-toys"})
        assert "Decision" in out.columns


# ---------------------------------------------------------------------------
# Reviewer-flagged: NaN safety + missing columns
# ---------------------------------------------------------------------------


class TestNaNSafety:
    def test_nan_numeric_columns_coerce_to_zero(self):
        row = _make_phase5_row()
        # Replace key numeric fields with float NaN.
        row["Monthly Gross Profit"] = float("nan")
        row["Trade Price"] = float("nan")
        row["Max Cost 20% ROI"] = float("nan")
        df = pd.DataFrame([row])
        out = compute_decisions(df)
        # No crash; row produces some decision (likely KILL or WATCH given
        # zero economics + weak signals).
        assert out.iloc[0]["Decision"] in {"BUY", "NEGOTIATE", "WATCH", "KILL"}

    def test_nan_categorical_columns_treated_as_empty(self):
        row = _make_phase5_row()
        row["IP Risk Band"] = float("nan")
        row["Opportunity Lane"] = float("nan")
        df = pd.DataFrame([row])
        out = compute_decisions(df)
        reason = out.iloc[0]["Decision Reason"]
        # NaN lane -> "UNASSIGNED" in reason; NaN IP -> "unknown IP risk".
        assert "UNASSIGNED" in reason
        assert "unknown ip risk" in reason.lower()

    def test_run_step_handles_numeric_columns(self):
        # The CLI uses dtype=str so NaN never appears, but the step 5 runner
        # may pass a DataFrame with raw numeric columns. Verify the scorer
        # tolerates int/float values, not just strings.
        row = _make_phase5_row()
        row["Monthly Gross Profit"] = 600.0
        row["Bought per Month"] = 100
        row["Trade Price"] = 6.0
        row["Max Cost 20% ROI"] = 10.0
        df = pd.DataFrame([row])
        out = run_step(df, {})
        assert out.iloc[0]["Decision"] in {"BUY", "NEGOTIATE", "WATCH", "KILL"}

    def test_warns_when_required_columns_missing(self):
        df = pd.DataFrame([{"ASIN": "B0BARE"}])
        with pytest.warns(UserWarning, match="missing input columns"):
            _ = compute_decisions(df)


# ---------------------------------------------------------------------------
# Stats / handoff
# ---------------------------------------------------------------------------


class TestStats:
    def test_stats_includes_niche_and_decision_distribution(self):
        df = pd.DataFrame([_make_phase5_row(), _make_phase5_row(ASIN="B002")])
        out = compute_decisions(df)
        text = build_stats(out, niche="kids-toys")
        assert "Niche: kids-toys" in text
        assert "Decision distribution" in text
        for verdict in ["BUY", "NEGOTIATE", "WATCH", "KILL"]:
            assert verdict in text

    def test_stats_lists_top_15_by_decision_score(self):
        df = pd.DataFrame(
            [_make_phase5_row(ASIN=f"B{i:03d}") for i in range(20)]
        )
        out = compute_decisions(df)
        text = build_stats(out, niche="kids-toys")
        assert "Top 15 by Decision Score" in text


class TestHandoff:
    def test_handoff_includes_summary_counts(self):
        df = pd.DataFrame([_make_phase5_row()])
        out = compute_decisions(df)
        text = build_handoff(
            out,
            niche="kids-toys",
            decision_csv_path="kids_toys_phase6_decisions.csv",
            stats_path="kids_toys_phase6_stats.txt",
            handoff_path="kids_toys_phase6_handoff.md",
            shortlist_xlsx_path="kids_toys_phase6_shortlist.xlsx",
        )
        assert "Phase 6 Handoff" in text
        assert "BUY:" in text
        assert "kids-toys" in text


# ---------------------------------------------------------------------------
# XLSX shortlist workbook
# ---------------------------------------------------------------------------


class TestBuildShortlistXlsx:
    def test_writes_workbook_with_two_sheets(self, tmp_path: Path):
        df = pd.DataFrame([_make_phase5_row(), _make_phase5_row(ASIN="B002")])
        enriched = compute_decisions(df)
        out_path = tmp_path / "test_shortlist.xlsx"
        build_shortlist_xlsx(enriched, out_path)
        assert out_path.exists()
        wb = openpyxl.load_workbook(out_path)
        assert "Shortlist" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_shortlist_sheet_includes_only_buy_and_negotiate_rows(
        self, tmp_path: Path
    ):
        df = pd.DataFrame(
            [
                _make_phase5_row(ASIN="B0BUY"),  # BUY
                _make_phase5_row(
                    ASIN="B0KILL", **{"IP Risk Band": "High"}
                ),  # KILL
            ]
        )
        enriched = compute_decisions(df)
        out_path = tmp_path / "shortlist_filtered.xlsx"
        build_shortlist_xlsx(enriched, out_path)
        wb = openpyxl.load_workbook(out_path)
        sheet = wb["Shortlist"]
        # Header row + only the BUY row.
        asin_col = [
            cell.value for cell in sheet["A"][1:]  # skip header
        ]
        assert "B0BUY" in asin_col
        assert "B0KILL" not in asin_col

    def test_summary_sheet_lists_decision_counts(self, tmp_path: Path):
        df = pd.DataFrame([_make_phase5_row()])
        enriched = compute_decisions(df)
        out_path = tmp_path / "summary.xlsx"
        build_shortlist_xlsx(enriched, out_path)
        wb = openpyxl.load_workbook(out_path)
        sheet = wb["Summary"]
        # First block title is "Decision".
        assert sheet.cell(1, 1).value == "Decision"


# ---------------------------------------------------------------------------
# Edge cases / Unicode
# ---------------------------------------------------------------------------


class TestEdgeCaseInputs:
    def test_unicode_brand_does_not_crash(self):
        df = pd.DataFrame(
            [
                _make_phase5_row(
                    Brand="Café Ø", **{"Product Name": "日本商品"}
                )
            ]
        )
        out = compute_decisions(df)
        assert len(out) == 1

    def test_decision_score_is_integer_string(self):
        # Output schema uses the stringified rounded score (no decimals).
        df = pd.DataFrame([_make_phase5_row()])
        out = compute_decisions(df)
        score_str = out.iloc[0]["Decision Score"]
        # Must be a digit-only string (or with leading minus, though scores
        # are non-negative). No decimal point allowed.
        assert "." not in score_str
        assert score_str.lstrip("-").isdigit()
